"""
Example usage of the ExcelManager class.

This example demonstrates how to use the optimized ExcelManager
for better performance and memory management.
"""

import sys
import os

# Add src to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from utils.excel_manager import ExcelManager, ExcelContextManager
from openpyxl.styles import PatternFill
import logging

# Set up logging
logging.basicConfig(level=logging.INFO)


def example_basic_usage():
    """Basic usage example of ExcelManager."""
    print("=== Basic ExcelManager Usage ===")
    
    # Create a test Excel file
    test_file = "test_excel_manager.xlsx"
    
    try:
        # Create new Excel file
        excel = ExcelManager.create_excel_file(test_file, "TestSheet")
        
        # Get file info
        info = excel.get_file_info()
        print(f"File created: {info}")
        
        # Add some data
        sheet = excel.active_sheet
        sheet['A1'] = "Name"
        sheet['B1'] = "Value"
        sheet['A2'] = "Test Data"
        sheet['B2'] = 42
        
        # Save changes
        excel.save()
        print("Data added and saved successfully")
        
        # Close the workbook
        excel.close()
        
    except Exception as e:
        print(f"Error in basic usage: {e}")
    finally:
        # Clean up
        if os.path.exists(test_file):
            os.remove(test_file)


def example_context_manager():
    """Example using context manager for automatic cleanup."""
    print("\n=== Context Manager Usage ===")
    
    test_file = "test_context.xlsx"
    
    try:
        # Create test file
        ExcelManager.create_excel_file(test_file, "DataSheet")
        
        # Use context manager
        with ExcelContextManager(test_file) as excel:
            print(f"File info: {excel.get_file_info()}")
            
            # Work with sheets
            sheet = excel.get_sheet_by_name("DataSheet")
            if sheet:
                sheet['A1'] = "Context Manager Test"
                sheet['A2'] = "Auto-cleanup enabled"
                
                # Add some formatting
                fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                sheet['A1'].fill = fill
                
                excel.save()
                print("File processed with context manager")
        
        # File is automatically closed here
        print("Context manager automatically closed the file")
        
    except Exception as e:
        print(f"Error in context manager usage: {e}")
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)


def example_performance_comparison():
    """Compare performance with traditional vs optimized approach."""
    print("\n=== Performance Comparison ===")
    
    test_file = "test_performance.xlsx"
    
    try:
        # Create a larger test file
        excel = ExcelManager.create_excel_file(test_file, "LargeData")
        
        # Add 1000 rows of data
        sheet = excel.active_sheet
        for i in range(1, 1001):
            sheet[f'A{i}'] = f"Row {i}"
            sheet[f'B{i}'] = i * 2
            sheet[f'C{i}'] = f"Data_{i}"
        
        excel.save()
        excel.close()
        
        # Test optimized loading
        print("Testing optimized loading...")
        
        import time
        
        # Method 1: Traditional loading
        start_time = time.time()
        from openpyxl import load_workbook
        wb1 = load_workbook(test_file)
        sheet1 = wb1.active
        data1 = [[cell.value for cell in row] for row in sheet1.iter_rows()]
        wb1.close()
        traditional_time = time.time() - start_time
        
        # Method 2: Optimized loading with ExcelManager
        start_time = time.time()
        with ExcelContextManager(test_file) as excel:
            sheet = excel.active_sheet
            data2 = [[cell.value for cell in row] for row in sheet.iter_rows()]
        optimized_time = time.time() - start_time
        
        print(f"Traditional loading: {traditional_time:.3f}s")
        print(f"Optimized loading: {optimized_time:.3f}s")
        print(f"Improvement: {((traditional_time - optimized_time) / traditional_time * 100):.1f}%")
        
        # Verify data integrity
        assert data1 == data2, "Data mismatch between methods"
        print("✓ Data integrity verified")
        
    except Exception as e:
        print(f"Error in performance comparison: {e}")
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)


def example_sheet_operations():
    """Example of sheet operations."""
    print("\n=== Sheet Operations ===")
    
    test_file = "test_sheets.xlsx"
    
    try:
        # Create file with multiple sheets
        excel = ExcelManager.create_excel_file(test_file, "Main")
        
        # Add additional sheets
        excel.create_sheet("Summary")
        excel.create_sheet("Data")
        
        print(f"Available sheets: {excel.get_sheet_names()}")
        
        # Check if sheet exists
        print(f"'Main' exists: {excel.sheet_exists('Main')}")
        print(f"'NonExistent' exists: {excel.sheet_exists('NonExistent')}")
        
        # Get specific sheet
        summary_sheet = excel.get_sheet_by_name("Summary")
        if summary_sheet:
            summary_sheet['A1'] = "This is the summary sheet"
        
        excel.save()
        excel.close()
        print("Sheet operations completed")
        
    except Exception as e:
        print(f"Error in sheet operations: {e}")
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)


if __name__ == "__main__":
    print("ExcelManager Usage Examples")
    print("=" * 50)
    
    try:
        example_basic_usage()
        example_context_manager()
        example_sheet_operations()
        example_performance_comparison()
        
        print("\n" + "=" * 50)
        print("All examples completed successfully!")
        
    except Exception as e:
        print(f"Example execution failed: {e}")
        import traceback
        traceback.print_exc()
