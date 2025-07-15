"""
Caching system usage examples for ProjectBudgetinator.

This module demonstrates how to use the caching system to improve
performance by reducing repeated Excel file reads and operations.
"""

import sys
import os
import time
import tempfile
import shutil

# Add src to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from utils.cache_manager import CacheManager, cache_manager, cached_with_invalidation
from utils.excel_manager import ExcelManager
from core.excel_service import ExcelService
import logging

# Set up logging
logging.basicConfig(level=logging.INFO)


def example_basic_caching():
    """Basic caching usage examples."""
    print("=== Basic Caching Examples ===")
    
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Create test Excel file
        test_file = os.path.join(temp_dir, "test_cache.xlsx")
        ExcelManager.create_excel_file(test_file, "TestSheet")
        
        # Add some test data
        with ExcelManager(test_file) as excel:
            sheet = excel.active_sheet
            sheet['A1'] = "Name"
            sheet['B1'] = "Value"
            sheet['A2'] = "Test Data"
            sheet['B2'] = 42
            excel.save()
        
        # Test file validation caching
        print("Testing file validation caching...")
        
        # First call - cache miss
        start_time = time.time()
        is_valid, error_msg = cache_manager.validate_excel_file(test_file)
        first_time = time.time() - start_time
        print(f"First validation: {is_valid}, took {first_time:.4f}s")
        
        # Second call - cache hit
        start_time = time.time()
        is_valid, error_msg = cache_manager.validate_excel_file(test_file)
        second_time = time.time() - start_time
        print(f"Second validation: {is_valid}, took {second_time:.4f}s")
        print(f"Speed improvement: {((first_time - second_time) / first_time * 100):.1f}%")
        
        # Test sheet names caching
        print("\nTesting sheet names caching...")
        
        # First call
        start_time = time.time()
        sheet_names = cache_manager.get_excel_sheet_names(test_file)
        first_time = time.time() - start_time
        print(f"First call: {sheet_names}, took {first_time:.4f}s")
        
        # Second call
        start_time = time.time()
        sheet_names = cache_manager.get_excel_sheet_names(test_file)
        second_time = time.time() - start_time
        print(f"Second call: {sheet_names}, took {second_time:.4f}s")
        
        # Test file metadata caching
        print("\nTesting file metadata caching...")
        metadata = cache_manager.get_file_metadata(test_file)
        print(f"File metadata: {metadata}")
        
    finally:
        shutil.rmtree(temp_dir)


def example_partner_caching():
    """Partner data caching examples."""
    print("\n=== Partner Data Caching ===")
    
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Create test Excel file with partner sheets
        test_file = os.path.join(temp_dir, "partners_cache.xlsx")
        excel = ExcelManager.create_excel_file(test_file, "PM Summary")
        
        # Create partner sheets
        p2_sheet = excel.create_sheet("P2 ABC")
        p2_sheet['D4'] = "P2-12345"
        p2_sheet['D5'] = "ABC Corporation"
        p2_sheet['D6'] = "Germany"
        p2_sheet['D7'] = "Coordinator"
        
        p3_sheet = excel.create_sheet("P3 XYZ")
        p3_sheet['D4'] = "P3-67890"
        p3_sheet['D5'] = "XYZ Ltd"
        p3_sheet['D6'] = "France"
        p3_sheet['D7'] = "Partner"
        
        excel.save()
        excel.close()
        
        # Test partner data caching
        print("Testing partner data caching...")
        
        # First call for P2
        start_time = time.time()
        p2_data = cache_manager.get_partner_data(test_file, "2")
        first_time = time.time() - start_time
        print(f"P2 data (first): {p2_data}, took {first_time:.4f}s")
        
        # Second call for P2 (cache hit)
        start_time = time.time()
        p2_data = cache_manager.get_partner_data(test_file, "2")
        second_time = time.time() - start_time
        print(f"P2 data (second): {p2_data}, took {second_time:.4f}s")
        
        # First call for P3
        start_time = time.time()
        p3_data = cache_manager.get_partner_data(test_file, "3")
        first_time = time.time() - start_time
        print(f"P3 data (first): {p3_data}, took {first_time:.4f}s")
        
    finally:
        shutil.rmtree(temp_dir)


def example_cache_decorator():
    """Example of using cache decorator."""
    print("\n=== Cache Decorator Usage ===")
    
    # Create a custom cache manager
    custom_cache = CacheManager()
    
    @cached_with_invalidation(custom_cache, ttl=60)  # 1 minute TTL
    def expensive_calculation(file_path: str, multiplier: int) -> int:
        """Simulate expensive calculation."""
        print(f"Performing expensive calculation for {file_path} * {multiplier}")
        time.sleep(0.1)  # Simulate expensive operation
        return os.path.getsize(file_path) * multiplier
    
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Create test file
        test_file = os.path.join(temp_dir, "decorator_test.xlsx")
        ExcelManager.create_excel_file(test_file, "Test")
        
        # First call - will execute
        result1 = expensive_calculation(test_file, 2)
        print(f"First call result: {result1}")
        
        # Second call - will use cache
        result2 = expensive_calculation(test_file, 2)
        print(f"Second call result: {result2}")
        
        # Different parameters - will execute
        result3 = expensive_calculation(test_file, 3)
        print(f"Different param result: {result3}")
        
    finally:
        shutil.rmtree(temp_dir)


def example_excel_service_caching():
    """ExcelService with caching examples."""
    print("\n=== ExcelService with Caching ===")
    
    service = ExcelService()
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Create test file
        test_file = os.path.join(temp_dir, "service_cache.xlsx")
        excel = ExcelManager.create_excel_file(test_file, "PM Summary")
        
        # Create partner sheets
        for i in range(2, 5):
            sheet = excel.create_sheet(f"P{i} Partner{i}")
            sheet['D4'] = f"P{i}-ID"
            sheet['D5'] = f"Partner {i}"
        
        excel.save()
        excel.close()
        
        # Test cached operations
        print("Testing ExcelService caching...")
        
        # Get file info (cached)
        start_time = time.time()
        file_info = service.get_file_info(test_file)
        first_time = time.time() - start_time
        print(f"File info (first): {file_info['sheet_count']} sheets, "
              f"took {first_time:.4f}s")
        
        # Get partner list (cached)
        start_time = time.time()
        partners = service.get_partner_list(test_file)
        second_time = time.time() - start_time
        print(f"Partner list: {len(partners)} partners, took {second_time:.4f}s")
        
        # Cache statistics
        stats = service.get_cache_stats()
        print(f"Cache stats: {stats}")
        
    finally:
        service.close_all_workbooks()
        shutil.rmtree(temp_dir)


def example_cache_invalidation():
    """Cache invalidation examples."""
    print("\n=== Cache Invalidation ===")
    
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Create test file
        test_file = os.path.join(temp_dir, "invalidation_test.xlsx")
        excel = ExcelManager.create_excel_file(test_file, "Sheet1")
        excel.save()
        excel.close()
        
        # Get initial cache stats
        initial_stats = cache_manager.get_cache_stats()
        print(f"Initial cache stats: {initial_stats}")
        
        # Populate cache
        sheet_names = cache_manager.get_excel_sheet_names(test_file)
        print(f"Cached sheet names: {sheet_names}")
        
        # Invalidate cache for this file
        cache_manager.invalidate_cache_for_file(test_file)
        print("Cache invalidated for file")
        
        # Check cache stats after invalidation
        new_stats = cache_manager.get_cache_stats()
        print(f"Cache stats after invalidation: {new_stats}")
        
        # Clear all caches
        cleared = cache_manager.clear_cache()
        print(f"Cleared {cleared} cache files")
        
    finally:
        shutil.rmtree(temp_dir)


def example_performance_comparison():
    """Compare performance with and without caching."""
    print("\n=== Performance Comparison ===")
    
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Create test file with multiple sheets
        test_file = os.path.join(temp_dir, "perf_test.xlsx")
        excel = ExcelManager.create_excel_file(test_file, "PM Summary")
        
        # Create 10 partner sheets
        for i in range(2, 12):
            sheet = excel.create_sheet(f"P{i} Partner{i}")
            sheet['D4'] = f"P{i}-ID"
            sheet['D5'] = f"Partner {i} Name"
            sheet['D6'] = f"Country{i}"
            sheet['D7'] = f"Role{i}"
        
        excel.save()
        excel.close()
        
        # Test without caching (simulate)
        print("Testing without caching...")
        start_time = time.time()
        
        # Simulate multiple operations
        for i in range(5):
            from openpyxl import load_workbook
            with load_workbook(test_file, read_only=True) as wb:
                sheets = wb.sheetnames
                partner_sheets = [s for s in sheets if s.startswith('P')]
        
        no_cache_time = time.time() - start_time
        print(f"Without caching: {no_cache_time:.4f}s")
        
        # Test with caching
        print("Testing with caching...")
        start_time = time.time()
        
        for i in range(5):
            sheet_names = cache_manager.get_excel_sheet_names(test_file)
            partner_sheets = [s for s in sheet_names if s.startswith('P')]
        
        with_cache_time = time.time() - start_time
        print(f"With caching: {with_cache_time:.4f}s")
        
        improvement = ((no_cache_time - with_cache_time) / no_cache_time * 100)
        print(f"Performance improvement: {improvement:.1f}%")
        
    finally:
        shutil.rmtree(temp_dir)


if __name__ == "__main__":
    print("Caching System Usage Examples")
    print("=" * 50)
    
    try:
        example_basic_caching()
        example_partner_caching()
        example_cache_decorator()
        example_excel_service_caching()
        example_cache_invalidation()
        example_performance_comparison()
        
        print("\n" + "=" * 50)
        print("All caching examples completed successfully!")
        
        # Final cache stats
        stats = cache_manager.get_cache_stats()
        print(f"Final cache statistics: {stats}")
        
    except Exception as e:
        print(f"Example execution failed: {e}")
        import traceback
        traceback.print_exc()
