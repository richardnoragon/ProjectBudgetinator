"""
GUI Memory Management Examples.

This module demonstrates how to use the GUI memory management system
to prevent memory leaks and ensure proper window lifecycle management.
"""

import tkinter as tk
from tkinter import ttk
import sys
import os
import tempfile
import shutil

# Add src to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from gui.window_manager import (
    BaseDialog, window_manager, weak_callback_manager, force_gui_cleanup
)
from handlers.base_handler import (
    BaseHandler, ExcelOperationHandler, ValidationResult, OperationResult
)


class ExampleDialog(BaseDialog):
    """Example dialog demonstrating GUI memory management."""
    
    def __init__(self, parent, title="Example Dialog"):
        super().__init__(parent, title)
        
        self._result = None
        self.setup_ui()
    
    def setup_ui(self):
        """Setup the dialog UI."""
        # Create main frame
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Add content
        ttk.Label(main_frame, text="This is a memory-managed dialog").pack(pady=10)
        
        # Add input field
        self.input_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.input_var).pack(pady=5)
        
        # Add buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=10)
        
        ttk.Button(button_frame, text="OK", 
                  command=self.on_ok).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", 
                  command=self.on_cancel).pack(side=tk.LEFT, padx=5)
    
    def on_ok(self):
        """Handle OK button click."""
        self._result = self.input_var.get()
        self.destroy()
    
    def on_cancel(self):
        """Handle Cancel button click."""
        self._result = None
        self.destroy()
    
    def _on_cleanup(self):
        """Custom cleanup logic."""
        print(f"Cleaning up dialog: {self.window_id}")


class ExampleHandler(BaseHandler):
    """Example handler demonstrating base handler usage."""
    
    def __init__(self, parent_window):
        super().__init__(parent_window)
    
    def validate_input(self, data: Dict[str, Any]) -> ValidationResult:
        """Validate input data."""
        result = ValidationResult()
        
        if not data.get('name'):
            result.add_error("Name is required")
        
        if not isinstance(data.get('age'), int) or data.get('age', 0) < 0:
            result.add_error("Age must be a positive integer")
        
        if data.get('age', 0) > 100:
            result.add_warning("Age seems unusually high")
        
        return result
    
    def process(self, data: Dict[str, Any]) -> OperationResult:
        """Process the operation."""
        try:
            # Simulate processing
            processed_data = {
                'name': data['name'].upper(),
                'age': data['age'],
                'processed_at': str(self.timestamp)
            }
            
            return OperationResult(
                success=True,
                message=f"Processed {data['name']} successfully",
                data=processed_data
            )
        except Exception as e:
            return OperationResult(
                success=False,
                message=f"Processing failed: {str(e)}",
                errors=[str(e)]
            )


class PartnerHandler(ExcelOperationHandler):
    """Example partner handler using ExcelOperationHandler."""
    
    def __init__(self, parent_window, workbook_path):
        super().__init__(parent_window, workbook_path)
    
    def validate_input(self, data: Dict[str, Any]) -> ValidationResult:
        """Validate partner data."""
        result = super().validate_required_fields(data, [
            'partner_number', 'partner_name', 'budget'
        ])
        
        numeric_result = super().validate_numeric_fields(data, ['budget'])
        
        return result.merge(numeric_result)
    
    def process(self, data: Dict[str, Any]) -> OperationResult:
        """Process partner addition."""
        try:
            with self.with_workbook_context(self._add_partner_to_sheet, data) as result:
                return result
        except Exception as e:
            return OperationResult(
                success=False,
                message=f"Failed to add partner: {str(e)}",
                errors=[str(e)]
            )
    
    def _add_partner_to_sheet(self, workbook, data):
        """Add partner to Excel sheet."""
        # This is a simplified example
        sheet = workbook.active
        next_row = sheet.max_row + 1
        
        sheet.cell(row=next_row, column=1, value=data['partner_number'])
        sheet.cell(row=next_row, column=2, value=data['partner_name'])
        sheet.cell(row=next_row, column=3, value=data['budget'])
        
        return OperationResult(
            success=True,
            message=f"Added partner {data['partner_name']}",
            data={'row': next_row}
        )


class MainWindow:
    """Main window demonstrating GUI memory management."""
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("GUI Memory Management Demo")
        self.root.geometry("400x500")
        
        self.setup_ui()
        self.setup_callbacks()
        
    def setup_ui(self):
        """Setup the main UI."""
        # Create notebook for tabs
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Memory management tab
        memory_frame = ttk.Frame(notebook)
        notebook.add(memory_frame, text="Memory Management")
        
        ttk.Label(memory_frame, text="GUI Memory Management Demo",
                 font=("Arial", 14, "bold")).pack(pady=10)
        
        # Buttons for testing
        button_frame = ttk.Frame(memory_frame)
        button_frame.pack(pady=10)
        
        ttk.Button(button_frame, text="Open Dialog",
                  command=self.open_dialog).pack(pady=5)
        
        ttk.Button(button_frame, text="Open Multiple Dialogs",
                  command=self.open_multiple_dialogs).pack(pady=5)
        
        ttk.Button(button_frame, text="Show Window Stats",
                  command=self.show_window_stats).pack(pady=5)
        
        ttk.Button(button_frame, text="Force Cleanup",
                  command=self.force_cleanup).pack(pady=5)
        
        # Handler tab
        handler_frame = ttk.Frame(notebook)
        notebook.add(handler_frame, text="Base Handler")
        
        ttk.Label(handler_frame, text="Base Handler Demo",
                 font=("Arial", 14, "bold")).pack(pady=10)
        
        # Input fields
        input_frame = ttk.LabelFrame(handler_frame, text="Input Data")
        input_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(input_frame, text="Name:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.name_var = tk.StringVar()
        ttk.Entry(input_frame, textvariable=self.name_var).grid(row=0, column=1, padx=5, pady=2)
        
        ttk.Label(input_frame, text="Age:").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        self.age_var = tk.StringVar()
        ttk.Entry(input_frame, textvariable=self.age_var).grid(row=1, column=1, padx=5, pady=2)
        
        # Buttons
        handler_buttons = ttk.Frame(handler_frame)
        handler_buttons.pack(pady=10)
        
        ttk.Button(handler_buttons, text="Process with BaseHandler",
                  command=self.process_with_handler).pack(pady=5)
        
        ttk.Button(handler_buttons, text="Process with ExcelHandler",
                  command=self.process_with_excel_handler).pack(pady=5)
        
        # Status display
        self.status_text = tk.Text(handler_frame, height=10, width=50)
        self.status_text.pack(pady=10, padx=10)
        
    def setup_callbacks(self):
        """Setup weak reference callbacks."""
        from gui.window_manager import weak_callback_manager
        
        # Register weak callback
        weak_callback_manager.register_callback(
            "main_window_close", self.on_main_close
        )
    
    def on_main_close(self):
        """Callback for main window close."""
        print("Main window is closing - cleanup initiated")
    
    def open_dialog(self):
        """Open a single dialog."""
        dialog = ExampleDialog(self.root, "Single Dialog")
        result = dialog.show_modal()
        print(f"Dialog result: {result}")
    
    def open_multiple_dialogs(self):
        """Open multiple dialogs to test memory management."""
        for i in range(3):
            dialog = ExampleDialog(self.root, f"Dialog {i+1}")
            # Note: In real usage, you'd handle these differently
            # This is just for demonstration
            self.root.after(1000 * (i + 1), lambda d=dialog: d.destroy())
    
    def show_window_stats(self):
        """Show current window statistics."""
        stats = window_manager.get_memory_stats()
        message = f"""
Window Statistics:
- Total windows: {stats['total_windows']}
- Registered windows: {stats['registered_windows']}
- Cleanup callbacks: {stats['cleanup_callbacks']}
        """
        self.show_status(message)
    
    def force_cleanup(self):
        """Force cleanup of all GUI resources."""
        force_gui_cleanup()
        self.show_status("GUI cleanup completed")
    
    def process_with_handler(self):
        """Process data using BaseHandler."""
        handler = ExampleHandler(self.root)
        
        data = {
            'name': self.name_var.get(),
            'age': int(self.age_var.get()) if self.age_var.get().isdigit() else None
        }
        
        result = handler.execute(data)
        self.show_status(f"Result: {result.message}")
        
        if result.errors:
            self.show_status(f"Errors: {result.errors}")
    
    def process_with_excel_handler(self):
        """Process data using ExcelOperationHandler."""
        temp_dir = tempfile.mkdtemp()
        try:
            # Create temporary Excel file
            from openpyxl import Workbook
            test_file = os.path.join(temp_dir, 'test.xlsx')
            
            wb = Workbook()
            wb.active.title = "Partners"
            wb.save(test_file)
            wb.close()
            
            handler = PartnerHandler(self.root, test_file)
            
            data = {
                'partner_number': 'P001',
                'partner_name': self.name_var.get() or 'Test Partner',
                'budget': float(self.age_var.get()) if self.age_var.get().isdigit() else 1000.0
            }
            
            result = handler.execute(data)
            self.show_status(f"Excel Result: {result.message}")
            
        except Exception as e:
            self.show_status(f"Error: {str(e)}")
        finally:
            shutil.rmtree(temp_dir)
    
    def show_status(self, message):
        """Show status message."""
        self.status_text.delete(1.0, tk.END)
        self.status_text.insert(1.0, message)
    
    def run(self):
        """Run the application."""
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.root.mainloop()
    
    def on_closing(self):
        """Handle application closing."""
        force_gui_cleanup()
        self.root.destroy()


def main():
    """Main function."""
    print("Starting GUI Memory Management Demo...")
    
    app = MainWindow()
    try:
        app.run()
    except KeyboardInterrupt:
        print("Application interrupted")
    finally:
        print("Application closed")


if __name__ == "__main__":
    main()
