"""
Comprehensive test suite for ProjectBudgetinator handlers.

This module contains unit tests for all handler classes, following
the testing recommendations from OPTIMIZATION_RECOMMENDATIONS.md.
"""

import pytest
import tkinter as tk
from unittest.mock import Mock, MagicMock, patch, mock_open
from pathlib import Path
import tempfile
import os
import openpyxl
from typing import Dict, Any

# Import the modules we're testing
from handlers.base_handler import (
    BaseHandler, ExcelOperationHandler, DialogHandler, 
    BatchOperationHandler, ValidationResult, OperationResult
)


class MockHandler(BaseHandler):
    """Mock implementation of BaseHandler for testing."""
    
    def __init__(self, parent_window=None, workbook_path=None):
        super().__init__(parent_window, workbook_path)
        self._validation_result = ValidationResult()
        self._operation_result = OperationResult()
    
    def validate_input(self, data: Dict[str, Any]) -> ValidationResult:
        return self._validation_result
    
    def process(self, data: Dict[str, Any]) -> OperationResult:
        return self._operation_result
    
    def set_validation_result(self, result: ValidationResult):
        self._validation_result = result
    
    def set_operation_result(self, result: OperationResult):
        self._operation_result = result


class TestValidationResult:
    """Test ValidationResult class."""
    
    def test_init_default(self):
        """Test default initialization."""
        result = ValidationResult()
        assert result.valid is True
        assert result.errors == []
        assert result.warnings == []
    
    def test_init_with_values(self):
        """Test initialization with values."""
        errors = ["Error 1", "Error 2"]
        warnings = ["Warning 1"]
        result = ValidationResult(valid=False, errors=errors, warnings=warnings)
        
        assert result.valid is False
        assert result.errors == errors
        assert result.warnings == warnings
    
    def test_add_error(self):
        """Test adding an error."""
        result = ValidationResult()
        result.add_error("Test error")
        
        assert result.valid is False
        assert "Test error" in result.errors
    
    def test_add_warning(self):
        """Test adding a warning."""
        result = ValidationResult()
        result.add_warning("Test warning")
        
        assert result.valid is True  # Warnings don't affect validity
        assert "Test warning" in result.warnings
    
    def test_merge_valid_results(self):
        """Test merging two valid results."""
        result1 = ValidationResult(warnings=["Warning 1"])
        result2 = ValidationResult(warnings=["Warning 2"])
        
        merged = result1.merge(result2)
        
        assert merged.valid is True
        assert merged.warnings == ["Warning 1", "Warning 2"]
    
    def test_merge_invalid_results(self):
        """Test merging with invalid result."""
        result1 = ValidationResult(errors=["Error 1"])
        result2 = ValidationResult(errors=["Error 2"])
        
        merged = result1.merge(result2)
        
        assert merged.valid is False
        assert merged.errors == ["Error 1", "Error 2"]


class TestOperationResult:
    """Test OperationResult class."""
    
    def test_init_default(self):
        """Test default initialization."""
        result = OperationResult()
        assert result.success is True
        assert result.message == ""
        assert result.data == {}
        assert result.errors == []
        assert result.timestamp is not None
    
    def test_init_with_values(self):
        """Test initialization with values."""
        data = {"key": "value"}
        errors = ["Error 1"]
        result = OperationResult(
            success=False, 
            message="Test message",
            data=data,
            errors=errors
        )
        
        assert result.success is False
        assert result.message == "Test message"
        assert result.data == data
        assert result.errors == errors


class TestBaseHandler:
    """Test BaseHandler class."""
    
    @pytest.fixture
    def mock_parent(self):
        """Create mock parent window."""
        return Mock(spec=tk.Widget)
    
    @pytest.fixture
    def handler(self, mock_parent):
        """Create mock handler instance."""
        return MockHandler(mock_parent, "/test/path.xlsx")
    
    def test_init(self, mock_parent):
        """Test handler initialization."""
        handler = MockHandler(mock_parent, "/test/path.xlsx")
        
        assert handler.parent == mock_parent
        assert handler.workbook_path == "/test/path.xlsx"
        assert handler.logger is not None
        assert handler._operation_history == []
    
    def test_execute_successful_operation(self, handler):
        """Test successful operation execution."""
        # Setup successful validation and operation
        validation_result = ValidationResult(valid=True)
        operation_result = OperationResult(success=True, message="Success")
        
        handler.set_validation_result(validation_result)
        handler.set_operation_result(operation_result)
        
        # Execute operation
        data = {"test": "data"}
        result = handler.execute(data)
        
        assert result.success is True
        assert result.message == "Success"
        assert len(handler.get_operation_history()) == 1
    
    def test_execute_validation_failure(self, handler):
        """Test operation with validation failure."""
        # Setup failed validation
        validation_result = ValidationResult(valid=False, errors=["Invalid data"])
        handler.set_validation_result(validation_result)
        
        # Execute operation
        data = {"test": "data"}
        result = handler.execute(data)
        
        assert result.success is False
        assert "Validation failed" in result.message
        assert result.errors == ["Invalid data"]
    
    def test_execute_with_warnings(self, handler):
        """Test operation with validation warnings."""
        # Setup validation with warnings
        validation_result = ValidationResult(
            valid=True, 
            warnings=["Warning message"]
        )
        operation_result = OperationResult(success=True)
        
        handler.set_validation_result(validation_result)
        handler.set_operation_result(operation_result)
        
        # Execute operation
        with patch.object(handler.logger, 'warning') as mock_warning:
            result = handler.execute({"test": "data"})
            
            assert result.success is True
            mock_warning.assert_called_with("Warning message")
    
    def test_execute_exception_handling(self, handler):
        """Test exception handling during execution."""
        # Setup validation to pass but operation to fail
        validation_result = ValidationResult(valid=True)
        handler.set_validation_result(validation_result)
        
        # Make process method raise exception
        def failing_process(data):
            raise ValueError("Test exception")
        
        handler.process = failing_process
        
        # Execute operation
        result = handler.execute({"test": "data"})
        
        assert result.success is False
        assert "Unexpected error" in result.message
        assert "Test exception" in result.errors[0]
    
    @patch('tkinter.messagebox.showerror')
    def test_show_error(self, mock_messagebox, handler):
        """Test error message display."""
        handler.show_error("Test error", "Test Title")
        
        mock_messagebox.assert_called_once_with("Test Title", "Test error")
    
    @patch('tkinter.messagebox.showinfo')
    def test_show_success(self, mock_messagebox, handler):
        """Test success message display."""
        handler.show_success("Test success", "Test Title")
        
        mock_messagebox.assert_called_once_with("Test Title", "Test success")
    
    @patch('tkinter.messagebox.askyesno')
    def test_show_confirmation(self, mock_messagebox, handler):
        """Test confirmation dialog."""
        mock_messagebox.return_value = True
        
        result = handler.show_confirmation("Confirm?", "Confirmation")
        
        assert result is True
        mock_messagebox.assert_called_once_with("Confirmation", "Confirm?")
    
    def test_operation_history(self, handler):
        """Test operation history management."""
        # Initially empty
        assert handler.get_operation_history() == []
        
        # Execute operations
        validation_result = ValidationResult(valid=True)
        operation_result = OperationResult(success=True)
        
        handler.set_validation_result(validation_result)
        handler.set_operation_result(operation_result)
        
        handler.execute({"test": "data1"})
        handler.execute({"test": "data2"})
        
        # Check history
        history = handler.get_operation_history()
        assert len(history) == 2
        
        # Clear history
        handler.clear_history()
        assert len(handler.get_operation_history()) == 0
    
    def test_validate_workbook_path_valid(self, handler):
        """Test workbook path validation with valid path."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            result = handler.validate_workbook_path(tmp_path)
            assert result.valid is True
            assert result.errors == []
        finally:
            os.unlink(tmp_path)
    
    def test_validate_workbook_path_invalid(self, handler):
        """Test workbook path validation with invalid paths."""
        # Empty path
        result = handler.validate_workbook_path("")
        assert result.valid is False
        assert "required" in result.errors[0].lower()
        
        # Non-existent file
        result = handler.validate_workbook_path("/nonexistent/file.xlsx")
        assert result.valid is False
        assert "does not exist" in result.errors[0]
        
        # Wrong extension
        with tempfile.NamedTemporaryFile(suffix='.txt', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            result = handler.validate_workbook_path(tmp_path)
            assert result.valid is False
            assert "Excel file" in result.errors[0]
        finally:
            os.unlink(tmp_path)
    
    def test_validate_required_fields(self, handler):
        """Test required fields validation."""
        required_fields = ["name", "email", "phone"]
        
        # Valid data
        data = {"name": "John", "email": "john@example.com", "phone": "123456"}
        result = handler.validate_required_fields(data, required_fields)
        assert result.valid is True
        
        # Missing field
        data = {"name": "John", "email": "john@example.com"}
        result = handler.validate_required_fields(data, required_fields)
        assert result.valid is False
        assert "phone is required" in result.errors
        
        # Empty field
        data = {"name": "", "email": "john@example.com", "phone": "123456"}
        result = handler.validate_required_fields(data, required_fields)
        assert result.valid is False
        assert "name is required" in result.errors
    
    def test_validate_numeric_fields(self, handler):
        """Test numeric fields validation."""
        numeric_fields = ["age", "salary"]
        
        # Valid data
        data = {"age": "25", "salary": "50000.50"}
        result = handler.validate_numeric_fields(data, numeric_fields)
        assert result.valid is True
        
        # Invalid numeric value
        data = {"age": "not_a_number", "salary": "50000"}
        result = handler.validate_numeric_fields(data, numeric_fields)
        assert result.valid is False
        assert "age must be a valid number" in result.errors
        
        # None values (should be ignored)
        data = {"age": None, "salary": "50000"}
        result = handler.validate_numeric_fields(data, numeric_fields)
        assert result.valid is True


class TestExcelOperationHandler:
    """Test ExcelOperationHandler class."""
    
    @pytest.fixture
    def mock_parent(self):
        """Create mock parent window."""
        return Mock(spec=tk.Widget)
    
    @pytest.fixture
    def temp_excel_file(self):
        """Create temporary Excel file for testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            # Create a simple Excel file
            wb = openpyxl.Workbook()
            wb.active.title = "Sheet1"
            wb.create_sheet("Sheet2")
            wb.save(tmp.name)
            wb.close()
            yield tmp.name
        os.unlink(tmp.name)
    
    def test_init(self, mock_parent, temp_excel_file):
        """Test Excel handler initialization."""
        handler = ExcelOperationHandler(mock_parent, temp_excel_file)
        
        assert handler.parent == mock_parent
        assert handler.workbook_path == temp_excel_file
        assert handler._workbook is None
    
    @patch('handlers.base_handler.ExcelManager')
    def test_open_workbook(self, mock_excel_manager, mock_parent, temp_excel_file):
        """Test workbook opening."""
        handler = ExcelOperationHandler(mock_parent, temp_excel_file)
        
        result = handler.open_workbook()
        
        mock_excel_manager.assert_called_once_with(temp_excel_file)
        assert handler._workbook is not None
    
    def test_open_workbook_no_path(self, mock_parent):
        """Test workbook opening without path."""
        handler = ExcelOperationHandler(mock_parent, None)
        
        with pytest.raises(ValueError, match="Workbook path not specified"):
            handler.open_workbook()
    
    def test_close_workbook(self, mock_parent, temp_excel_file):
        """Test workbook closing."""
        handler = ExcelOperationHandler(mock_parent, temp_excel_file)
        
        # Mock workbook
        mock_workbook = Mock()
        handler._workbook = mock_workbook
        
        handler.close_workbook()
        
        mock_workbook.force_close.assert_called_once()
        assert handler._workbook is None
    
    def test_get_sheet_names(self, mock_parent, temp_excel_file):
        """Test getting sheet names."""
        handler = ExcelOperationHandler(mock_parent, temp_excel_file)
        
        sheet_names = handler.get_sheet_names()
        
        assert "Sheet1" in sheet_names
        assert "Sheet2" in sheet_names
    
    def test_sheet_exists(self, mock_parent, temp_excel_file):
        """Test sheet existence check."""
        handler = ExcelOperationHandler(mock_parent, temp_excel_file)
        
        assert handler.sheet_exists("Sheet1") is True
        assert handler.sheet_exists("NonExistent") is False


class TestDialogHandler:
    """Test DialogHandler class."""
    
    @pytest.fixture
    def mock_parent(self):
        """Create mock parent window."""
        return Mock(spec=tk.Widget)
    
    @pytest.fixture
    def mock_dialog_class(self):
        """Create mock dialog class."""
        mock_dialog = Mock()
        mock_dialog.show_modal.return_value = "dialog_result"
        
        mock_dialog_class = Mock()
        mock_dialog_class.return_value = mock_dialog
        
        return mock_dialog_class
    
    def test_init(self, mock_parent, mock_dialog_class):
        """Test dialog handler initialization."""
        handler = DialogHandler(mock_parent, mock_dialog_class)
        
        assert handler.parent == mock_parent
        assert handler.dialog_class == mock_dialog_class
        assert handler._current_dialog is None
    
    def test_show_dialog(self, mock_parent, mock_dialog_class):
        """Test showing dialog."""
        handler = DialogHandler(mock_parent, mock_dialog_class)
        
        result = handler.show_dialog("arg1", kwarg1="value1")
        
        mock_dialog_class.assert_called_once_with(mock_parent, "arg1", kwarg1="value1")
        assert result == "dialog_result"
        assert handler._current_dialog is None  # Cleaned up after use
    
    def test_show_dialog_exception(self, mock_parent, mock_dialog_class):
        """Test dialog showing with exception."""
        mock_dialog_class.side_effect = Exception("Dialog error")
        
        handler = DialogHandler(mock_parent, mock_dialog_class)
        
        with patch.object(handler, 'show_error') as mock_show_error:
            result = handler.show_dialog()
            
            assert result is None
            mock_show_error.assert_called_once()


class TestBatchOperationHandler:
    """Test BatchOperationHandler class."""
    
    @pytest.fixture
    def mock_parent(self):
        """Create mock parent window."""
        return Mock(spec=tk.Widget)
    
    @pytest.fixture
    def sample_items(self):
        """Create sample items for batch processing."""
        return [
            {"id": 1, "name": "Item 1"},
            {"id": 2, "name": "Item 2"},
            {"id": 3, "name": "Item 3"}
        ]
    
    def test_init(self, mock_parent, sample_items):
        """Test batch handler initialization."""
        handler = BatchOperationHandler(mock_parent, sample_items)
        
        assert handler.parent == mock_parent
        assert handler.items == sample_items
        assert handler.processed_count == 0
        assert handler.error_count == 0
    
    def test_process_batch_success(self, mock_parent, sample_items):
        """Test successful batch processing."""
        handler = BatchOperationHandler(mock_parent, sample_items)
        
        def mock_process_func(item):
            return OperationResult(success=True, message=f"Processed {item['name']}")
        
        results = handler.process_batch(mock_process_func)
        
        assert len(results) == 3
        assert all(result.success for result in results)
        assert handler.processed_count == 3
        assert handler.error_count == 0
    
    def test_process_batch_with_errors(self, mock_parent, sample_items):
        """Test batch processing with errors."""
        handler = BatchOperationHandler(mock_parent, sample_items)
        
        def mock_process_func(item):
            if item['id'] == 2:
                return OperationResult(success=False, message="Failed")
            return OperationResult(success=True, message="Success")
        
        results = handler.process_batch(mock_process_func)
        
        assert len(results) == 3
        assert handler.processed_count == 2
        assert handler.error_count == 1
        assert not results[1].success  # Second item failed
    
    def test_process_batch_with_exceptions(self, mock_parent, sample_items):
        """Test batch processing with exceptions."""
        handler = BatchOperationHandler(mock_parent, sample_items)
        
        def mock_process_func(item):
            if item['id'] == 2:
                raise ValueError("Processing error")
            return OperationResult(success=True, message="Success")
        
        results = handler.process_batch(mock_process_func)
        
        assert len(results) == 3
        assert handler.processed_count == 2
        assert handler.error_count == 1
        assert not results[1].success  # Second item failed with exception
    
    def test_get_batch_summary(self, mock_parent, sample_items):
        """Test batch summary generation."""
        handler = BatchOperationHandler(mock_parent, sample_items)
        
        def mock_process_func(item):
            if item['id'] == 3:
                return OperationResult(success=False, message="Failed")
            return OperationResult(success=True, message="Success")
        
        handler.process_batch(mock_process_func)
        summary = handler.get_batch_summary()
        
        assert summary['total_items'] == 3
        assert summary['processed'] == 2
        assert summary['errors'] == 1
        assert summary['success_rate'] == pytest.approx(66.67, rel=1e-2)


@pytest.mark.integration
class TestHandlerIntegration:
    """Integration tests for handler classes."""
    
    @pytest.fixture
    def temp_excel_file(self):
        """Create temporary Excel file for integration testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Partners"
            
            # Add some test data
            ws['A1'] = "Partner Number"
            ws['B1'] = "Partner Name"
            ws['A2'] = "P1"
            ws['B2'] = "Lead Partner"
            
            wb.save(tmp.name)
            wb.close()
            yield tmp.name
        os.unlink(tmp.name)
    
    def test_excel_handler_with_real_file(self, temp_excel_file):
        """Test Excel handler with real Excel file."""
        handler = ExcelOperationHandler(None, temp_excel_file)
        
        # Test opening workbook
        workbook = handler.open_workbook()
        assert workbook is not None
        
        # Test getting sheet names
        sheet_names = handler.get_sheet_names()
        assert "Partners" in sheet_names
        
        # Test sheet existence
        assert handler.sheet_exists("Partners") is True
        assert handler.sheet_exists("NonExistent") is False
        
        # Test closing
        handler.close_workbook()
        assert handler._workbook is None
