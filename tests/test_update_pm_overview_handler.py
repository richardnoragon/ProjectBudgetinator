"""
Test cases for PM Overview update handler functionality.

This module contains comprehensive tests for the PM Overview update handler,
including formula copying, cell mapping, debug functionality, and integration tests.
"""

import unittest
from unittest.mock import Mock, patch, MagicMock
import tkinter as tk
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Import the modules to test
from handlers.update_pm_overview_handler import (
    UpdatePMOverviewHandler,
    PMOverviewDebugWindow,
    get_pm_overview_row,
    get_partner_number_from_sheet_name,
    adjust_formula_references,
    update_pm_overview_after_partner_operation,
    update_pm_overview_with_progress,
    PM_OVERVIEW_CELL_MAPPINGS
)
from handlers.pm_overview_format import (
    PMOverviewFormatter,
    RowAnalyzer,
    RowStatus,
    apply_pm_overview_formatting
)
from handlers.base_handler import ValidationResult, OperationResult


class TestPMOverviewCellMappings(unittest.TestCase):
    """Test PM Overview cell mapping calculations."""
    
    def test_get_pm_overview_row(self):
        """Test PM Overview row calculation."""
        # Test the mapping: Partner 2 → Row 6, Partner 3 → Row 7, etc.
        self.assertEqual(get_pm_overview_row(2), 6)
        self.assertEqual(get_pm_overview_row(3), 7)
        self.assertEqual(get_pm_overview_row(4), 8)
        self.assertEqual(get_pm_overview_row(10), 14)
        self.assertEqual(get_pm_overview_row(20), 24)
    
    def test_get_partner_number_from_sheet_name(self):
        """Test partner number extraction from sheet names."""
        # Valid sheet names
        self.assertEqual(get_partner_number_from_sheet_name("P2-ACME"), 2)
        self.assertEqual(get_partner_number_from_sheet_name("P3-University"), 3)
        self.assertEqual(get_partner_number_from_sheet_name("P10-LongPartnerName"), 10)
        self.assertEqual(get_partner_number_from_sheet_name("P20-LastPartner"), 20)
        
        # Invalid sheet names
        self.assertIsNone(get_partner_number_from_sheet_name("P1-Coordinator"))  # P1 not allowed
        self.assertIsNone(get_partner_number_from_sheet_name("P21-TooHigh"))     # P21 not allowed
        self.assertIsNone(get_partner_number_from_sheet_name("NotAPartner"))     # Doesn't start with P
        self.assertIsNone(get_partner_number_from_sheet_name("P-NoNumber"))      # No number
        self.assertIsNone(get_partner_number_from_sheet_name("PA-Invalid"))      # Invalid number
    
    def test_pm_overview_cell_mappings(self):
        """Test that PM Overview cell mappings are correctly defined."""
        expected_mappings = {
            'C18': 'C', 'D18': 'D', 'E18': 'E', 'F18': 'F', 'G18': 'G',
            'H18': 'H', 'I18': 'I', 'J18': 'J', 'K18': 'K', 'L18': 'L',
            'M18': 'M', 'N18': 'N', 'O18': 'O', 'P18': 'P', 'Q18': 'Q'
        }
        
        self.assertEqual(PM_OVERVIEW_CELL_MAPPINGS, expected_mappings)
        self.assertEqual(len(PM_OVERVIEW_CELL_MAPPINGS), 15)  # WP1-WP15


class TestFormulaAdjustment(unittest.TestCase):
    """Test formula reference adjustment functionality."""
    
    def test_adjust_formula_references_simple(self):
        """Test simple formula reference adjustment."""
        # Test relative references
        formula = "=SUM(A1:A10)"
        adjusted = adjust_formula_references(formula, 18, 6, "P2", "PM Overview")
        # Row difference: 6 - 18 = -12, so A1 becomes A-11 (invalid), A10 becomes A-2 (invalid)
        # This is expected behavior for relative references
        expected = "=SUM(A-11:A-2)"
        self.assertEqual(adjusted, expected)
        
        # Test with positive row difference
        formula = "=SUM(B5:B15)"
        adjusted = adjust_formula_references(formula, 6, 18, "PM Overview", "P2")
        # Row difference: 18 - 6 = 12, so B5 becomes B17, B15 becomes B27
        expected = "=SUM(B17:B27)"
        self.assertEqual(adjusted, expected)
    
    def test_adjust_formula_references_absolute(self):
        """Test absolute formula reference adjustment."""
        # Test absolute row references (should not change)
        formula = "=SUM($A$1:$A$10)"
        adjusted = adjust_formula_references(formula, 18, 6, "P2", "PM Overview")
        expected = "=SUM($A$1:$A$10)"  # Absolute references unchanged
        self.assertEqual(adjusted, expected)
        
        # Test mixed references
        formula = "=SUM(A$1:$A10)"
        adjusted = adjust_formula_references(formula, 18, 6, "P2", "PM Overview")
        expected = "=SUM(A$1:$A-2)"  # Only relative parts change
        self.assertEqual(adjusted, expected)
    
    def test_adjust_formula_references_complex(self):
        """Test complex formula reference adjustment."""
        # Test formula with multiple references
        formula = "=IF(A1>0, SUM(B1:B5), C1*2)"
        adjusted = adjust_formula_references(formula, 18, 6, "P2", "PM Overview")
        expected = "=IF(A-11>0, SUM(B-11:B-7), C-11*2)"
        self.assertEqual(adjusted, expected)
    
    def test_adjust_formula_references_no_formula(self):
        """Test adjustment with non-formula values."""
        # Test non-formula values
        self.assertEqual(adjust_formula_references("123", 18, 6, "P2", "PM Overview"), "123")
        self.assertEqual(adjust_formula_references("Text", 18, 6, "P2", "PM Overview"), "Text")
        self.assertEqual(adjust_formula_references("", 18, 6, "P2", "PM Overview"), "")
        self.assertEqual(adjust_formula_references(None, 18, 6, "P2", "PM Overview"), None)


class TestPMOverviewDebugWindow(unittest.TestCase):
    """Test PM Overview debug window functionality."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.root = tk.Tk()
        self.root.withdraw()  # Hide the window during tests
        self.debug_window = PMOverviewDebugWindow(self.root)
    
    def tearDown(self):
        """Clean up test fixtures."""
        if self.debug_window.debug_window:
            self.debug_window.debug_window.destroy()
        self.root.destroy()
    
    def test_debug_window_creation(self):
        """Test debug window creation."""
        debug_data = [
            {
                'source_sheet': 'P2-ACME',
                'source_cell': 'C18',
                'target_cell': 'C6',
                'original_formula': '=SUM(A1:A10)',
                'adjusted_formula': '=SUM(A-11:A-2)',
                'value': 100
            }
        ]
        
        # This should not raise an exception
        self.debug_window.show_debug_info("Test Debug", debug_data)
        self.assertIsNotNone(self.debug_window.debug_window)
    
    def test_generate_debug_content(self):
        """Test debug content generation."""
        debug_data = [
            {
                'source_sheet': 'P2-ACME',
                'source_cell': 'C18',
                'target_cell': 'C6',
                'original_formula': '=SUM(A1:A10)',
                'adjusted_formula': '=SUM(A-11:A-2)',
                'value': 100
            },
            {
                'source_sheet': 'P3-University',
                'source_cell': 'D18',
                'target_cell': 'D7',
                'original_formula': '=B1*2',
                'value': 200,
                'error': 'Test error'
            }
        ]
        
        content = self.debug_window._generate_debug_content(debug_data)
        
        # Check that content contains expected elements
        self.assertIn("PM OVERVIEW UPDATE DEBUG INFORMATION", content)
        self.assertIn("Total Operations: 2", content)
        self.assertIn("P2-ACME", content)
        self.assertIn("P3-University", content)
        self.assertIn("C18", content)
        self.assertIn("D18", content)
        self.assertIn("Test error", content)


class TestUpdatePMOverviewHandler(unittest.TestCase):
    """Test the main PM Overview update handler."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.root = tk.Tk()
        self.root.withdraw()  # Hide the window during tests
        self.handler = UpdatePMOverviewHandler(self.root)
        
        # Create a mock workbook
        self.workbook = Mock(spec=Workbook)
        self.workbook.sheetnames = ["PM Overview", "P2-ACME", "P3-University"]
        
        # Create mock worksheets
        self.pm_overview_ws = Mock(spec=Worksheet)
        self.partner2_ws = Mock(spec=Worksheet)
        self.partner3_ws = Mock(spec=Worksheet)
        
        # Configure workbook to return appropriate worksheets
        def get_worksheet(name):
            if name == "PM Overview":
                return self.pm_overview_ws
            elif name == "P2-ACME":
                return self.partner2_ws
            elif name == "P3-University":
                return self.partner3_ws
            else:
                raise KeyError(f"Worksheet {name} not found")
        
        self.workbook.__getitem__ = get_worksheet
        
        # Configure partner worksheets
        self.partner2_ws.title = "P2-ACME"
        self.partner3_ws.title = "P3-University"
    
    def tearDown(self):
        """Clean up test fixtures."""
        self.root.destroy()
    
    def test_validate_input_success(self):
        """Test successful input validation."""
        data = {'workbook': self.workbook}
        result = self.handler.validate_input(data)
        
        self.assertTrue(result.valid)
        self.assertEqual(len(result.errors), 0)
    
    def test_validate_input_no_workbook(self):
        """Test validation failure when no workbook provided."""
        data = {}
        result = self.handler.validate_input(data)
        
        self.assertFalse(result.valid)
        self.assertIn("Workbook is required", result.errors)
    
    def test_validate_input_no_pm_overview(self):
        """Test validation failure when PM Overview worksheet missing."""
        self.workbook.sheetnames = ["P2-ACME", "P3-University"]  # No PM Overview
        data = {'workbook': self.workbook}
        result = self.handler.validate_input(data)
        
        self.assertFalse(result.valid)
        self.assertIn("'PM Overview' worksheet not found", result.errors)
    
    def test_get_partner_worksheets(self):
        """Test partner worksheet discovery."""
        partner_sheets = self.handler.get_partner_worksheets(self.workbook)
        
        expected = [("P2-ACME", 2), ("P3-University", 3)]
        self.assertEqual(partner_sheets, expected)
    
    def test_extract_partner_formulas(self):
        """Test formula extraction from partner worksheet."""
        # Mock cell values
        mock_cells = {}
        for source_cell in PM_OVERVIEW_CELL_MAPPINGS.keys():
            mock_cell = Mock()
            mock_cell.value = f"=SUM(A1:A10)"  # Simple formula
            mock_cell.data_type = 'f'  # Formula type
            mock_cell.displayed_value = "100"
            mock_cells[source_cell] = mock_cell
        
        def get_cell(cell_ref):
            return mock_cells.get(cell_ref, Mock(value=None))
        
        self.partner2_ws.__getitem__ = get_cell
        
        partner_data, debug_data = self.handler.extract_partner_formulas(self.partner2_ws, 2)
        
        # Check partner data structure
        self.assertEqual(partner_data['partner_number'], 2)
        self.assertIn('formula_mappings', partner_data)
        
        # Check that all expected cells were processed
        self.assertEqual(len(partner_data['formula_mappings']), len(PM_OVERVIEW_CELL_MAPPINGS))
        
        # Check debug data
        self.assertEqual(len(debug_data), len(PM_OVERVIEW_CELL_MAPPINGS))
        
        # Check specific mapping
        c18_mapping = partner_data['formula_mappings']['C18']
        self.assertEqual(c18_mapping['formula'], "=SUM(A1:A10)")
        self.assertEqual(c18_mapping['target_col'], 'C')
    
    @patch('handlers.update_pm_overview_handler.logger')
    def test_update_pm_overview_row(self, mock_logger):
        """Test updating a PM Overview row with formulas."""
        # Prepare partner data
        partner_data = {
            'partner_number': 2,
            'formula_mappings': {
                'C18': {'formula': '=SUM(A1:A10)', 'target_col': 'C'},
                'D18': {'formula': '=B1*2', 'target_col': 'D'}
            }
        }
        
        debug_data = [
            {'source_cell': 'C18', 'target_cell': 'C6'},
            {'source_cell': 'D18', 'target_cell': 'D6'}
        ]
        
        # Mock PM Overview worksheet
        mock_cells = {}
        def set_cell(cell_ref, value):
            mock_cells[cell_ref] = value
        
        def get_cell(cell_ref):
            mock_cell = Mock()
            mock_cell.value = mock_cells.get(cell_ref)
            return mock_cell
        
        self.pm_overview_ws.__setitem__ = set_cell
        self.pm_overview_ws.__getitem__ = get_cell
        
        # Execute the update
        self.handler.update_pm_overview_row(self.pm_overview_ws, partner_data, debug_data)
        
        # Check that formulas were adjusted and set
        self.assertIn('C6', mock_cells)
        self.assertIn('D6', mock_cells)
        
        # The formulas should be adjusted (row 18 -> row 6, difference = -12)
        self.assertEqual(mock_cells['C6'], '=SUM(A-11:A-2)')
        self.assertEqual(mock_cells['D6'], '=B-11*2')


class TestPMOverviewFormatter(unittest.TestCase):
    """Test PM Overview formatting functionality."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.root = tk.Tk()
        self.root.withdraw()  # Hide the window during tests
        self.formatter = PMOverviewFormatter(self.root)
        
        # Create a mock workbook with PM Overview
        self.workbook = Mock(spec=Workbook)
        self.workbook.sheetnames = ["PM Overview", "P2-ACME", "P3-University"]
        
        self.pm_overview_ws = Mock(spec=Worksheet)
        self.workbook.__getitem__ = lambda name: self.pm_overview_ws if name == "PM Overview" else Mock()
    
    def tearDown(self):
        """Clean up test fixtures."""
        self.root.destroy()
    
    def test_row_analyzer_calculations(self):
        """Test row analyzer calculations."""
        # Test PM Overview row calculations (different from Budget Overview)
        self.assertEqual(RowAnalyzer.get_pm_overview_row(2), 6)
        self.assertEqual(RowAnalyzer.get_pm_overview_row(3), 7)
        
        self.assertEqual(RowAnalyzer.get_partner_number_from_row(6), 2)
        self.assertEqual(RowAnalyzer.get_partner_number_from_row(7), 3)
    
    def test_get_partner_rows(self):
        """Test partner row discovery in PM Overview."""
        # Mock worksheet cells
        mock_cells = {}
        for row in range(6, 10):  # Rows 6-9 (partners 2-5)
            mock_cell = Mock()
            mock_cell.value = f"Partner {row-4} data"
            mock_cells[f'C{row}'] = mock_cell
        
        def get_cell(cell_ref):
            return mock_cells.get(cell_ref, Mock(value=None))
        
        self.pm_overview_ws.__getitem__ = get_cell
        
        partner_rows = self.formatter.get_partner_rows(self.pm_overview_ws)
        expected_rows = [6, 7, 8, 9]  # Partners 2, 3, 4, 5
        self.assertEqual(partner_rows, expected_rows)
    
    @patch('handlers.pm_overview_format.messagebox')
    def test_apply_conditional_formatting_no_openpyxl(self, mock_messagebox):
        """Test formatting when openpyxl is not available."""
        with patch('handlers.pm_overview_format.OPENPYXL_AVAILABLE', False):
            result = self.formatter.apply_conditional_formatting(self.workbook)
            
            self.assertFalse(result)
            mock_messagebox.showerror.assert_called_once()
    
    def test_apply_conditional_formatting_no_pm_overview(self):
        """Test formatting when PM Overview worksheet is missing."""
        self.workbook.sheetnames = ["P2-ACME", "P3-University"]  # No PM Overview
        
        result = self.formatter.apply_conditional_formatting(self.workbook)
        self.assertFalse(result)


class TestIntegrationFunctions(unittest.TestCase):
    """Test integration functions for automatic updates."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.workbook = Mock(spec=Workbook)
        self.workbook.sheetnames = ["PM Overview", "P2-ACME"]
    
    @patch('handlers.update_pm_overview_handler.UpdatePMOverviewHandler')
    def test_update_pm_overview_after_partner_operation(self, mock_handler_class):
        """Test automatic PM Overview update after partner operation."""
        # Mock handler instance
        mock_handler = Mock()
        mock_handler.execute.return_value = OperationResult(success=True, message="Success")
        mock_handler_class.return_value = mock_handler
        
        # Test successful update
        result = update_pm_overview_after_partner_operation(self.workbook, 2)
        
        self.assertTrue(result)
        mock_handler_class.assert_called_once()
        mock_handler.execute.assert_called_once_with({
            'workbook': self.workbook,
            'partner_number': 2
        })
    
    @patch('handlers.update_pm_overview_handler.UpdatePMOverviewHandler')
    def test_update_pm_overview_after_partner_operation_failure(self, mock_handler_class):
        """Test automatic PM Overview update failure."""
        # Mock handler instance that fails
        mock_handler = Mock()
        mock_handler.execute.return_value = OperationResult(success=False, message="Failed")
        mock_handler_class.return_value = mock_handler
        
        # Test failed update
        result = update_pm_overview_after_partner_operation(self.workbook, 2)
        
        self.assertFalse(result)
    
    @patch('handlers.update_pm_overview_handler.show_progress_for_operation')
    @patch('handlers.update_pm_overview_handler.UpdatePMOverviewHandler')
    def test_update_pm_overview_with_progress(self, mock_handler_class, mock_progress):
        """Test PM Overview update with progress dialog."""
        # Mock progress operation
        mock_progress.return_value = True
        
        # Mock parent window
        parent_window = Mock()
        
        result = update_pm_overview_with_progress(parent_window, self.workbook)
        
        self.assertTrue(result)
        mock_progress.assert_called_once()


class TestPMOverviewEndToEnd(unittest.TestCase):
    """End-to-end integration tests for PM Overview functionality."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.root = tk.Tk()
        self.root.withdraw()  # Hide the window during tests
    
    def tearDown(self):
        """Clean up test fixtures."""
        self.root.destroy()
    
    def test_complete_pm_overview_workflow(self):
        """Test complete PM Overview update workflow."""
        # Create a real workbook for testing
        workbook = Workbook()
        
        # Create PM Overview worksheet
        pm_overview = workbook.create_sheet("PM Overview")
        
        # Create partner worksheets
        partner2 = workbook.create_sheet("P2-ACME")
        partner3 = workbook.create_sheet("P3-University")
        
        # Add some test formulas to partner worksheets
        partner2['C18'] = "=SUM(A1:A10)"
        partner2['D18'] = "=B1*2"
        partner3['C18'] = "=SUM(C1:C5)"
        partner3['D18'] = "=D1+100"
        
        # Remove default worksheet
        workbook.remove(workbook['Sheet'])
        
        # Create handler and perform update
        handler = UpdatePMOverviewHandler(self.root)
        
        # Test validation
        validation = handler.validate_input({'workbook': workbook})
        self.assertTrue(validation.valid)
        
        # Test partner worksheet discovery
        partner_sheets = handler.get_partner_worksheets(workbook)
        expected = [("P2-ACME", 2), ("P3-University", 3)]
        self.assertEqual(partner_sheets, expected)
        
        # Test formula extraction
        partner_data, debug_data = handler.extract_partner_formulas(partner2, 2)
        self.assertEqual(partner_data['partner_number'], 2)
        self.assertIn('formula_mappings', partner_data)
        
        # Test PM Overview update
        handler.update_pm_overview_row(pm_overview, partner_data, debug_data)
        
        # Verify that formulas were copied and adjusted
        # Partner 2 should go to row 6
        self.assertIsNotNone(pm_overview['C6'].value)
        self.assertIsNotNone(pm_overview['D6'].value)
        
        # Check that formulas were adjusted (row 18 -> row 6, difference = -12)
        self.assertEqual(pm_overview['C6'].value, "=SUM(A-11:A-2)")
        self.assertEqual(pm_overview['D6'].value, "=B-11*2")


if __name__ == '__main__':
    # Run the tests
    unittest.main()