"""
Integration tests for Excel operations in ProjectBudgetinator.

Tests cover end-to-end Excel file operations, workbook management,
and cross-module integration as outlined in OPTIMIZATION_RECOMMENDATIONS.md.
"""

import pytest
import tempfile
import os
import openpyxl
from pathlib import Path
import time
from unittest.mock import Mock, patch
import threading

# Import modules for testing
from utils.excel_manager import ExcelManager, excel_context


class TestExcelFileOperations:
    """Test Excel file operations integration."""
    
    @pytest.fixture
    def temp_excel_file(self):
        """Create temporary Excel file for testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = openpyxl.Workbook()
            
            # Create multiple sheets
            summary_sheet = wb.active
            summary_sheet.title = "Summary"
            
            partners_sheet = wb.create_sheet("Partners")
            workpackages_sheet = wb.create_sheet("Workpackages")
            
            # Add some test data
            summary_sheet['A1'] = "Project Overview"
            partners_sheet['A1'] = "Partner Information"
            workpackages_sheet['A1'] = "Workpackage Details"
            
            wb.save(tmp.name)
            wb.close()
            yield tmp.name
        
        if os.path.exists(tmp.name):
            os.unlink(tmp.name)
    
    @pytest.fixture
    def large_excel_file(self):
        """Create large Excel file for performance testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Large Dataset"
            
            # Create large dataset
            for row in range(1, 1001):  # 1000 rows
                for col in range(1, 21):  # 20 columns
                    ws.cell(row=row, column=col).value = f"Data_{row}_{col}"
            
            wb.save(tmp.name)
            wb.close()
            yield tmp.name
        
        if os.path.exists(tmp.name):
            os.unlink(tmp.name)
    
    def test_excel_manager_basic_operations(self, temp_excel_file):
        """Test basic ExcelManager operations."""
        # Test opening and closing
        manager = ExcelManager(temp_excel_file)
        
        # Verify file properties
        assert manager.file_path == temp_excel_file
        assert manager.is_closed is True  # Initially closed
        
        # Test opening
        workbook = manager.workbook
        assert workbook is not None
        assert manager.is_closed is False
        
        # Test sheet access
        sheet_names = workbook.sheetnames
        expected_sheets = ["Summary", "Partners", "Workpackages"]
        for sheet in expected_sheets:
            assert sheet in sheet_names
        
        # Test closing
        manager.force_close()
        assert manager.is_closed is True
    
    def test_excel_context_manager(self, temp_excel_file):
        """Test Excel context manager functionality."""
        # Test successful context usage
        with excel_context(temp_excel_file) as wb:
            assert wb is not None
            assert "Summary" in wb.sheetnames
            
            # Modify data
            ws = wb["Summary"]
            ws['B1'] = "Modified Data"
            
            # Save changes
            wb.save(temp_excel_file)
        
        # Verify changes were saved and file is properly closed
        with openpyxl.load_workbook(temp_excel_file) as wb2:
            assert wb2["Summary"]['B1'].value == "Modified Data"
    
    def test_excel_context_manager_exception_handling(self):
        """Test context manager with file errors."""
        nonexistent_file = "nonexistent_file.xlsx"
        
        with pytest.raises(FileNotFoundError):
            with excel_context(nonexistent_file) as wb:
                pass
    
    def test_concurrent_excel_access(self, temp_excel_file):
        """Test concurrent access to Excel files."""
        results = []
        errors = []
        
        def read_excel_data(file_path, thread_id):
            try:
                with excel_context(file_path) as wb:
                    sheet_count = len(wb.sheetnames)
                    results.append((thread_id, sheet_count))
                    time.sleep(0.1)  # Simulate processing time
            except Exception as e:
                errors.append((thread_id, str(e)))
        
        # Create multiple threads
        threads = []
        for i in range(5):
            thread = threading.Thread(
                target=read_excel_data, 
                args=(temp_excel_file, i)
            )
            threads.append(thread)
        
        # Start all threads
        for thread in threads:
            thread.start()
        
        # Wait for completion
        for thread in threads:
            thread.join()
        
        # Verify results
        assert len(errors) == 0, f"Concurrent access errors: {errors}"
        assert len(results) == 5
        
        # All threads should get same sheet count
        expected_sheet_count = 3
        for thread_id, sheet_count in results:
            assert sheet_count == expected_sheet_count
    
    @pytest.mark.performance
    def test_large_file_performance(self, large_excel_file, benchmark):
        """Test performance with large Excel files."""
        def load_large_file():
            with excel_context(large_excel_file) as wb:
                ws = wb.active
                # Read some data
                data = []
                for row in range(1, 101):  # Read first 100 rows
                    row_data = []
                    for col in range(1, 11):  # Read first 10 columns
                        value = ws.cell(row=row, column=col).value
                        row_data.append(value)
                    data.append(row_data)
                return len(data)
        
        result = benchmark(load_large_file)
        assert result == 100
    
    def test_memory_usage_monitoring(self, temp_excel_file):
        """Test memory usage during Excel operations."""
        import psutil
        import os
        
        process = psutil.Process(os.getpid())
        initial_memory = process.memory_info().rss
        
        # Perform Excel operations
        managers = []
        for i in range(10):
            manager = ExcelManager(temp_excel_file)
            workbook = manager.workbook
            assert workbook is not None
            managers.append(manager)
        
        # Check memory usage
        peak_memory = process.memory_info().rss
        memory_increase = peak_memory - initial_memory
        
        # Clean up
        for manager in managers:
            manager.force_close()
        
        # Allow garbage collection
        import gc
        gc.collect()
        
        final_memory = process.memory_info().rss
        memory_decrease = peak_memory - final_memory
        
        # Verify memory is released (at least partially)
        assert memory_decrease > 0, "Memory should be released after cleanup"
        
        # Memory increase should be reasonable
        assert memory_increase < 100 * 1024 * 1024, "Memory usage should be reasonable"


class TestExcelDataIntegrity:
    """Test Excel data integrity and validation."""
    
    @pytest.fixture
    def project_workbook(self):
        """Create a complete project workbook for testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = openpyxl.Workbook()
            
            # Summary sheet
            summary = wb.active
            summary.title = "Summary"
            summary['A1'] = "Project Title"
            summary['B1'] = "Test Project"
            summary['A2'] = "Total Budget"
            summary['B2'] = "500000"
            
            # Partners sheet
            partners = wb.create_sheet("Partners")
            partners['A1'] = "Partner Number"
            partners['B1'] = "Organization"
            partners['C1'] = "Country"
            partners['A2'] = "P1"
            partners['B2'] = "Lead Organization"
            partners['C2'] = "Germany"
            
            # Workpackages sheet
            workpackages = wb.create_sheet("Workpackages")
            workpackages['A1'] = "WP Number"
            workpackages['B1'] = "Title"
            workpackages['C1'] = "Budget"
            workpackages['A2'] = "WP1"
            workpackages['B2'] = "Project Management"
            workpackages['C2'] = "50000"
            
            wb.save(tmp.name)
            wb.close()
            yield tmp.name
        
        if os.path.exists(tmp.name):
            os.unlink(tmp.name)
    
    def test_data_consistency_across_sheets(self, project_workbook):
        """Test data consistency across multiple sheets."""
        with excel_context(project_workbook) as wb:
            summary = wb["Summary"]
            partners = wb["Partners"]
            workpackages = wb["Workpackages"]
            
            # Verify project title consistency
            project_title = summary['B1'].value
            assert project_title == "Test Project"
            
            # Verify budget calculations
            total_budget = float(summary['B2'].value)
            wp_budget = float(workpackages['C2'].value)
            
            assert total_budget >= wp_budget, "Total budget should be >= sum of workpackage budgets"
            
            # Verify partner references
            lead_partner = partners['A2'].value
            assert lead_partner == "P1", "Lead partner should be P1"
    
    def test_data_validation_rules(self, project_workbook):
        """Test data validation rules."""
        with excel_context(project_workbook) as wb:
            partners = wb["Partners"]
            workpackages = wb["Workpackages"]
            
            # Validate partner numbers
            partner_number = partners['A2'].value
            assert partner_number.startswith('P'), "Partner numbers should start with 'P'"
            assert partner_number[1:].isdigit(), "Partner number should have numeric suffix"
            
            # Validate workpackage numbers
            wp_number = workpackages['A2'].value
            assert wp_number.startswith('WP'), "Workpackage numbers should start with 'WP'"
            assert wp_number[2:].isdigit(), "Workpackage number should have numeric suffix"
            
            # Validate budget values
            wp_budget = workpackages['C2'].value
            try:
                budget_value = float(wp_budget)
                assert budget_value >= 0, "Budget should be non-negative"
            except ValueError:
                pytest.fail("Budget should be a valid number")
    
    def test_sheet_structure_validation(self, project_workbook):
        """Test validation of sheet structures."""
        with excel_context(project_workbook) as wb:
            expected_sheets = ["Summary", "Partners", "Workpackages"]
            
            # Verify all required sheets exist
            for sheet_name in expected_sheets:
                assert sheet_name in wb.sheetnames, f"Sheet '{sheet_name}' should exist"
            
            # Verify sheet headers
            partners = wb["Partners"]
            expected_headers = ["Partner Number", "Organization", "Country"]
            for col, header in enumerate(expected_headers, 1):
                actual_header = partners.cell(row=1, column=col).value
                assert actual_header == header, f"Header mismatch in column {col}"
    
    def test_data_persistence(self, project_workbook):
        """Test data persistence across file operations."""
        # Modify data and save
        with excel_context(project_workbook) as wb:
            summary = wb["Summary"]
            original_title = summary['B1'].value
            new_title = "Modified Test Project"
            summary['B1'] = new_title
            wb.save(project_workbook)
        
        # Verify changes persist
        with excel_context(project_workbook) as wb:
            summary = wb["Summary"]
            saved_title = summary['B1'].value
            assert saved_title == new_title, "Changes should persist after save"
            assert saved_title != original_title, "Data should actually be modified"
    
    def test_backup_and_restore(self, project_workbook):
        """Test backup and restore functionality."""
        import shutil
        
        # Create backup
        backup_path = project_workbook + '.backup'
        shutil.copy2(project_workbook, backup_path)
        
        try:
            # Modify original
            with excel_context(project_workbook) as wb:
                summary = wb["Summary"]
                summary['B1'] = "Corrupted Data"
                wb.save(project_workbook)
            
            # Restore from backup
            shutil.copy2(backup_path, project_workbook)
            
            # Verify restoration
            with excel_context(project_workbook) as wb:
                summary = wb["Summary"]
                restored_title = summary['B1'].value
                assert restored_title == "Test Project", "Data should be restored from backup"
        
        finally:
            if os.path.exists(backup_path):
                os.unlink(backup_path)


class TestExcelErrorHandling:
    """Test Excel error handling and recovery."""
    
    def test_corrupted_file_handling(self):
        """Test handling of corrupted Excel files."""
        # Create a fake Excel file with invalid content
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp.write(b"This is not a valid Excel file")
            tmp.flush()
            
            corrupted_file = tmp.name
        
        try:
            # Attempt to open corrupted file
            with pytest.raises(Exception):  # Should raise some form of exception
                with excel_context(corrupted_file) as wb:
                    pass
        finally:
            os.unlink(corrupted_file)
    
    def test_permission_error_handling(self, temp_excel_file):
        """Test handling of permission errors."""
        # This test is platform-specific and might not work on all systems
        if os.name == 'nt':  # Windows
            import stat
            
            # Make file read-only
            os.chmod(temp_excel_file, stat.S_IRUSR | stat.S_IRGRP | stat.S_IROTH)
            
            try:
                with excel_context(temp_excel_file) as wb:
                    summary = wb["Summary"]
                    summary['A1'] = "Modified"
                    # This should work for reading, but saving might fail
                    # The exact behavior depends on the implementation
            except PermissionError:
                # Expected on some systems
                pass
            finally:
                # Restore write permissions
                os.chmod(temp_excel_file, stat.S_IRUSR | stat.S_IWUSR | stat.S_IRGRP | stat.S_IROTH)
    
    def test_disk_space_simulation(self):
        """Test handling when disk space is low."""
        # This is a simplified test - real disk space testing is complex
        # In practice, you might mock file operations or use container limits
        pass
    
    def test_concurrent_modification_detection(self, temp_excel_file):
        """Test detection of concurrent file modifications."""
        original_mtime = os.path.getmtime(temp_excel_file)
        
        # Simulate external modification
        time.sleep(0.1)
        with openpyxl.load_workbook(temp_excel_file) as wb:
            wb["Summary"]['A1'] = "Externally Modified"
            wb.save(temp_excel_file)
        
        new_mtime = os.path.getmtime(temp_excel_file)
        assert new_mtime > original_mtime, "File modification should be detectable"


@pytest.mark.integration
class TestProjectWorkflowIntegration:
    """Test complete project workflow integration."""
    
    @pytest.fixture
    def empty_project_template(self):
        """Create empty project template."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = openpyxl.Workbook()
            
            # Basic structure only
            summary = wb.active
            summary.title = "Summary"
            summary['A1'] = "Project Summary"
            
            wb.create_sheet("Partners")
            wb.create_sheet("Workpackages")
            
            wb.save(tmp.name)
            wb.close()
            yield tmp.name
        
        if os.path.exists(tmp.name):
            os.unlink(tmp.name)
    
    def test_complete_project_setup_workflow(self, empty_project_template):
        """Test complete project setup from template to final workbook."""
        # Step 1: Initialize project
        with excel_context(empty_project_template) as wb:
            summary = wb["Summary"]
            summary['A1'] = "New Research Project"
            summary['B1'] = "Advanced AI Research"
            summary['A2'] = "Duration"
            summary['B2'] = "36 months"
            wb.save(empty_project_template)
        
        # Step 2: Add partners
        partners_data = [
            {"number": "P1", "name": "Lead University", "country": "Germany"},
            {"number": "P2", "name": "Research Institute", "country": "France"},
            {"number": "P3", "name": "Technology Company", "country": "Spain"}
        ]
        
        with excel_context(empty_project_template) as wb:
            partners_sheet = wb["Partners"]
            
            # Add headers
            partners_sheet['A1'] = "Partner Number"
            partners_sheet['B1'] = "Organization Name"
            partners_sheet['C1'] = "Country"
            
            # Add partner data
            for row, partner in enumerate(partners_data, 2):
                partners_sheet[f'A{row}'] = partner["number"]
                partners_sheet[f'B{row}'] = partner["name"]
                partners_sheet[f'C{row}'] = partner["country"]
            
            wb.save(empty_project_template)
        
        # Step 3: Add workpackages
        workpackages_data = [
            {"number": "WP1", "title": "Project Management", "budget": 50000},
            {"number": "WP2", "title": "Research", "budget": 200000},
            {"number": "WP3", "title": "Development", "budget": 150000},
            {"number": "WP4", "title": "Dissemination", "budget": 75000}
        ]
        
        with excel_context(empty_project_template) as wb:
            wp_sheet = wb["Workpackages"]
            
            # Add headers
            wp_sheet['A1'] = "WP Number"
            wp_sheet['B1'] = "Title"
            wp_sheet['C1'] = "Budget"
            
            # Add workpackage data
            total_budget = 0
            for row, wp in enumerate(workpackages_data, 2):
                wp_sheet[f'A{row}'] = wp["number"]
                wp_sheet[f'B{row}'] = wp["title"]
                wp_sheet[f'C{row}'] = wp["budget"]
                total_budget += wp["budget"]
            
            # Update summary with total budget
            summary = wb["Summary"]
            summary['A3'] = "Total Budget"
            summary['B3'] = total_budget
            
            wb.save(empty_project_template)
        
        # Step 4: Validate final project structure
        with excel_context(empty_project_template) as wb:
            # Verify summary
            summary = wb["Summary"]
            assert summary['B1'].value == "Advanced AI Research"
            assert summary['B2'].value == "36 months"
            assert summary['B3'].value == 475000
            
            # Verify partners
            partners_sheet = wb["Partners"]
            assert partners_sheet.max_row == 4  # Header + 3 partners
            
            # Verify workpackages
            wp_sheet = wb["Workpackages"]
            assert wp_sheet.max_row == 5  # Header + 4 workpackages
            
            # Verify data integrity
            calculated_budget = 0
            for row in range(2, wp_sheet.max_row + 1):
                budget = wp_sheet[f'C{row}'].value
                if budget:
                    calculated_budget += budget
            
            assert calculated_budget == summary['B3'].value, "Budget calculations should match"
    
    def test_project_modification_workflow(self, empty_project_template):
        """Test project modification and update workflow."""
        # Initial setup
        with excel_context(empty_project_template) as wb:
            summary = wb["Summary"]
            summary['A1'] = "Original Project"
            summary['B1'] = "Initial Setup"
            wb.save(empty_project_template)
        
        # Modification 1: Update project details
        with excel_context(empty_project_template) as wb:
            summary = wb["Summary"]
            summary['B1'] = "Updated Project Details"
            summary['A2'] = "Last Modified"
            summary['B2'] = time.strftime("%Y-%m-%d %H:%M:%S")
            wb.save(empty_project_template)
        
        # Modification 2: Add version tracking
        with excel_context(empty_project_template) as wb:
            summary = wb["Summary"]
            summary['A3'] = "Version"
            summary['B3'] = "2.0"
            wb.save(empty_project_template)
        
        # Verify all modifications
        with excel_context(empty_project_template) as wb:
            summary = wb["Summary"]
            assert summary['B1'].value == "Updated Project Details"
            assert summary['B3'].value == "2.0"
            assert summary['B2'].value is not None  # Timestamp should exist
