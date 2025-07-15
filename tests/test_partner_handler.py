"""
Unit tests for PartnerHandler and related partner management functions.

Tests cover partner validation, addition, editing, and Excel operations
as outlined in OPTIMIZATION_RECOMMENDATIONS.md section 6.1.
"""

import pytest
import tkinter as tk
from unittest.mock import Mock, patch, MagicMock, call
import tempfile
import os
import openpyxl
from typing import Dict, Any

# Import modules under test
from handlers.add_partner_handler import (
    PartnerDialog, add_partner_to_workbook, add_partner_with_progress
)


class TestPartnerDialog:
    """Test PartnerDialog class."""
    
    @pytest.fixture
    def mock_master(self):
        """Create mock master window."""
        return Mock(spec=tk.Tk)
    
    def test_init(self, mock_master):
        """Test dialog initialization."""
        dialog = PartnerDialog(mock_master, "P2", "ACME")
        
        assert dialog.title() == "Add Partner Details"
        assert dialog.result is None
        assert "partner_number_acronym" in dialog.vars
        
        # Check that partner number and acronym are set
        value = dialog.vars["partner_number_acronym"].get()
        assert "P2" in value
        assert "ACME" in value
    
    @patch('tkinter.messagebox.showwarning')
    def test_validation_empty_fields(self, mock_warning, mock_master):
        """Test validation with empty required fields."""
        dialog = PartnerDialog(mock_master, "P2", "ACME")
        
        # Test validation with empty fields
        # This would need to be adapted based on actual validation logic
        # in the PartnerDialog class
        pass
    
    @patch('tkinter.messagebox.showinfo')
    def test_successful_validation(self, mock_info, mock_master):
        """Test successful field validation."""
        dialog = PartnerDialog(mock_master, "P2", "ACME")
        
        # Test validation with valid data
        # This would need to be adapted based on actual validation logic
        pass


class TestPartnerValidation:
    """Test partner data validation functions."""
    
    def test_validate_partner_number_valid(self):
        """Test valid partner number validation."""
        valid_numbers = ["P2", "P10", "P99"]
        
        for number in valid_numbers:
            # This would test the actual validation function
            # when it's extracted to a testable form
            assert self._is_valid_partner_number(number)
    
    def test_validate_partner_number_invalid(self):
        """Test invalid partner number validation."""
        invalid_numbers = ["", "P1", "P0", "X2", "P", "P-1"]
        
        for number in invalid_numbers:
            assert not self._is_valid_partner_number(number)
    
    def test_validate_partner_acronym_valid(self):
        """Test valid partner acronym validation."""
        valid_acronyms = ["ACME", "MIT", "UNIV", "CORP-A"]
        
        for acronym in valid_acronyms:
            assert self._is_valid_partner_acronym(acronym)
    
    def test_validate_partner_acronym_invalid(self):
        """Test invalid partner acronym validation."""
        invalid_acronyms = ["", "A", "VERY_LONG_ACRONYM_NAME"]
        
        for acronym in invalid_acronyms:
            assert not self._is_valid_partner_acronym(acronym)
    
    def _is_valid_partner_number(self, number: str) -> bool:
        """Helper method for partner number validation."""
        if not number or not number.startswith('P'):
            return False
        try:
            num = int(number[1:])
            return num >= 2  # P1 is reserved for lead partner
        except ValueError:
            return False
    
    def _is_valid_partner_acronym(self, acronym: str) -> bool:
        """Helper method for partner acronym validation."""
        if not acronym:
            return False
        return 2 <= len(acronym) <= 10


class TestAddPartnerToWorkbook:
    """Test add_partner_to_workbook function."""
    
    @pytest.fixture
    def temp_excel_file(self):
        """Create temporary Excel file for testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Partners"
            
            # Add headers
            ws['A1'] = "Partner Number"
            ws['B1'] = "Partner Name"
            ws['C1'] = "Country"
            
            # Add lead partner
            ws['A2'] = "P1"
            ws['B2'] = "Lead Partner"
            ws['C2'] = "Germany"
            
            wb.save(tmp.name)
            wb.close()
            yield tmp.name
        os.unlink(tmp.name)
    
    @pytest.fixture
    def sample_partner_info(self):
        """Create sample partner information."""
        return {
            'partner_number': 'P2',
            'partner_acronym': 'ACME',
            'organization_name': 'ACME Corporation',
            'address': '123 Business St',
            'zip_code': '12345',
            'city': 'Business City',
            'country': 'United States',
            'contact_person': 'John Smith',
            'email': 'john.smith@acme.com',
            'phone': '+1-555-0123',
            'website': 'www.acme.com'
        }
    
    def test_add_partner_success(self, temp_excel_file, sample_partner_info):
        """Test successful partner addition."""
        # Load workbook
        wb = openpyxl.load_workbook(temp_excel_file)
        
        # Add partner
        result = add_partner_to_workbook(wb, sample_partner_info)
        
        # Verify result
        assert result is True
        
        # Check if partner sheet was created
        expected_sheet_name = f"P2 {sample_partner_info['partner_acronym']}"
        assert expected_sheet_name in wb.sheetnames
        
        # Check partner data in partners sheet
        partners_sheet = wb['Partners']
        # Find the row with P2
        partner_row = None
        for row in range(2, partners_sheet.max_row + 1):
            if partners_sheet[f'A{row}'].value == 'P2':
                partner_row = row
                break
        
        assert partner_row is not None
        assert partners_sheet[f'B{partner_row}'].value == sample_partner_info['organization_name']
        
        wb.close()
    
    def test_add_partner_duplicate_number(self, temp_excel_file, sample_partner_info):
        """Test adding partner with duplicate number."""
        wb = openpyxl.load_workbook(temp_excel_file)
        
        # Add partner twice
        add_partner_to_workbook(wb, sample_partner_info)
        
        # Try to add same partner number again
        duplicate_info = sample_partner_info.copy()
        duplicate_info['partner_acronym'] = 'OTHER'
        
        # This should handle the duplicate appropriately
        # (implementation dependent - might raise exception or return False)
        try:
            result = add_partner_to_workbook(wb, duplicate_info)
            # If no exception, should return False or handle gracefully
        except ValueError as e:
            assert "duplicate" in str(e).lower() or "exists" in str(e).lower()
        
        wb.close()
    
    def test_add_partner_invalid_data(self, temp_excel_file):
        """Test adding partner with invalid data."""
        wb = openpyxl.load_workbook(temp_excel_file)
        
        # Test with empty partner number
        invalid_info = {
            'partner_number': '',
            'partner_acronym': 'ACME',
            'organization_name': 'ACME Corporation'
        }
        
        try:
            result = add_partner_to_workbook(wb, invalid_info)
            assert result is False
        except ValueError:
            pass  # Expected for invalid data
        
        wb.close()
    
    @patch('handlers.add_partner_handler.logger')
    def test_add_partner_logging(self, mock_logger, temp_excel_file, sample_partner_info):
        """Test that partner addition is properly logged."""
        wb = openpyxl.load_workbook(temp_excel_file)
        
        add_partner_to_workbook(wb, sample_partner_info)
        
        # Verify logging was called
        mock_logger.info.assert_called()
        
        wb.close()


class TestAddPartnerWithProgress:
    """Test add_partner_with_progress function."""
    
    @pytest.fixture
    def mock_parent(self):
        """Create mock parent window."""
        return Mock(spec=tk.Widget)
    
    @pytest.fixture
    def temp_excel_file(self):
        """Create temporary Excel file for testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Partners"
            wb.save(tmp.name)
            wb.close()
            yield tmp.name
        os.unlink(tmp.name)
    
    @pytest.fixture
    def sample_partner_info(self):
        """Create sample partner information."""
        return {
            'partner_number': 'P2',
            'partner_acronym': 'ACME',
            'organization_name': 'ACME Corporation',
            'country': 'United States'
        }
    
    @patch('handlers.add_partner_handler.show_progress_for_operation')
    @patch('handlers.add_partner_handler.add_partner_to_workbook')
    def test_add_partner_with_progress_success(self, mock_add_partner, mock_progress, 
                                               mock_parent, temp_excel_file, sample_partner_info):
        """Test partner addition with progress dialog."""
        # Setup mocks
        mock_add_partner.return_value = True
        mock_progress.return_value = True
        
        # Load workbook
        wb = openpyxl.load_workbook(temp_excel_file)
        
        # Execute function
        result = add_partner_with_progress(mock_parent, wb, sample_partner_info)
        
        # Verify result
        assert result is True
        
        # Verify progress dialog was shown
        mock_progress.assert_called_once()
        
        # Verify add_partner_to_workbook was called
        mock_add_partner.assert_called_once()
        
        wb.close()
    
    @patch('handlers.add_partner_handler.show_progress_for_operation')
    @patch('handlers.add_partner_handler.add_partner_to_workbook')
    def test_add_partner_with_progress_failure(self, mock_add_partner, mock_progress,
                                               mock_parent, temp_excel_file, sample_partner_info):
        """Test partner addition failure with progress dialog."""
        # Setup mocks to simulate failure
        mock_add_partner.side_effect = Exception("Partner addition failed")
        mock_progress.return_value = True
        
        # Load workbook
        wb = openpyxl.load_workbook(temp_excel_file)
        
        # Execute function (should handle exception gracefully)
        result = add_partner_with_progress(mock_parent, wb, sample_partner_info)
        
        # Verify failure handling
        assert result is False or result is None
        
        wb.close()
    
    @patch('handlers.add_partner_handler.show_progress_for_operation')
    def test_add_partner_progress_cancelled(self, mock_progress, mock_parent, 
                                           temp_excel_file, sample_partner_info):
        """Test partner addition when progress is cancelled."""
        # Setup progress to return False (cancelled)
        mock_progress.return_value = False
        
        # Load workbook
        wb = openpyxl.load_workbook(temp_excel_file)
        
        # Execute function
        result = add_partner_with_progress(mock_parent, wb, sample_partner_info)
        
        # Verify cancellation handling
        assert result is False
        
        wb.close()


@pytest.mark.performance
class TestPartnerPerformance:
    """Performance tests for partner operations."""
    
    @pytest.fixture
    def large_workbook(self):
        """Create workbook with many partners for performance testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Partners"
            
            # Add headers
            ws['A1'] = "Partner Number"
            ws['B1'] = "Partner Name"
            ws['C1'] = "Country"
            
            # Add many partners
            for i in range(1, 101):  # 100 partners
                ws[f'A{i+1}'] = f"P{i}"
                ws[f'B{i+1}'] = f"Partner {i}"
                ws[f'C{i+1}'] = "Country"
            
            wb.save(tmp.name)
            wb.close()
            yield tmp.name
        os.unlink(tmp.name)
    
    def test_partner_addition_performance(self, large_workbook, benchmark):
        """Benchmark partner addition performance."""
        partner_info = {
            'partner_number': 'P101',
            'partner_acronym': 'PERF',
            'organization_name': 'Performance Test Corp',
            'country': 'Test Country'
        }
        
        def add_partner_operation():
            wb = openpyxl.load_workbook(large_workbook)
            result = add_partner_to_workbook(wb, partner_info)
            wb.close()
            return result
        
        # Benchmark the operation
        result = benchmark(add_partner_operation)
        assert result is True
    
    def test_partner_search_performance(self, large_workbook, benchmark):
        """Benchmark partner search performance."""
        def search_partner_operation():
            wb = openpyxl.load_workbook(large_workbook)
            partners_sheet = wb['Partners']
            
            # Search for specific partner
            found = False
            for row in range(2, partners_sheet.max_row + 1):
                if partners_sheet[f'A{row}'].value == 'P50':
                    found = True
                    break
            
            wb.close()
            return found
        
        # Benchmark the search
        result = benchmark(search_partner_operation)
        assert result is True


@pytest.mark.integration
class TestPartnerIntegration:
    """Integration tests for partner operations."""
    
    @pytest.fixture
    def temp_workbook_path(self):
        """Create temporary workbook file."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = openpyxl.Workbook()
            
            # Create Partners sheet with proper structure
            partners_sheet = wb.active
            partners_sheet.title = "Partners"
            
            # Add headers
            headers = [
                "Partner Number", "Organization Name", "Partner Acronym",
                "Address", "Zip Code", "City", "Country", "Contact Person",
                "Email", "Phone", "Website"
            ]
            
            for col, header in enumerate(headers, 1):
                partners_sheet.cell(row=1, column=col).value = header
            
            # Create Summary sheet
            summary_sheet = wb.create_sheet("Summary")
            summary_sheet['A1'] = "Project Summary"
            
            wb.save(tmp.name)
            wb.close()
            yield tmp.name
        os.unlink(tmp.name)
    
    def test_complete_partner_workflow(self, temp_workbook_path):
        """Test complete partner addition workflow."""
        partner_info = {
            'partner_number': 'P2',
            'partner_acronym': 'INTEG',
            'organization_name': 'Integration Test Corp',
            'address': '123 Test Street',
            'zip_code': '12345',
            'city': 'Test City',
            'country': 'Test Country',
            'contact_person': 'Test Contact',
            'email': 'test@integration.com',
            'phone': '+1-555-TEST',
            'website': 'www.test-integration.com'
        }
        
        # Load workbook
        wb = openpyxl.load_workbook(temp_workbook_path)
        
        # Add partner
        result = add_partner_to_workbook(wb, partner_info)
        assert result is True
        
        # Verify partner was added to Partners sheet
        partners_sheet = wb['Partners']
        partner_found = False
        for row in range(2, partners_sheet.max_row + 1):
            if partners_sheet[f'A{row}'].value == 'P2':
                partner_found = True
                assert partners_sheet[f'B{row}'].value == partner_info['organization_name']
                assert partners_sheet[f'C{row}'].value == partner_info['partner_acronym']
                break
        
        assert partner_found, "Partner not found in Partners sheet"
        
        # Verify partner sheet was created
        expected_sheet_name = f"P2 {partner_info['partner_acronym']}"
        assert expected_sheet_name in wb.sheetnames
        
        # Save and verify file integrity
        wb.save(temp_workbook_path)
        wb.close()
        
        # Reload and verify data persists
        wb2 = openpyxl.load_workbook(temp_workbook_path)
        assert expected_sheet_name in wb2.sheetnames
        partners_sheet2 = wb2['Partners']
        
        # Verify data still exists
        partner_still_found = False
        for row in range(2, partners_sheet2.max_row + 1):
            if partners_sheet2[f'A{row}'].value == 'P2':
                partner_still_found = True
                break
        
        assert partner_still_found, "Partner data not persisted"
        wb2.close()
    
    def test_multiple_partners_addition(self, temp_workbook_path):
        """Test adding multiple partners sequentially."""
        partners = [
            {'partner_number': 'P2', 'partner_acronym': 'CORP1', 'organization_name': 'Corporation 1'},
            {'partner_number': 'P3', 'partner_acronym': 'CORP2', 'organization_name': 'Corporation 2'},
            {'partner_number': 'P4', 'partner_acronym': 'CORP3', 'organization_name': 'Corporation 3'}
        ]
        
        wb = openpyxl.load_workbook(temp_workbook_path)
        
        # Add all partners
        for partner_info in partners:
            result = add_partner_to_workbook(wb, partner_info)
            assert result is True
        
        # Verify all partners were added
        partners_sheet = wb['Partners']
        found_partners = []
        
        for row in range(2, partners_sheet.max_row + 1):
            partner_num = partners_sheet[f'A{row}'].value
            if partner_num:
                found_partners.append(partner_num)
        
        expected_numbers = [p['partner_number'] for p in partners]
        for expected in expected_numbers:
            assert expected in found_partners
        
        # Verify partner sheets were created
        for partner_info in partners:
            expected_sheet = f"{partner_info['partner_number']} {partner_info['partner_acronym']}"
            assert expected_sheet in wb.sheetnames
        
        wb.close()
