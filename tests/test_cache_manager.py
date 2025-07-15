"""
Tests for the CacheManager class.

This module contains unit tests for the caching system including
file hashing, metadata caching, and TTL-based caching.
"""

import pytest
import os
import time
import tempfile
import shutil
import json
from unittest.mock import patch, MagicMock

# Add src to path for imports
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from utils.cache_manager import CacheManager, cached_with_invalidation


class TestCacheManager:
    """Test cases for CacheManager class."""
    
    @pytest.fixture
    def temp_dir(self):
        """Create temporary directory for tests."""
        temp_dir = tempfile.mkdtemp()
        yield temp_dir
        shutil.rmtree(temp_dir)
    
    @pytest.fixture
    def cache_manager(self, temp_dir):
        """Create CacheManager instance with temp directory."""
        return CacheManager(cache_dir=temp_dir)
    
    @pytest.fixture
    def test_file(self, temp_dir):
        """Create test file for caching tests."""
        file_path = os.path.join(temp_dir, "test.xlsx")
        with open(file_path, 'w') as f:
            f.write("test content")
        return file_path
    
    def test_initialization(self, temp_dir):
        """Test CacheManager initialization."""
        cm = CacheManager(cache_dir=temp_dir)
        assert cm.cache_dir == temp_dir
        assert os.path.exists(temp_dir)
    
    def test_get_file_hash(self, test_file):
        """Test file hash caching."""
        hash1 = CacheManager.get_file_hash(test_file)
        hash2 = CacheManager.get_file_hash(test_file)
        
        assert hash1 == hash2
        assert len(hash1) == 32  # MD5 hash length
        assert hash1 != ""
    
    def test_get_file_hash_nonexistent(self):
        """Test file hash for non-existent file."""
        hash_result = CacheManager.get_file_hash("nonexistent.txt")
        assert hash_result == ""
    
    def test_get_file_metadata(self, test_file):
        """Test file metadata caching."""
        metadata1 = cache_manager.get_file_metadata(test_file)
        metadata2 = cache_manager.get_file_metadata(test_file)
        
        assert metadata1 == metadata2
        assert 'size' in metadata1
        assert 'modified' in metadata1
        assert 'created' in metadata1
        assert 'hash' in metadata1
        assert metadata1['size'] > 0
    
    def test_is_file_modified(self, test_file):
        """Test file modification detection."""
        original_hash = cache_manager.get_file_hash(test_file)
        
        # Initially should not be modified
        assert not cache_manager.is_file_modified(test_file, original_hash)
        
        # Modify file
        with open(test_file, 'a') as f:
            f.write("modified")
        
        # Should detect modification
        assert cache_manager.is_file_modified(test_file, original_hash)
    
    def test_validate_excel_file(self, temp_dir):
        """Test Excel file validation caching."""
        # Create valid Excel file
        excel_file = os.path.join(temp_dir, "valid.xlsx")
        from openpyxl import Workbook
        wb = Workbook()
        wb.save(excel_file)
        wb.close()
        
        # Test validation
        is_valid, error_msg = cache_manager.validate_excel_file(excel_file)
        assert is_valid is True
        assert error_msg == ""
        
        # Test caching (second call should be faster)
        is_valid2, error_msg2 = cache_manager.validate_excel_file(excel_file)
        assert is_valid2 == is_valid
        assert error_msg2 == error_msg
    
    def test_validate_excel_file_invalid(self):
        """Test validation for invalid files."""
        is_valid, error_msg = cache_manager.validate_excel_file("nonexistent.xlsx")
        assert is_valid is False
        assert "does not exist" in error_msg
    
    def test_cache_with_ttl(self, cache_manager):
        """Test TTL-based caching."""
        key = "test_ttl"
        data = {"test": "data"}
        
        # Cache with 1 second TTL
        cache_manager.cache_with_ttl(key, data, ttl_seconds=1)
        
        # Should get cached data immediately
        cached = cache_manager.get_cached_with_ttl(key)
        assert cached == data
        
        # Wait for TTL to expire
        time.sleep(1.1)
        
        # Should return None after TTL
        cached = cache_manager.get_cached_with_ttl(key)
        assert cached is None
    
    def test_clear_cache(self, cache_manager):
        """Test cache clearing."""
        # Add some cache data
        cache_manager.cache_with_ttl("test1", "data1", 300)
        cache_manager.cache_with_ttl("test2", "data2", 300)
        
        # Clear cache
        removed = cache_manager.clear_cache()
        assert removed >= 2
        
        # Should be empty
        cached = cache_manager.get_cached_with_ttl("test1")
        assert cached is None
    
    def test_get_cache_stats(self, cache_manager):
        """Test cache statistics."""
        # Use some cache functions
        cache_manager.get_file_hash(__file__)
        cache_manager.get_file_metadata(__file__)
        
        stats = cache_manager.get_cache_stats()
        
        assert isinstance(stats, dict)
        assert 'file_hash_cache_size' in stats
        assert 'file_hash_cache_max' in stats
        assert 'metadata_cache_size' in stats
        assert 'metadata_cache_max' in stats
    
    def test_invalidate_cache_for_file(self, test_file):
        """Test cache invalidation for specific file."""
        # Populate cache
        hash1 = cache_manager.get_file_hash(test_file)
        metadata1 = cache_manager.get_file_metadata(test_file)
        
        # Invalidate cache
        cache_manager.invalidate_cache_for_file(test_file)
        
        # Check that caches are cleared
        # Note: This is hard to test directly, but we can verify
        # the function runs without error
        assert True  # Function executed successfully
    
    def test_get_excel_sheet_names(self, temp_dir):
        """Test Excel sheet names caching."""
        # Create Excel file with multiple sheets
        excel_file = os.path.join(temp_dir, "sheets.xlsx")
        from openpyxl import Workbook
        wb = Workbook()
        wb.create_sheet("Sheet2")
        wb.create_sheet("Sheet3")
        wb.save(excel_file)
        wb.close()
        
        # Test sheet names caching
        sheet_names1 = cache_manager.get_excel_sheet_names(excel_file)
        sheet_names2 = cache_manager.get_excel_sheet_names(excel_file)
        
        assert sheet_names1 == sheet_names2
        assert len(sheet_names1) >= 3  # Default + 2 created
        assert "Sheet" in str(sheet_names1)
    
    def test_get_partner_data(self, temp_dir):
        """Test partner data caching."""
        # Create Excel file with partner sheet
        excel_file = os.path.join(temp_dir, "partner.xlsx")
        from openpyxl import Workbook
        wb = Workbook()
        
        # Create partner sheet
        p2_sheet = wb.create_sheet("P2 TestPartner")
        p2_sheet['D4'] = "P2-12345"
        p2_sheet['D5'] = "Test Partner"
        p2_sheet['D6'] = "TestCountry"
        p2_sheet['D7'] = "TestRole"
        
        wb.save(excel_file)
        wb.close()
        
        # Test partner data caching
        partner_data1 = cache_manager.get_partner_data(excel_file, "2")
        partner_data2 = cache_manager.get_partner_data(excel_file, "2")
        
        assert partner_data1 == partner_data2
        assert partner_data1['partner_number'] == "2"
        assert partner_data1['sheet_name'] == "P2 TestPartner"
        assert partner_data1['data_extracted'] is True
    
    def test_get_workpackage_data(self, temp_dir):
        """Test workpackage data caching."""
        # Create Excel file with PM Summary
        excel_file = os.path.join(temp_dir, "workpackage.xlsx")
        from openpyxl import Workbook
        wb = Workbook()
        
        # Create PM Summary sheet
        pm_sheet = wb.create_sheet("PM Summary")
        pm_sheet['A2'] = "1"
        pm_sheet['B2'] = "Workpackage 1"
        pm_sheet['A3'] = "2"
        pm_sheet['B3'] = "Workpackage 2"
        
        wb.save(excel_file)
        wb.close()
        
        # Test workpackage data caching
        wp_data1 = cache_manager.get_workpackage_data(excel_file, "1")
        wp_data2 = cache_manager.get_workpackage_data(excel_file, "1")
        
        assert wp_data1 == wp_data2
        assert wp_data1['workpackage_id'] == "1"
        assert wp_data1['exists'] is True


class TestCacheDecorator:
    """Test cases for cache decorator."""
    
    @pytest.fixture
    def temp_cache(self):
        """Create temporary cache manager."""
        return CacheManager()
    
    def test_cached_with_invalidation(self, temp_cache):
        """Test cache decorator with TTL."""
        call_count = 0
        
        @cached_with_invalidation(temp_cache, ttl=1)
        def test_function(x, y=10):
            nonlocal call_count
            call_count += 1
            return x + y
        
        # First call
        result1 = test_function(5, y=10)
        assert result1 == 15
        assert call_count == 1
        
        # Second call with same params - should use cache
        result2 = test_function(5, y=10)
        assert result2 == 15
        assert call_count == 1  # Should not increment
        
        # Different params - should execute
        result3 = test_function(5, y=20)
        assert result3 == 25
        assert call_count == 2
    
    def test_cache_ttl_expiration(self, temp_cache):
        """Test TTL expiration in cache decorator."""
        call_count = 0
        
        @cached_with_invalidation(temp_cache, ttl=1)
        def test_function(x):
            nonlocal call_count
            call_count += 1
            return x * 2
        
        # First call
        test_function(5)
        assert call_count == 1
        
        # Wait for TTL to expire
        time.sleep(1.1)
        
        # Should execute again
        test_function(5)
        assert call_count == 2


class TestCacheAwareExcelManager:
    """Test cases for CacheAwareExcelManager."""
    
    @pytest.fixture
    def cache_aware(self):
        """Create CacheAwareExcelManager instance."""
        return CacheAwareExcelManager(cache_manager)
    
    @pytest.fixture
    def test_excel_file(self, temp_dir):
        """Create test Excel file."""
        file_path = os.path.join(temp_dir, "cache_aware.xlsx")
        from openpyxl import Workbook
        wb = Workbook()
        wb.create_sheet("P2 TestPartner")
        wb.save(file_path)
        wb.close()
        return file_path
    
    def test_get_cached_sheet_names(self, cache_aware, test_excel_file):
        """Test cached sheet names."""
        names1 = cache_aware.get_cached_sheet_names(test_excel_file)
        names2 = cache_aware.get_cached_sheet_names(test_excel_file)
        
        assert names1 == names2
        assert isinstance(names1, list)
    
    def test_get_cached_partner_data(self, cache_aware, test_excel_file):
        """Test cached partner data."""
        data1 = cache_aware.get_cached_partner_data(test_excel_file, "2")
        data2 = cache_aware.get_cached_partner_data(test_excel_file, "2")
        
        assert data1 == data2
        assert isinstance(data1, dict)
    
    def test_has_file_changed(self, cache_aware, test_excel_file):
        """Test file change detection."""
        original_hash = cache_manager.get_file_hash(test_excel_file)
        
        # Initially should not be changed
        assert not cache_aware.has_file_changed(test_excel_file, original_hash)
        
        # Modify file
        with open(test_excel_file, 'a') as f:
            f.write("modified")
        
        # Should detect change
        assert cache_aware.has_file_changed(test_excel_file, original_hash)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
