"""
Shared test configuration and fixtures for ProjectBudgetinator tests.

This module provides common fixtures, utilities, and configuration
for all test modules as outlined in OPTIMIZATION_RECOMMENDATIONS.md.
"""

import pytest
import tempfile
import os
import openpyxl
from pathlib import Path
import tkinter as tk
from unittest.mock import Mock
import sys
import time
import shutil

# Add src directory to path for imports
src_path = Path(__file__).parent.parent / "src"
sys.path.insert(0, str(src_path))


@pytest.fixture(scope="session")
def test_data_directory():
    """Create temporary directory for test data."""
    temp_dir = tempfile.mkdtemp(prefix="projectbudgetinator_tests_")
    yield temp_dir
    shutil.rmtree(temp_dir, ignore_errors=True)


@pytest.fixture
def mock_tk_root():
    """Create mock Tkinter root window."""
    try:
        root = tk.Tk()
        root.withdraw()  # Hide the window
        yield root
        root.destroy()
    except tk.TclError:
        # If no display available, use mock
        yield Mock(spec=tk.Tk)


@pytest.fixture
def mock_parent_window():
    """Create mock parent window for dialogs."""
    return Mock(spec=tk.Widget)


@pytest.fixture
def simple_excel_file(test_data_directory):
    """Create a simple Excel file for testing."""
    file_path = os.path.join(test_data_directory, "simple_test.xlsx")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Sheet"
    ws['A1'] = "Test Data"
    ws['B1'] = "Value"
    
    wb.save(file_path)
    wb.close()
    
    yield file_path
    
    if os.path.exists(file_path):
        os.unlink(file_path)


@pytest.fixture
def project_excel_template(test_data_directory):
    """Create a project Excel template for testing."""
    file_path = os.path.join(test_data_directory, "project_template.xlsx")
    
    wb = openpyxl.Workbook()
    
    # Summary sheet
    summary = wb.active
    summary.title = "Summary"
    summary['A1'] = "Project Title"
    summary['B1'] = "Test Project"
    summary['A2'] = "Duration"
    summary['B2'] = "36 months"
    summary['A3'] = "Total Budget"
    summary['B3'] = "500000"
    
    # Partners sheet
    partners = wb.create_sheet("Partners")
    partners['A1'] = "Partner Number"
    partners['B1'] = "Organization Name"
    partners['C1'] = "Country"
    partners['D1'] = "Contact Email"
    
    # Add lead partner
    partners['A2'] = "P1"
    partners['B2'] = "Lead Organization"
    partners['C2'] = "Germany"
    partners['D2'] = "lead@example.com"
    
    # Workpackages sheet
    workpackages = wb.create_sheet("Workpackages")
    workpackages['A1'] = "WP Number"
    workpackages['B1'] = "Title"
    workpackages['C1'] = "Lead Partner"
    workpackages['D1'] = "Budget"
    workpackages['E1'] = "Start Month"
    workpackages['F1'] = "End Month"
    
    wb.save(file_path)
    wb.close()
    
    yield file_path
    
    if os.path.exists(file_path):
        os.unlink(file_path)


@pytest.fixture
def large_excel_file(test_data_directory):
    """Create large Excel file for performance testing."""
    file_path = os.path.join(test_data_directory, "large_test.xlsx")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Large Data"
    
    # Create headers
    for col in range(1, 21):  # 20 columns
        ws.cell(row=1, column=col).value = f"Column_{col}"
    
    # Create data rows
    for row in range(2, 1002):  # 1000 data rows
        for col in range(1, 21):
            ws.cell(row=row, column=col).value = f"Data_{row}_{col}"
    
    wb.save(file_path)
    wb.close()
    
    yield file_path
    
    if os.path.exists(file_path):
        os.unlink(file_path)


@pytest.fixture
def sample_partner_data():
    """Provide sample partner data for testing."""
    return {
        'partner_number': 'P2',
        'partner_acronym': 'TEST',
        'organization_name': 'Test Organization',
        'address': '123 Test Street',
        'zip_code': '12345',
        'city': 'Test City',
        'country': 'Test Country',
        'contact_person': 'Test Contact',
        'email': 'test@example.com',
        'phone': '+1-555-0123',
        'website': 'www.test.com'
    }


@pytest.fixture
def sample_workpackage_data():
    """Provide sample workpackage data for testing."""
    return {
        'workpackage_number': 'WP1',
        'title': 'Test Workpackage',
        'lead_partner': 'P1',
        'budget': '50000.00',
        'start_month': '1',
        'end_month': '12',
        'description': 'Test workpackage description'
    }


@pytest.fixture
def mock_logger():
    """Create mock logger for testing."""
    return Mock()


@pytest.fixture
def mock_progress_dialog():
    """Create mock progress dialog for testing."""
    dialog = Mock()
    dialog.update_progress.return_value = None
    dialog.close.return_value = None
    dialog.cancelled = False
    return dialog


class TestUtilities:
    """Utility functions for tests."""
    
    @staticmethod
    def create_test_excel_file(file_path: str, sheets_data: dict):
        """
        Create Excel file with specified data.
        
        Args:
            file_path: Path for the Excel file
            sheets_data: Dict with sheet names as keys and data as values
        """
        wb = openpyxl.Workbook()
        
        # Remove default sheet if we're adding custom sheets
        if "Sheet" in sheets_data and len(sheets_data) > 1:
            wb.remove(wb.active)
        
        for sheet_name, data in sheets_data.items():
            if sheet_name == "Sheet":
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(sheet_name)
            
            # Add data
            for row_idx, row_data in enumerate(data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx).value = value
        
        wb.save(file_path)
        wb.close()
    
    @staticmethod
    def verify_excel_data(file_path: str, sheet_name: str, expected_data: list):
        """
        Verify data in Excel file matches expected data.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet to check
            expected_data: Expected data as list of lists
        
        Returns:
            bool: True if data matches
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]
        
        for row_idx, expected_row in enumerate(expected_data, 1):
            for col_idx, expected_value in enumerate(expected_row, 1):
                actual_value = ws.cell(row=row_idx, column=col_idx).value
                if actual_value != expected_value:
                    wb.close()
                    return False
        
        wb.close()
        return True
    
    @staticmethod
    def get_file_modification_time(file_path: str) -> float:
        """Get file modification time."""
        return os.path.getmtime(file_path)
    
    @staticmethod
    def wait_for_file_change(file_path: str, original_time: float, timeout: float = 5.0) -> bool:
        """
        Wait for file modification time to change.
        
        Args:
            file_path: Path to file
            original_time: Original modification time
            timeout: Maximum time to wait
            
        Returns:
            bool: True if file was modified within timeout
        """
        start_time = time.time()
        
        while time.time() - start_time < timeout:
            current_time = os.path.getmtime(file_path)
            if current_time > original_time:
                return True
            time.sleep(0.1)
        
        return False


@pytest.fixture
def test_utilities():
    """Provide test utilities."""
    return TestUtilities


# Performance test markers
def pytest_configure(config):
    """Configure pytest with custom markers."""
    config.addinivalue_line(
        "markers", "unit: marks tests as unit tests"
    )
    config.addinivalue_line(
        "markers", "integration: marks tests as integration tests"
    )
    config.addinivalue_line(
        "markers", "performance: marks tests as performance tests"
    )
    config.addinivalue_line(
        "markers", "slow: marks tests as slow running"
    )


# Performance test configuration
@pytest.fixture
def benchmark_config():
    """Configuration for benchmark tests."""
    return {
        'min_rounds': 3,
        'max_time': 10.0,
        'min_time': 0.1,
        'warmup': True,
        'warmup_iterations': 1
    }


# Test data generators
@pytest.fixture
def generate_partner_list():
    """Generate list of test partners."""
    def _generate(count: int = 5):
        partners = []
        for i in range(2, count + 2):  # Start from P2
            partners.append({
                'partner_number': f'P{i}',
                'partner_acronym': f'ORG{i}',
                'organization_name': f'Organization {i}',
                'country': f'Country {i}',
                'email': f'contact{i}@example.com'
            })
        return partners
    return _generate


@pytest.fixture
def generate_workpackage_list():
    """Generate list of test workpackages."""
    def _generate(count: int = 5):
        workpackages = []
        for i in range(1, count + 1):
            workpackages.append({
                'workpackage_number': f'WP{i}',
                'title': f'Workpackage {i}',
                'lead_partner': 'P1',
                'budget': str(10000 * i),
                'start_month': str(i),
                'end_month': str(i + 12)
            })
        return workpackages
    return _generate


# Error simulation fixtures
@pytest.fixture
def mock_file_error():
    """Mock file operation errors."""
    class FileErrorSimulator:
        def __init__(self):
            self.should_fail = False
            self.error_type = Exception
            self.error_message = "Simulated file error"
        
        def set_error(self, error_type: type, message: str):
            self.should_fail = True
            self.error_type = error_type
            self.error_message = message
        
        def clear_error(self):
            self.should_fail = False
        
        def check_and_raise(self):
            if self.should_fail:
                raise self.error_type(self.error_message)
    
    return FileErrorSimulator()


# Memory monitoring fixture
@pytest.fixture
def memory_monitor():
    """Monitor memory usage during tests."""
    import psutil
    import os
    
    class MemoryMonitor:
        def __init__(self):
            self.process = psutil.Process(os.getpid())
            self.initial_memory = self.process.memory_info().rss
            self.peak_memory = self.initial_memory
        
        def update_peak(self):
            current_memory = self.process.memory_info().rss
            self.peak_memory = max(self.peak_memory, current_memory)
        
        def get_memory_usage(self):
            current_memory = self.process.memory_info().rss
            return {
                'initial': self.initial_memory,
                'current': current_memory,
                'peak': self.peak_memory,
                'increase': current_memory - self.initial_memory,
                'peak_increase': self.peak_memory - self.initial_memory
            }
    
    return MemoryMonitor()


# Clean up after tests
@pytest.fixture(autouse=True)
def cleanup_test_files():
    """Automatically clean up test files after each test."""
    yield
    
    # Clean up any temporary files that might have been created
    import glob
    import tempfile
    
    temp_dir = tempfile.gettempdir()
    test_files = glob.glob(os.path.join(temp_dir, "projectbudgetinator_test_*"))
    
    for file_path in test_files:
        try:
            if os.path.isfile(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path, ignore_errors=True)
        except Exception:
            pass  # Ignore cleanup errors
