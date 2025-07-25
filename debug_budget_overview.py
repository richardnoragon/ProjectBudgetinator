#!/usr/bin/env python3
"""
Debug script for Budget Overview Handler.

This script helps troubleshoot issues with the Budget Overview update process
by providing detailed logging of source cells, target cells, and values.
"""

import sys
import os
from pathlib import Path

# Add the src directory to the Python path
src_path = Path(__file__).parent / "src"
sys.path.insert(0, str(src_path))

from openpyxl import load_workbook
from handlers.update_budget_overview_handler import (
    UpdateBudgetOverviewHandler,
    BUDGET_OVERVIEW_CELL_MAPPINGS,
    get_budget_overview_row,
    get_partner_number_from_sheet_name
)


def debug_workbook_structure(workbook_path: str):
    """Debug the workbook structure and available worksheets."""
    print("🔍 DEBUGGING WORKBOOK STRUCTURE")
    print("=" * 80)
    
    try:
        workbook = load_workbook(workbook_path, data_only=True)
        print(f"📁 Workbook: {workbook_path}")
        print(f"📊 Total worksheets: {len(workbook.sheetnames)}")
        print()
        
        print("📋 Available worksheets:")
        for i, sheet_name in enumerate(workbook.sheetnames, 1):
            partner_num = get_partner_number_from_sheet_name(sheet_name)
            if partner_num:
                target_row = get_budget_overview_row(partner_num)
                print(f"  {i:2d}. {sheet_name:<20} → Partner {partner_num} → Budget Row {target_row}")
            else:
                print(f"  {i:2d}. {sheet_name:<20} → (Not a partner sheet)")
        
        print()
        return workbook
        
    except Exception as e:
        print(f"❌ ERROR loading workbook: {e}")
        return None


def debug_cell_mappings():
    """Debug the cell mapping configuration."""
    print("🗺️  DEBUGGING CELL MAPPINGS")
    print("=" * 80)
    
    print("📍 Cell Mapping Configuration:")
    print("   SOURCE → TARGET")
    print("   " + "-" * 20)
    
    for source_cell, target_col in BUDGET_OVERVIEW_CELL_MAPPINGS.items():
        print(f"   {source_cell:>6} → Column {target_col}")
    
    print(f"\n📊 Total mappings: {len(BUDGET_OVERVIEW_CELL_MAPPINGS)}")
    print()


def debug_partner_data_extraction(workbook, partner_number: int):
    """Debug data extraction from a specific partner worksheet."""
    print(f"🔬 DEBUGGING PARTNER {partner_number} DATA EXTRACTION")
    print("=" * 80)
    
    try:
        # Find partner worksheet
        partner_sheet_name = None
        for sheet_name in workbook.sheetnames:
            if get_partner_number_from_sheet_name(sheet_name) == partner_number:
                partner_sheet_name = sheet_name
                break
        
        if not partner_sheet_name:
            print(f"❌ Partner {partner_number} worksheet not found!")
            return None
        
        print(f"📋 Partner worksheet: {partner_sheet_name}")
        partner_ws = workbook[partner_sheet_name]
        
        print("\n📍 Cell Values:")
        print("   SOURCE | VALUE                    | TYPE")
        print("   " + "-" * 45)
        
        extracted_data = {}
        for source_cell, target_col in BUDGET_OVERVIEW_CELL_MAPPINGS.items():
            try:
                cell = partner_ws[source_cell]
                value = cell.value
                value_type = type(value).__name__
                display_value = str(value) if value is not None else "None"
                
                print(f"   {source_cell:>6} | {display_value:<20} | {value_type}")
                
                extracted_data[source_cell] = {
                    'value': value,
                    'target_col': target_col,
                    'formatted_value': value if value is not None else ''
                }
                
            except Exception as e:
                print(f"   {source_cell:>6} | ERROR: {str(e):<15} | ERROR")
                extracted_data[source_cell] = {
                    'value': None,
                    'target_col': target_col,
                    'formatted_value': '',
                    'error': str(e)
                }
        
        print(f"\n✅ Extracted {len(extracted_data)} cell values")
        return extracted_data
        
    except Exception as e:
        print(f"❌ ERROR extracting partner {partner_number} data: {e}")
        return None


def debug_budget_overview_update(workbook, partner_number: int, extracted_data: dict):
    """Debug the Budget Overview update process."""
    print(f"🎯 DEBUGGING BUDGET OVERVIEW UPDATE FOR PARTNER {partner_number}")
    print("=" * 80)
    
    try:
        if "Budget Overview" not in workbook.sheetnames:
            print("❌ Budget Overview worksheet not found!")
            return False
        
        budget_ws = workbook["Budget Overview"]
        target_row = get_budget_overview_row(partner_number)
        
        print("📋 Budget Overview worksheet found")
        print(f"🎯 Target row: {target_row}")
        print()
        
        print("📝 Update Operations:")
        print("   SOURCE → TARGET     | VALUE                    | STATUS")
        print("   " + "-" * 60)
        
        success_count = 0
        for source_cell, data in extracted_data.items():
            target_col = data['target_col']
            value = data['formatted_value']
            target_cell = f"{target_col}{target_row}"
            
            try:
                # Show the update operation
                display_value = str(value) if value != '' else "EMPTY"
                status = "✅ READY" if 'error' not in data else "❌ ERROR"
                
                print(f"   {source_cell:>6} → {target_cell:<8} | {display_value:<20} | {status}")
                
                if 'error' not in data:
                    success_count += 1
                    
            except Exception as e:
                print(f"   {source_cell:>6} → {target_cell:<8} | ERROR: {str(e):<15} | ❌ FAIL")
        
        print(f"\n📊 Summary: {success_count}/{len(extracted_data)} operations ready")
        return success_count == len(extracted_data)
        
    except Exception as e:
        print(f"❌ ERROR during Budget Overview update debug: {e}")
        return False


def main():
    """Main debug function."""
    print("🐛 BUDGET OVERVIEW HANDLER DEBUG TOOL")
    print("=" * 80)
    
    # Get workbook path from command line or prompt
    if len(sys.argv) > 1:
        workbook_path = sys.argv[1]
    else:
        workbook_path = input("Enter path to Excel workbook: ").strip()
    
    if not os.path.exists(workbook_path):
        print(f"❌ File not found: {workbook_path}")
        return
    
    # Debug workbook structure
    workbook = debug_workbook_structure(workbook_path)
    if not workbook:
        return
    
    print()
    
    # Debug cell mappings
    debug_cell_mappings()
    
    # Get partner number to debug
    try:
        partner_input = input("Enter partner number to debug (e.g., 2, 3, 4): ").strip()
        partner_number = int(partner_input)
    except ValueError:
        print("❌ Invalid partner number")
        return
    
    print()
    
    # Debug partner data extraction
    extracted_data = debug_partner_data_extraction(workbook, partner_number)
    if not extracted_data:
        return
    
    print()
    
    # Debug Budget Overview update
    debug_budget_overview_update(workbook, partner_number, extracted_data)
    
    print()
    print("🎉 Debug complete! Check the output above for any issues.")


if __name__ == "__main__":
    main()