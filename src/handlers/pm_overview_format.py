"""
PM Overview conditional formatting module for ProjectBudgetinator.

This module applies conditional formatting to PM Overview worksheets based on
row completion status: complete rows (green), partial rows (blue/gray), and empty rows (red).
"""

import tkinter as tk
from tkinter import messagebox
from typing import Dict, Any, List, Optional, Tuple, NamedTuple
import traceback
import logging
from dataclasses import dataclass
from enum import Enum

# Import custom exceptions
from ..exceptions import (
    BudgetError,
    BudgetFormatError,
    BudgetDataError,
    WorksheetAccessError,
    CellAccessError,
    StyleApplicationError,
    handle_budget_exception
)

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.styles import PatternFill, Font, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None
    Worksheet = None
    PatternFill = None
    Font = None
    Border = None
    Side = None

# Set up logging for this module
logger = logging.getLogger(__name__)


class RowStatus(Enum):
    """Enumeration for row completion status."""
    COMPLETE = "complete"
    PARTIAL = "partial"
    EMPTY = "empty"


@dataclass
class RowAnalysis:
    """Data class for row analysis results."""
    row_number: int
    partner_number: int
    status: RowStatus
    filled_cells: List[str]
    empty_cells: List[str]
    total_cells: int
    
    @property
    def filled_count(self) -> int:
        return len(self.filled_cells)
    
    @property
    def empty_count(self) -> int:
        return len(self.empty_cells)


class StyleDefinitions:
    """Style definitions for different row completion states."""
    
    # Complete rows (all cells filled)
    COMPLETE_STYLE = {
        'fill': PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid') if PatternFill else None,
        'font': Font(color='2E7D32', bold=True) if Font else None,
        'border': Border(
            left=Side(style='thin', color='4CAF50'),
            right=Side(style='thin', color='4CAF50'),
            top=Side(style='thin', color='4CAF50'),
            bottom=Side(style='thin', color='4CAF50')
        ) if Border and Side else None
    }
    
    # Partial rows - filled cells
    PARTIAL_FILLED_STYLE = {
        'fill': PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid') if PatternFill else None,
        'font': Font(color='1565C0') if Font else None,
        'border': Border(
            left=Side(style='thin', color='2196F3'),
            right=Side(style='thin', color='2196F3'),
            top=Side(style='thin', color='2196F3'),
            bottom=Side(style='thin', color='2196F3')
        ) if Border and Side else None
    }
    
    # Partial rows - empty cells
    PARTIAL_EMPTY_STYLE = {
        'fill': PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid') if PatternFill else None,
        'font': Font(color='757575', italic=True) if Font else None,
        'border': Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        ) if Border and Side else None
    }
    
    # Empty rows (no cells filled)
    EMPTY_STYLE = {
        'fill': PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid') if PatternFill else None,
        'font': Font(color='C62828') if Font else None,
        'border': Border(
            left=Side(style='thin', color='FFCDD2'),
            right=Side(style='thin', color='FFCDD2'),
            top=Side(style='thin', color='FFCDD2'),
            bottom=Side(style='thin', color='FFCDD2')
        ) if Border and Side else None
    }


class RowAnalyzer:
    """Analyzer for PM Overview row completion status."""
    
    # PM Overview columns to analyze (C through Q for WP1-WP15)
    ANALYSIS_COLUMNS = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
                        'N', 'O', 'P', 'Q']
    
    @staticmethod
    def get_pm_overview_row(partner_number: int) -> int:
        """Calculate PM Overview row number from partner number."""
        return partner_number + 4  # P2->Row6, P3->Row7, etc.
    
    @staticmethod
    def get_partner_number_from_row(row_number: int) -> int:
        """Calculate partner number from PM Overview row number."""
        return row_number - 4  # Row6->P2, Row7->P3, etc.
    
    @staticmethod
    def analyze_row(worksheet, row_number: int) -> RowAnalysis:
        """Analyze a single row for completion status."""
        partner_number = RowAnalyzer.get_partner_number_from_row(row_number)
        filled_cells = []
        empty_cells = []
        
        for col in RowAnalyzer.ANALYSIS_COLUMNS:
            cell_ref = f"{col}{row_number}"
            try:
                cell = worksheet[cell_ref]
                cell_value = cell.value
                
                # Check if cell has meaningful content
                if cell_value is not None and str(cell_value).strip():
                    filled_cells.append(cell_ref)
                else:
                    empty_cells.append(cell_ref)
                    
            except AttributeError as e:
                logger.warning(f"Worksheet missing attribute for cell {cell_ref}: {e}")
                empty_cells.append(cell_ref)
            except KeyError as e:
                logger.warning(f"Invalid cell reference {cell_ref}: {e}")
                empty_cells.append(cell_ref)
            except ValueError as e:
                logger.warning(f"Invalid cell value for {cell_ref}: {e}")
                empty_cells.append(cell_ref)
            except Exception as e:
                logger.error(f"Unexpected error accessing cell {cell_ref}: {e}")
                empty_cells.append(cell_ref)
        
        # Determine status
        if len(filled_cells) == len(RowAnalyzer.ANALYSIS_COLUMNS):
            status = RowStatus.COMPLETE
        elif len(filled_cells) == 0:
            status = RowStatus.EMPTY
        else:
            status = RowStatus.PARTIAL
        
        return RowAnalysis(
            row_number=row_number,
            partner_number=partner_number,
            status=status,
            filled_cells=filled_cells,
            empty_cells=empty_cells,
            total_cells=len(RowAnalyzer.ANALYSIS_COLUMNS)
        )


class PMOverviewFormatter:
    """Main formatter class for applying conditional formatting to PM Overview."""
    
    def __init__(self, parent_window: Optional[tk.Widget] = None):
        """Initialize the formatter with optional parent window for debug display."""
        self.parent = parent_window
        self.pm_overview_sheet_name = "PM Overview"
        self.debug_window = None
    
    def show_debug_window(self, title: str, content: str):
        """Show debug information in a window."""
        if self.debug_window:
            self.debug_window.destroy()
        
        # Use shared root pattern to avoid multiple Tk() instances
        if self.parent:
            self.debug_window = tk.Toplevel(self.parent)
        else:
            # Import shared root utility
            from ..utils.window_positioning import ScreenInfo
            if not ScreenInfo._shared_root:
                ScreenInfo._shared_root = tk.Tk()
                ScreenInfo._shared_root.withdraw()  # Hide the main root
            self.debug_window = tk.Toplevel(ScreenInfo._shared_root)
        self.debug_window.title(title)
        self.debug_window.geometry("1200x800")
        
        # Create text widget with scrollbar
        frame = tk.Frame(self.debug_window)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        text_widget = tk.Text(frame, wrap=tk.WORD, font=("Consolas", 9))
        scrollbar = tk.Scrollbar(frame, orient=tk.VERTICAL, command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        text_widget.insert(tk.END, content)
        text_widget.config(state=tk.DISABLED)
        
        # Close button
        close_btn = tk.Button(self.debug_window, text="Close",
                              command=self.debug_window.destroy)
        close_btn.pack(pady=5)
    
    def get_partner_rows(self, worksheet) -> List[int]:
        """Get list of partner row numbers in PM Overview (rows 6-24 for partners 2-20)."""
        partner_rows = []
        
        # Check rows 6-24 (partners 2-20)
        for row_num in range(6, 25):
            try:
                # Check if this row has any partner identifier (column C typically has WP1 data)
                partner_cell = worksheet[f'C{row_num}']
                if partner_cell.value is not None:
                    partner_rows.append(row_num)
            except AttributeError as e:
                logger.warning(f"Worksheet missing attribute for row {row_num}: {e}")
                continue
            except KeyError as e:
                logger.warning(f"Invalid cell reference C{row_num}: {e}")
                continue
            except Exception as e:
                logger.error(f"Unexpected error checking partner row {row_num}: {e}")
                continue
        
        return partner_rows
    
    def analyze_all_rows(self, worksheet) -> Dict[int, RowAnalysis]:
        """Analyze all partner rows in the PM Overview."""
        partner_rows = self.get_partner_rows(worksheet)
        analysis_results = {}
        
        for row_num in partner_rows:
            try:
                analysis = RowAnalyzer.analyze_row(worksheet, row_num)
                analysis_results[row_num] = analysis
            except BudgetError as e:
                logger.error(f"Budget error analyzing row {row_num}: {e}")
                continue
            except AttributeError as e:
                logger.warning(f"Missing worksheet attribute for row {row_num}: {e}")
                continue
            except ValueError as e:
                logger.warning(f"Invalid data in row {row_num}: {e}")
                continue
            except Exception as e:
                logger.error(f"Unexpected error analyzing row {row_num}: {e}")
                continue
        
        return analysis_results
    
    def generate_debug_report(self, analysis_results: Dict[int, RowAnalysis]) -> str:
        """Generate comprehensive debug report for styling analysis."""
        debug_lines = []
        
        # Header
        debug_lines.append("🎨 PM OVERVIEW STYLING ANALYSIS")
        debug_lines.append("=" * 80)
        debug_lines.append("")
        
        # Summary statistics
        complete_count = sum(1 for a in analysis_results.values() if a.status == RowStatus.COMPLETE)
        partial_count = sum(1 for a in analysis_results.values() if a.status == RowStatus.PARTIAL)
        empty_count = sum(1 for a in analysis_results.values() if a.status == RowStatus.EMPTY)
        
        debug_lines.append("📊 Row Analysis Summary:")
        debug_lines.append(f"   Total Rows Analyzed: {len(analysis_results)}")
        debug_lines.append(f"   Complete Rows: {complete_count}")
        debug_lines.append(f"   Partial Rows: {partial_count}")
        debug_lines.append(f"   Empty Rows: {empty_count}")
        debug_lines.append("")
        
        # Detailed row analysis
        debug_lines.append("📋 Detailed Row Analysis:")
        debug_lines.append("=" * 80)
        debug_lines.append("ROW | PARTNER | STATUS   | FILLED | EMPTY | STYLE TO APPLY")
        debug_lines.append("-" * 80)
        
        for row_num in sorted(analysis_results.keys()):
            analysis = analysis_results[row_num]
            
            # Status emoji and description
            if analysis.status == RowStatus.COMPLETE:
                status_display = "COMPLETE"
                style_desc = "✅ Complete Style (Green)"
            elif analysis.status == RowStatus.PARTIAL:
                status_display = "PARTIAL "
                style_desc = "🔵 Partial Style (Blue/Gray)"
            else:
                status_display = "EMPTY   "
                style_desc = "❌ Empty Style (Red)"
            
            filled_ratio = f"{analysis.filled_count}/{analysis.total_cells}"
            empty_ratio = f"{analysis.empty_count}/{analysis.total_cells}"
            
            debug_lines.append(
                f"{row_num:>3} | P{analysis.partner_number:<6} | {status_display} | {filled_ratio:>6} | {empty_ratio:>5} | {style_desc}"
            )
        
        debug_lines.append("-" * 80)
        debug_lines.append("")
        
        # Style application preview
        debug_lines.append("🎨 Style Application Preview:")
        debug_lines.append("=" * 80)
        
        # Group by status for cleaner display
        complete_rows = [a for a in analysis_results.values() if a.status == RowStatus.COMPLETE]
        partial_rows = [a for a in analysis_results.values() if a.status == RowStatus.PARTIAL]
        empty_rows = [a for a in analysis_results.values() if a.status == RowStatus.EMPTY]
        
        if complete_rows:
            debug_lines.append(f"Complete Rows ({len(complete_rows)}):")
            for analysis in sorted(complete_rows, key=lambda x: x.row_number):
                debug_lines.append(f"   Row {analysis.row_number} (P{analysis.partner_number}): Background=#E8F5E8, Text=#2E7D32, Font=Bold")
            debug_lines.append("")
        
        if partial_rows:
            debug_lines.append(f"Partial Rows ({len(partial_rows)}):")
            for analysis in sorted(partial_rows, key=lambda x: x.row_number):
                debug_lines.append(f"   Row {analysis.row_number} (P{analysis.partner_number}): Filled cells=#E3F2FD, Empty cells=#F5F5F5")
            debug_lines.append("")
        
        if empty_rows:
            debug_lines.append(f"Empty Rows ({len(empty_rows)}):")
            for analysis in sorted(empty_rows, key=lambda x: x.row_number):
                debug_lines.append(f"   Row {analysis.row_number} (P{analysis.partner_number}): Background=#FFEBEE, Text=#C62828")
            debug_lines.append("")
        
        # Cell-level details for partial rows
        if partial_rows:
            debug_lines.append("🔍 Partial Row Cell Details:")
            debug_lines.append("=" * 80)
            for analysis in sorted(partial_rows, key=lambda x: x.row_number)[:3]:  # Show first 3 for brevity
                debug_lines.append(f"Row {analysis.row_number} (P{analysis.partner_number}):")
                debug_lines.append(f"   Filled: {', '.join(analysis.filled_cells[:10])}{'...' if len(analysis.filled_cells) > 10 else ''}")
                debug_lines.append(f"   Empty:  {', '.join(analysis.empty_cells[:10])}{'...' if len(analysis.empty_cells) > 10 else ''}")
                debug_lines.append("")
        
        return "\n".join(debug_lines)
    
    def apply_cell_style(self, worksheet, cell_ref: str, style_dict: Dict[str, Any]):
        """Apply style to a single cell."""
        try:
            cell = worksheet[cell_ref]
            
            if style_dict.get('fill') and style_dict['fill'] is not None:
                cell.fill = style_dict['fill']
            
            if style_dict.get('font') and style_dict['font'] is not None:
                cell.font = style_dict['font']
            
            if style_dict.get('border') and style_dict['border'] is not None:
                cell.border = style_dict['border']
                
        except AttributeError as e:
            logger.warning(f"Cell {cell_ref} missing style attribute: {e}")
            raise StyleApplicationError(
                f"Failed to apply style to cell {cell_ref}: missing attribute",
                cell_ref=cell_ref,
                style_type="general"
            ) from e
        except KeyError as e:
            logger.warning(f"Invalid cell reference {cell_ref}: {e}")
            raise CellAccessError(
                f"Cannot access cell {cell_ref}",
                cell_ref=cell_ref
            ) from e
        except ValueError as e:
            logger.warning(f"Invalid style value for cell {cell_ref}: {e}")
            raise StyleApplicationError(
                f"Invalid style value for cell {cell_ref}",
                cell_ref=cell_ref,
                style_type="value_error"
            ) from e
        except Exception as e:
            logger.error(f"Unexpected error styling cell {cell_ref}: {e}")
            raise StyleApplicationError(
                f"Unexpected error applying style to cell {cell_ref}: {str(e)}",
                cell_ref=cell_ref,
                style_type="unexpected"
            ) from e
    
    def apply_styles_to_worksheet(self, worksheet, analysis_results: Dict[int, RowAnalysis]) -> int:
        """Apply the actual styles to the worksheet based on analysis results."""
        styled_cells = 0
        
        for row_num, analysis in analysis_results.items():
            try:
                if analysis.status == RowStatus.COMPLETE:
                    # Apply complete style to all cells in the row
                    for col in RowAnalyzer.ANALYSIS_COLUMNS:
                        cell_ref = f"{col}{row_num}"
                        self.apply_cell_style(worksheet, cell_ref, StyleDefinitions.COMPLETE_STYLE)
                        styled_cells += 1
                
                elif analysis.status == RowStatus.PARTIAL:
                    # Apply different styles to filled vs empty cells
                    for cell_ref in analysis.filled_cells:
                        self.apply_cell_style(worksheet, cell_ref, StyleDefinitions.PARTIAL_FILLED_STYLE)
                        styled_cells += 1
                    
                    for cell_ref in analysis.empty_cells:
                        self.apply_cell_style(worksheet, cell_ref, StyleDefinitions.PARTIAL_EMPTY_STYLE)
                        styled_cells += 1
                
                elif analysis.status == RowStatus.EMPTY:
                    # Apply empty style to all cells in the row
                    for col in RowAnalyzer.ANALYSIS_COLUMNS:
                        cell_ref = f"{col}{row_num}"
                        self.apply_cell_style(worksheet, cell_ref, StyleDefinitions.EMPTY_STYLE)
                        styled_cells += 1
                        
            except StyleApplicationError as e:
                logger.error(f"Style application failed for row {row_num}: {e}")
                continue
            except CellAccessError as e:
                logger.error(f"Cell access failed for row {row_num}: {e}")
                continue
            except Exception as e:
                logger.error(f"Unexpected error styling row {row_num}: {e}")
                continue
        
        return styled_cells
    
    @handle_budget_exception
    def apply_conditional_formatting(self, workbook) -> bool:
        """Main method to apply conditional formatting to PM Overview."""
        logger.info("Starting PM Overview conditional formatting")
        
        try:
            if not OPENPYXL_AVAILABLE:
                messagebox.showerror("Error", "openpyxl library is not available. Please install it with: pip install openpyxl")
                return False
            
            # Validate PM Overview worksheet exists
            if self.pm_overview_sheet_name not in workbook.sheetnames:
                error_msg = f"'{self.pm_overview_sheet_name}' worksheet not found in workbook.\n\nAvailable worksheets:\n" + "\n".join(workbook.sheetnames)
                self.show_debug_window("Validation Error", error_msg)
                messagebox.showerror("Error", f"'{self.pm_overview_sheet_name}' worksheet not found.")
                return False
            
            # Get PM Overview worksheet
            worksheet = workbook[self.pm_overview_sheet_name]
            
            # Analyze all rows
            analysis_results = self.analyze_all_rows(worksheet)
            
            if not analysis_results:
                messagebox.showwarning("Warning", "No partner rows found in PM Overview for formatting.")
                return False
            
            # Generate and show debug report
            debug_report = self.generate_debug_report(analysis_results)
            self.show_debug_window("PM Overview Styling Analysis", debug_report)
            
            # Apply styles to worksheet
            styled_cells = self.apply_styles_to_worksheet(worksheet, analysis_results)
            
            if styled_cells > 0:
                complete_count = sum(1 for a in analysis_results.values() if a.status == RowStatus.COMPLETE)
                partial_count = sum(1 for a in analysis_results.values() if a.status == RowStatus.PARTIAL)
                empty_count = sum(1 for a in analysis_results.values() if a.status == RowStatus.EMPTY)
                
                success_msg = (
                    f"Applied conditional formatting to {len(analysis_results)} rows ({styled_cells} cells):\n\n"
                    f"✅ Complete Rows: {complete_count} (Green)\n"
                    f"🔵 Partial Rows: {partial_count} (Blue/Gray)\n"
                    f"❌ Empty Rows: {empty_count} (Red)\n\n"
                    f"Formatting reflects current row completion status."
                )
                messagebox.showinfo("Success", success_msg)
                return True
            else:
                messagebox.showerror("Error", "No cells were formatted successfully.")
                return False
                
        except BudgetError as e:
            logger.error(f"Budget error during PM formatting: {e}")
            error_msg = f"PM formatting error:\n{str(e)}\n\nContext: {e.context}"
            self.show_debug_window("Budget Error", error_msg)
            messagebox.showerror("Budget Error", f"Formatting failed: {str(e)}")
            return False
        except AttributeError as e:
            logger.error(f"Missing attribute during PM formatting: {e}")
            error_msg = f"Missing required attribute:\n{str(e)}\n\nThis may indicate a corrupted workbook or missing openpyxl features."
            self.show_debug_window("Attribute Error", error_msg)
            messagebox.showerror("Error", "Missing required workbook features. Please check your Excel file.")
            return False
        except KeyError as e:
            logger.error(f"Missing worksheet or data during PM formatting: {e}")
            error_msg = f"Missing required worksheet or data:\n{str(e)}\n\nPlease ensure the workbook contains the expected worksheets."
            self.show_debug_window("Data Error", error_msg)
            messagebox.showerror("Error", "Required worksheet or data not found.")
            return False
        except ValueError as e:
            logger.error(f"Invalid data during PM formatting: {e}")
            error_msg = f"Invalid data encountered:\n{str(e)}\n\nPlease check your data for invalid values."
            self.show_debug_window("Value Error", error_msg)
            messagebox.showerror("Error", "Invalid data found. Please check your worksheet data.")
            return False
        except Exception as e:
            logger.exception("Unexpected error during PM Overview formatting")
            error_msg = f"Critical error during PM Overview formatting:\n{str(e)}\n\nTraceback:\n{traceback.format_exc()}"
            self.show_debug_window("Critical Error", error_msg)
            messagebox.showerror("Error", f"Formatting failed: {str(e)}")
            return False


def apply_pm_overview_formatting(parent_window, workbook) -> bool:
    """Apply PM Overview conditional formatting with progress tracking."""
    try:
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Error", "openpyxl library is not available. Please install it with: pip install openpyxl")
            return False
        
        formatter = PMOverviewFormatter(parent_window)
        return formatter.apply_conditional_formatting(workbook)
        
    except Exception as e:
        error_msg = f"Failed to create PM Overview formatter:\n{str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        
        # Show error in a debug window
        # Use shared root pattern to avoid multiple Tk() instances
        if parent_window:
            debug_window = tk.Toplevel(parent_window)
        else:
            # Import shared root utility
            from ..utils.window_positioning import ScreenInfo
            if not ScreenInfo._shared_root:
                ScreenInfo._shared_root = tk.Tk()
                ScreenInfo._shared_root.withdraw()  # Hide the main root
            debug_window = tk.Toplevel(ScreenInfo._shared_root)
        debug_window.title("PM Overview Formatter Error")
        debug_window.geometry("600x400")
        
        text_widget = tk.Text(debug_window, wrap=tk.WORD, font=("Consolas", 10))
        text_widget.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        text_widget.insert(tk.END, error_msg)
        text_widget.config(state=tk.DISABLED)
        
        close_btn = tk.Button(debug_window, text="Close", command=debug_window.destroy)
        close_btn.pack(pady=5)
        
        messagebox.showerror("Error", f"Formatter creation failed: {str(e)}")
        return False