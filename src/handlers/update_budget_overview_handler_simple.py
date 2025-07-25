"""
Simplified Budget Overview update handler for ProjectBudgetinator.

This module provides a simplified, robust implementation that handles
missing dependencies gracefully and provides detailed error reporting.
"""

import tkinter as tk
from tkinter import messagebox
from typing import Dict, Any, List, Optional, Tuple
import traceback

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None
    Worksheet = None

# Budget Overview cell mappings configuration
BUDGET_OVERVIEW_CELL_MAPPINGS = {
    # Complete cell mapping from partner worksheets to Budget Overview
    # Format: source_cell -> target_column
    'D4': 'B',   # Partner ID Code
    'D13': 'C',  # Name of Beneficiary (from row 13)
    'D5': 'D',   # Name of Beneficiary (from row 5)
    'D6': 'E',   # Country
    
    # WP data from row 13 (G13 through U13) -> columns F through T
    'G13': 'F', 'H13': 'G', 'I13': 'H', 'J13': 'I', 'K13': 'J',
    'L13': 'K', 'M13': 'L', 'N13': 'M', 'O13': 'N', 'P13': 'O',
    'Q13': 'P', 'R13': 'Q', 'S13': 'R', 'T13': 'S', 'U13': 'T',
    
    # Additional data (skip column U in target, so V13->U, W13->V, X13->W)
    'V13': 'U', 'W13': 'V', 'X13': 'W'
}

# Debug configuration
DEBUG_ENABLED = True


def get_budget_overview_row(partner_number: int) -> int:
    """Calculate Budget Overview row number from partner number."""
    return partner_number + 7  # P2->Row9, P3->Row10, etc.


def get_partner_number_from_sheet_name(sheet_name: str) -> Optional[int]:
    """Extract partner number from sheet name."""
    if not sheet_name.startswith('P'):
        return None
    
    try:
        parts = sheet_name[1:].split('-', 1)
        if parts and parts[0].isdigit():
            partner_num = int(parts[0])
            if 2 <= partner_num <= 20:
                return partner_num
    except (ValueError, IndexError):
        pass
    
    return None


def validate_cell_reference(cell_ref: str) -> bool:
    """
    Validate Excel cell reference format.
    
    Args:
        cell_ref: Cell reference like 'A1', 'B2', etc.
        
    Returns:
        bool: True if valid, False otherwise
    """
    import re
    # Pattern for valid Excel cell reference (A1 notation)
    # Excel has max 16384 columns (XFD) and 1048576 rows
    pattern = r'^[A-Z]{1,3}[1-9]\d{0,6}$'
    
    if not re.match(pattern, cell_ref):
        return False
    
    # Additional validation for column limits
    col_part = ''.join(c for c in cell_ref if c.isalpha())
    row_part = ''.join(c for c in cell_ref if c.isdigit())
    
    # Check column limit (XFD is the max)
    if len(col_part) > 3:
        return False
    if len(col_part) == 3 and col_part > 'XFD':
        return False
    
    # Check row limit (1048576 is the max)
    if int(row_part) > 1048576:
        return False
    
    return True


def safe_get_cell_value(worksheet, cell_ref: str) -> tuple:
    """
    Safely get cell value with error handling.
    
    Args:
        worksheet: Excel worksheet
        cell_ref: Cell reference
        
    Returns:
        tuple: (success, value, error_message)
    """
    try:
        if not validate_cell_reference(cell_ref):
            return False, None, f"Invalid cell reference: {cell_ref}"
            
        cell = worksheet[cell_ref]
        
        # Handle formula cells properly
        if hasattr(cell, 'data_type') and cell.data_type == 'f':
            # This is a formula cell, get the calculated value
            calculated_value = cell.value
            return True, calculated_value, None
        else:
            # Regular value cell
            return True, cell.value, None
            
    except Exception as e:
        return False, None, str(e)


class SimpleBudgetOverviewHandler:
    """Simplified Budget Overview update handler."""
    
    def __init__(self, parent_window: Optional[tk.Widget] = None):
        """Initialize the handler."""
        self.parent = parent_window
        self.budget_overview_sheet_name = "Budget Overview"
        self.debug_window = None
    
    def show_debug_window(self, title: str, content: str):
        """Show debug information in a window."""
        if self.debug_window:
            self.debug_window.destroy()
        
        # Only create debug window if we have a parent window
        if self.parent:
            self.debug_window = tk.Toplevel(self.parent)
        else:
            # For automatic operations, don't create debug window
            self.debug_window = None
            return
        self.debug_window.title(title)
        self.debug_window.geometry("800x600")
        
        # Create text widget with scrollbar
        frame = tk.Frame(self.debug_window)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        text_widget = tk.Text(frame, wrap=tk.WORD, font=("Consolas", 10))
        scrollbar = tk.Scrollbar(frame, orient=tk.VERTICAL, command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        text_widget.insert(tk.END, content)
        text_widget.config(state=tk.DISABLED)
        
        # Close button
        close_btn = tk.Button(self.debug_window, text="Close",
                              command=self.debug_window.destroy)
        close_btn.pack(pady=5)
    
    def validate_dependencies(self) -> Tuple[bool, str]:
        """Validate that required dependencies are available."""
        debug_info = []
        debug_info.append("🔍 DEPENDENCY CHECK")
        debug_info.append("=" * 50)
        
        if not OPENPYXL_AVAILABLE:
            debug_info.append("❌ openpyxl not available")
            debug_info.append("   Please install: pip install openpyxl")
            return False, "\n".join(debug_info)
        
        debug_info.append("✅ openpyxl available")
        debug_info.append("✅ All dependencies satisfied")
        return True, "\n".join(debug_info)
    
    def get_partner_worksheets(self, workbook) -> List[Tuple[str, int]]:
        """Get list of partner worksheets in the workbook."""
        partner_sheets = []
        debug_info = []
        
        debug_info.append("🔍 DISCOVERING PARTNER WORKSHEETS")
        debug_info.append("=" * 50)
        debug_info.append(f"Total worksheets: {len(workbook.sheetnames)}")
        debug_info.append("")
        
        for sheet_name in workbook.sheetnames:
            partner_number = get_partner_number_from_sheet_name(sheet_name)
            if partner_number:
                partner_sheets.append((sheet_name, partner_number))
                target_row = get_budget_overview_row(partner_number)
                debug_info.append(f"✅ {sheet_name} → Partner {partner_number} → Budget Row {target_row}")
            else:
                debug_info.append(f"⚪ {sheet_name} (not a partner sheet)")
        
        partner_sheets.sort(key=lambda x: x[1])
        
        debug_info.append("")
        debug_info.append(f"📊 Found {len(partner_sheets)} partner worksheets")
        
        if DEBUG_ENABLED:
            self.show_debug_window("Partner Worksheets Discovery", "\n".join(debug_info))
        
        return partner_sheets
    
    def extract_partner_data(self, worksheet, partner_number: int) -> Dict[str, Any]:
        """Extract data from a partner worksheet with detailed debugging."""
        data = {
            'partner_number': partner_number,
            'cell_mappings': {}
        }
        
        debug_info = []
        debug_info.append(f"🔍 EXTRACTING DATA from Partner {partner_number}")
        debug_info.append(f"Worksheet: {worksheet.title}")
        debug_info.append("=" * 80)
        debug_info.append("SOURCE | VALUE                    | TARGET | STATUS")
        debug_info.append("-" * 80)
        
        for source_cell, target_col in BUDGET_OVERVIEW_CELL_MAPPINGS.items():
            # Use safe cell value extraction
            success, value, error_msg = safe_get_cell_value(worksheet, source_cell)
            
            if success:
                # Ensure we store the actual value, not a formula
                formatted_value = value if value is not None else ''
                
                data['cell_mappings'][source_cell] = {
                    'original_value': value,
                    'calculated_value': value,
                    'target_col': target_col,
                    'formatted_value': formatted_value,
                    'cell_type': 'calculated' if value is not None else 'empty'
                }
                
                status = "✅ OK"
                value_display = str(value)[:20] if value is not None else "None"
                debug_info.append(f"{source_cell:>6} | {value_display:<20} | {target_col:>6} | {status}")
                
            else:
                data['cell_mappings'][source_cell] = {
                    'value': None,
                    'target_col': target_col,
                    'formatted_value': '',
                    'error': error_msg
                }
                debug_info.append(f"{source_cell:>6} | ERROR: {str(error_msg):<15} | {target_col:>6} | ❌ FAIL")
        
        debug_info.append("-" * 80)
        debug_info.append(f"✅ Extracted {len(data['cell_mappings'])} cell mappings")
        
        if DEBUG_ENABLED:
            self.show_debug_window(f"Partner {partner_number} Data Extraction", "\n".join(debug_info))
        
        return data
    
    def update_budget_row(self, budget_ws, partner_data: Dict[str, Any]) -> bool:
        """Update a specific row in the Budget Overview worksheet."""
        partner_number = partner_data['partner_number']
        target_row = get_budget_overview_row(partner_number)
        cell_mappings = partner_data.get('cell_mappings', {})
        
        debug_info = []
        debug_info.append(f"🎯 UPDATING Budget Overview Row {target_row}")
        debug_info.append(f"Partner: {partner_number}")
        debug_info.append("=" * 80)
        debug_info.append("SOURCE | TARGET | VALUE                    | STATUS")
        debug_info.append("-" * 80)
        
        success_count = 0
        
        for source_cell, mapping_info in cell_mappings.items():
            try:
                target_col = mapping_info['target_col']
                value = mapping_info['formatted_value']
                target_cell = f"{target_col}{target_row}"
                
                # Get the target cell object
                target_cell_obj = budget_ws[target_cell]
                
                # Safe value assignment without manual data type manipulation
                try:
                    # Let openpyxl handle data type detection automatically
                    if value is None or value == '':
                        # Explicitly handle empty values
                        target_cell_obj.value = None
                    elif isinstance(value, str) and value.startswith('='):
                        # This is a formula - let openpyxl handle it
                        target_cell_obj.value = value
                        # Don't manually set data_type for formulas
                    else:
                        # For regular values, let openpyxl auto-detect
                        target_cell_obj.value = value
                        
                    debug_info.append(f"✅ Assigned value to {target_cell}: {value} (type: {type(value).__name__})")
                    
                except Exception as assign_error:
                    debug_info.append(f"❌ Failed to assign value to {target_cell}: {assign_error}")
                    raise assign_error
                
                success_count += 1
                
                status = "✅ OK"
                value_display = str(value)[:20] if value != '' else "EMPTY"
                cell_type = mapping_info.get('cell_type', '')
                debug_info.append(f"{source_cell:>6} | {target_cell:>6} | {value_display:<20} | {status} → {cell_type}")
                
            except Exception as e:
                target_col = mapping_info.get('target_col', 'UNKNOWN')
                status = f"❌ {str(e)[:15]}"
                debug_info.append(f"{source_cell:>6} | {target_col}{target_row} | ERROR               | {status}")
        
        debug_info.append("-" * 80)
        debug_info.append(f"✅ Updated {success_count}/{len(cell_mappings)} cells successfully")
        
        if DEBUG_ENABLED:
            self.show_debug_window(f"Budget Overview Row {target_row} Update", "\n".join(debug_info))
        
        return success_count > 0
    
    def update_budget_overview(self, workbook) -> bool:
        """Update Budget Overview worksheet with partner data."""
        try:
            # Check dependencies
            deps_ok, deps_info = self.validate_dependencies()
            if not deps_ok:
                if DEBUG_ENABLED:
                    self.show_debug_window("Dependency Check Failed", deps_info)
                messagebox.showerror("Error", "Missing dependencies. Please check the debug window for details.")
                return False
            
            # Validate Budget Overview worksheet exists
            if self.budget_overview_sheet_name not in workbook.sheetnames:
                error_msg = f"'{self.budget_overview_sheet_name}' worksheet not found in workbook.\n\nAvailable worksheets:\n" + "\n".join(workbook.sheetnames)
                if DEBUG_ENABLED:
                    self.show_debug_window("Validation Error", error_msg)
                messagebox.showerror("Error", f"'{self.budget_overview_sheet_name}' worksheet not found.")
                return False
            
            # Get partner worksheets
            partner_sheets = self.get_partner_worksheets(workbook)
            if not partner_sheets:
                messagebox.showwarning("Warning", "No partner worksheets found (P2-P20).")
                return False
            
            # Get Budget Overview worksheet
            budget_ws = workbook[self.budget_overview_sheet_name]
            
            # Update each partner
            updated_count = 0
            for sheet_name, partner_number in partner_sheets:
                try:
                    partner_ws = workbook[sheet_name]
                    partner_data = self.extract_partner_data(partner_ws, partner_number)
                    
                    if self.update_budget_row(budget_ws, partner_data):
                        updated_count += 1
                        
                except Exception as e:
                    error_msg = f"Failed to update partner {partner_number}: {str(e)}\n\nTraceback:\n{traceback.format_exc()}"
                    if DEBUG_ENABLED:
                        self.show_debug_window(f"Partner {partner_number} Update Error", error_msg)
                    continue
            
            if updated_count > 0:
                messagebox.showinfo("Success", f"Updated {updated_count} partner(s) in Budget Overview.")
                return True
            else:
                messagebox.showerror("Error", "No partners were updated successfully.")
                return False
                
        except Exception as e:
            error_msg = f"Critical error during Budget Overview update:\n{str(e)}\n\nTraceback:\n{traceback.format_exc()}"
            if DEBUG_ENABLED:
                self.show_debug_window("Critical Error", error_msg)
            messagebox.showerror("Error", f"Update failed: {str(e)}")
            return False


def update_budget_overview_with_progress(parent_window, workbook) -> bool:
    """Update Budget Overview with simplified progress handling."""
    try:
        handler = SimpleBudgetOverviewHandler(parent_window)
        return handler.update_budget_overview(workbook)
    except Exception as e:
        error_msg = f"Failed to create Budget Overview handler:\n{str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        
        # Show error in a debug window only if parent window exists
        if parent_window:
            debug_window = tk.Toplevel(parent_window)
            debug_window.title("Budget Overview Handler Error")
            debug_window.geometry("600x400")
            
            text_widget = tk.Text(debug_window, wrap=tk.WORD, font=("Consolas", 10))
            text_widget.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            text_widget.insert(tk.END, error_msg)
            text_widget.config(state=tk.DISABLED)
            
            close_btn = tk.Button(debug_window, text="Close", command=debug_window.destroy)
            close_btn.pack(pady=5)
            
            messagebox.showerror("Error", f"Handler creation failed: {str(e)}")
        else:
            # For automatic operations, just print the error (no logger available)
            print(f"Budget Overview Handler Error: {error_msg}")
        
        return False