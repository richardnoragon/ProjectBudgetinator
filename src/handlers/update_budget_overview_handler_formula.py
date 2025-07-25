"""
Formula-based Budget Overview update handler for ProjectBudgetinator.

This version updates existing formulas in the Budget Overview to reference
the correct partner worksheets, rather than copying values.
"""

import tkinter as tk
from tkinter import messagebox
from typing import Dict, Any, List, Optional, Tuple
import traceback
import re

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None
    Worksheet = None

# Budget Overview cell mappings configuration
BUDGET_OVERVIEW_CELL_MAPPINGS = {
    'D4': 'B',   # Partner ID Code
    'D13': 'C',  # Name of Beneficiary (from row 13)
    'D5': 'D',   # Name of Beneficiary (from row 5)
    'D6': 'E',   # Country
    
    # WP data from row 13 (G13 through U13) -> columns F through T
    'G13': 'F', 'H13': 'G', 'I13': 'H', 'J13': 'I', 'K13': 'J',
    'L13': 'K', 'M13': 'L', 'N13': 'M', 'O13': 'N', 'P13': 'O',
    'Q13': 'P', 'R13': 'Q', 'S13': 'R', 'T13': 'S', 'U13': 'T',
    
    # Additional data
    'V13': 'U', 'W13': 'V', 'X13': 'W'
}

DEBUG_ENABLED = True


def get_budget_overview_row(partner_number: int) -> int:
    """Calculate Budget Overview row number from partner number."""
    return partner_number + 7


def get_partner_number_from_sheet_name(sheet_name: str) -> Optional[int]:
    """Extract partner number from sheet name."""
    if not sheet_name.startswith('P'):
        return None
    
    try:
        parts = sheet_name[1:].split('-', 1)
        if parts and parts[0].isdigit():
            partner_num = int(parts[0])
            if 2 <= partner_num <= 20:
                return partner_num
    except (ValueError, IndexError):
        pass
    
    return None


class FormulaBudgetOverviewHandler:
    """Budget Overview handler that updates formula references instead of copying values."""
    
    def __init__(self, parent_window: Optional[tk.Widget] = None):
        """Initialize the handler."""
        self.parent = parent_window
        self.budget_overview_sheet_name = "Budget Overview"
        self.debug_window = None
    
    def show_debug_window(self, title: str, content: str):
        """Show debug information in a window."""
        if self.debug_window:
            self.debug_window.destroy()
        
        # Use shared root pattern to avoid multiple Tk() instances
        if self.parent:
            self.debug_window = tk.Toplevel(self.parent)
        else:
            # Import shared root utility
            from ..utils.window_positioning import ScreenInfo
            if not ScreenInfo._shared_root:
                ScreenInfo._shared_root = tk.Tk()
                ScreenInfo._shared_root.withdraw()  # Hide the main root
            self.debug_window = tk.Toplevel(ScreenInfo._shared_root)
        self.debug_window.title(title)
        self.debug_window.geometry("1000x700")
        
        # Create text widget with scrollbar
        frame = tk.Frame(self.debug_window)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        text_widget = tk.Text(frame, wrap=tk.WORD, font=("Consolas", 9))
        scrollbar = tk.Scrollbar(frame, orient=tk.VERTICAL, command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        text_widget.insert(tk.END, content)
        text_widget.config(state=tk.DISABLED)
        
        # Close button
        close_btn = tk.Button(self.debug_window, text="Close",
                              command=self.debug_window.destroy)
        close_btn.pack(pady=5)
    
    def create_formula_reference(self, partner_sheet_name: str, source_cell: str) -> str:
        """Create a formula reference to a partner worksheet cell."""
        # Handle sheet names with hyphens or spaces by wrapping in single quotes
        if '-' in partner_sheet_name or ' ' in partner_sheet_name:
            sheet_ref = f"'{partner_sheet_name}'"
        else:
            sheet_ref = partner_sheet_name
        
        return f"={sheet_ref}!{source_cell}"
    
    def get_partner_worksheets(self, workbook) -> List[Tuple[str, int]]:
        """Get list of partner worksheets in the workbook."""
        partner_sheets = []
        debug_info = []
        
        debug_info.append("🔍 DISCOVERING PARTNER WORKSHEETS")
        debug_info.append("=" * 70)
        debug_info.append(f"Total worksheets: {len(workbook.sheetnames)}")
        debug_info.append("")
        
        for sheet_name in workbook.sheetnames:
            partner_number = get_partner_number_from_sheet_name(sheet_name)
            if partner_number:
                partner_sheets.append((sheet_name, partner_number))
                target_row = get_budget_overview_row(partner_number)
                debug_info.append(f"✅ {sheet_name} → Partner {partner_number} → Budget Row {target_row}")
            else:
                debug_info.append(f"⚪ {sheet_name} (not a partner sheet)")
        
        partner_sheets.sort(key=lambda x: x[1])
        
        debug_info.append("")
        debug_info.append(f"📊 Found {len(partner_sheets)} partner worksheets")
        
        if DEBUG_ENABLED:
            self.show_debug_window("Partner Worksheets Discovery", "\n".join(debug_info))
        
        return partner_sheets
    
    def update_budget_row_formulas(self, budget_ws, partner_sheet_name: str, partner_number: int) -> bool:
        """Update formulas in Budget Overview to reference the partner worksheet."""
        target_row = get_budget_overview_row(partner_number)
        
        debug_info = []
        debug_info.append(f"🎯 UPDATING Budget Overview Row {target_row} FORMULAS")
        debug_info.append(f"Partner: {partner_number} ({partner_sheet_name})")
        debug_info.append("=" * 100)
        debug_info.append("SOURCE | TARGET | OLD FORMULA/VALUE            | NEW FORMULA                      | STATUS")
        debug_info.append("-" * 100)
        
        success_count = 0
        
        for source_cell, target_col in BUDGET_OVERVIEW_CELL_MAPPINGS.items():
            try:
                target_cell_ref = f"{target_col}{target_row}"
                target_cell = budget_ws[target_cell_ref]
                
                # Get current content
                old_content = target_cell.value if target_cell.value is not None else ""
                old_display = str(old_content)[:25] if old_content else "EMPTY"
                
                # Create new formula reference
                new_formula = self.create_formula_reference(partner_sheet_name, source_cell)
                
                # Update the cell with the new formula
                target_cell.value = new_formula
                
                success_count += 1
                status = "✅ OK"
                debug_info.append(f"{source_cell:>6} | {target_cell_ref:>6} | {old_display:<25} | {new_formula:<30} | {status}")
                
            except Exception as e:
                status = f"❌ {str(e)[:15]}"
                debug_info.append(f"{source_cell:>6} | {target_col}{target_row} | ERROR                    | ERROR                          | {status}")
        
        debug_info.append("-" * 100)
        debug_info.append(f"✅ Updated {success_count}/{len(BUDGET_OVERVIEW_CELL_MAPPINGS)} formulas successfully")
        debug_info.append("")
        debug_info.append("📋 Formula Examples:")
        debug_info.append(f"   D4 → {self.create_formula_reference(partner_sheet_name, 'D4')}")
        debug_info.append(f"   G13 → {self.create_formula_reference(partner_sheet_name, 'G13')}")
        
        if DEBUG_ENABLED:
            self.show_debug_window(f"Budget Overview Row {target_row} Formula Update", "\n".join(debug_info))
        
        return success_count > 0
    
    def analyze_existing_formulas(self, budget_ws, partner_sheets: List[Tuple[str, int]]) -> str:
        """Analyze existing formulas in Budget Overview to understand the current structure."""
        debug_info = []
        debug_info.append("🔍 ANALYZING EXISTING BUDGET OVERVIEW FORMULAS")
        debug_info.append("=" * 80)
        debug_info.append("")
        
        # Check a few sample cells to understand the current structure
        sample_partners = partner_sheets[:3]  # Check first 3 partners
        
        for partner_sheet_name, partner_number in sample_partners:
            target_row = get_budget_overview_row(partner_number)
            debug_info.append(f"📊 Partner {partner_number} (Row {target_row}):")
            
            # Check a few key cells
            sample_cells = ['B', 'C', 'D', 'F', 'G']
            for col in sample_cells:
                cell_ref = f"{col}{target_row}"
                try:
                    cell = budget_ws[cell_ref]
                    value = cell.value
                    if value:
                        value_str = str(value)[:50]
                        if value_str.startswith('='):
                            debug_info.append(f"   {cell_ref}: {value_str} (FORMULA)")
                        else:
                            debug_info.append(f"   {cell_ref}: {value_str} (VALUE)")
                    else:
                        debug_info.append(f"   {cell_ref}: EMPTY")
                except Exception as e:
                    debug_info.append(f"   {cell_ref}: ERROR - {str(e)}")
            debug_info.append("")
        
        return "\n".join(debug_info)
    
    def update_budget_overview(self, workbook) -> bool:
        """Update Budget Overview worksheet by updating formula references."""
        try:
            # Validate Budget Overview worksheet exists
            if self.budget_overview_sheet_name not in workbook.sheetnames:
                error_msg = f"'{self.budget_overview_sheet_name}' worksheet not found in workbook.\n\nAvailable worksheets:\n" + "\n".join(workbook.sheetnames)
                if DEBUG_ENABLED:
                    self.show_debug_window("Validation Error", error_msg)
                messagebox.showerror("Error", f"'{self.budget_overview_sheet_name}' worksheet not found.")
                return False
            
            # Get partner worksheets
            partner_sheets = self.get_partner_worksheets(workbook)
            if not partner_sheets:
                messagebox.showwarning("Warning", "No partner worksheets found (P2-P20).")
                return False
            
            # Get Budget Overview worksheet
            budget_ws = workbook[self.budget_overview_sheet_name]
            
            # Analyze existing formulas first
            if DEBUG_ENABLED:
                analysis = self.analyze_existing_formulas(budget_ws, partner_sheets)
                self.show_debug_window("Existing Formula Analysis", analysis)
            
            # Update formulas for each partner
            updated_count = 0
            for partner_sheet_name, partner_number in partner_sheets:
                try:
                    if self.update_budget_row_formulas(budget_ws, partner_sheet_name, partner_number):
                        updated_count += 1
                        
                except Exception as e:
                    error_msg = f"Failed to update formulas for partner {partner_number}: {str(e)}\n\nTraceback:\n{traceback.format_exc()}"
                    if DEBUG_ENABLED:
                        self.show_debug_window(f"Partner {partner_number} Formula Update Error", error_msg)
                    continue
            
            if updated_count > 0:
                # Apply conditional formatting after successful formula updates
                try:
                    from .budget_overview_format import BudgetOverviewFormatter
                    formatter = BudgetOverviewFormatter(self.parent)
                    formatting_success = formatter.apply_conditional_formatting(workbook)
                    
                    if formatting_success:
                        success_msg = (
                            f"✅ Updated formulas for {updated_count} partner(s) in Budget Overview.\n"
                            f"🎨 Applied conditional formatting based on row completion.\n\n"
                            f"Formulas now reference the correct partner worksheets with visual styling."
                        )
                    else:
                        success_msg = (
                            f"✅ Updated formulas for {updated_count} partner(s) in Budget Overview.\n"
                            f"⚠️ Formatting could not be applied (see debug window for details).\n\n"
                            f"Formulas now reference the correct partner worksheets."
                        )
                except ImportError as import_err:
                    success_msg = (
                        f"✅ Updated formulas for {updated_count} partner(s) in Budget Overview.\n"
                        f"⚠️ Formatting module not available: {str(import_err)}\n\n"
                        f"Formulas now reference the correct partner worksheets."
                    )
                except Exception as format_err:
                    success_msg = (
                        f"✅ Updated formulas for {updated_count} partner(s) in Budget Overview.\n"
                        f"⚠️ Formatting failed: {str(format_err)}\n\n"
                        f"Formulas now reference the correct partner worksheets."
                    )
                
                messagebox.showinfo("Success", success_msg)
                return True
            else:
                messagebox.showerror("Error", "No partner formulas were updated successfully.")
                return False
                
        except Exception as e:
            error_msg = f"Critical error during Budget Overview formula update:\n{str(e)}\n\nTraceback:\n{traceback.format_exc()}"
            if DEBUG_ENABLED:
                self.show_debug_window("Critical Error", error_msg)
            messagebox.showerror("Error", f"Formula update failed: {str(e)}")
            return False


def update_budget_overview_with_progress(parent_window, workbook) -> bool:
    """Update Budget Overview with formula-based approach."""
    try:
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Error", "openpyxl library is not available. Please install it with: pip install openpyxl")
            return False
        
        handler = FormulaBudgetOverviewHandler(parent_window)
        return handler.update_budget_overview(workbook)
        
    except Exception as e:
        error_msg = f"Failed to create Budget Overview handler:\n{str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        
        # Show error in a debug window
        # Use shared root pattern to avoid multiple Tk() instances
        if parent_window:
            debug_window = tk.Toplevel(parent_window)
        else:
            # Import shared root utility
            from ..utils.window_positioning import ScreenInfo
            if not ScreenInfo._shared_root:
                ScreenInfo._shared_root = tk.Tk()
                ScreenInfo._shared_root.withdraw()  # Hide the main root
            debug_window = tk.Toplevel(ScreenInfo._shared_root)
        debug_window.title("Budget Overview Handler Error")
        debug_window.geometry("600x400")
        
        text_widget = tk.Text(debug_window, wrap=tk.WORD, font=("Consolas", 10))
        text_widget.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        text_widget.insert(tk.END, error_msg)
        text_widget.config(state=tk.DISABLED)
        
        close_btn = tk.Button(debug_window, text="Close", command=debug_window.destroy)
        close_btn.pack(pady=5)
        
        messagebox.showerror("Error", f"Handler creation failed: {str(e)}")
        return False