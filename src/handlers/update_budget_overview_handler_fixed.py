"""
Fixed Budget Overview update handler for ProjectBudgetinator.

This version specifically addresses the formula vs. value issue by ensuring
we extract calculated values from source cells and write only values to target cells.
"""

import tkinter as tk
from tkinter import messagebox
from typing import Dict, Any, List, Optional, Tuple
import traceback

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None
    Worksheet = None

# Budget Overview cell mappings configuration
BUDGET_OVERVIEW_CELL_MAPPINGS = {
    'D4': 'B',   # Partner ID Code
    'D13': 'C',  # Name of Beneficiary (from row 13)
    'D5': 'D',   # Name of Beneficiary (from row 5)
    'D6': 'E',   # Country
    
    # WP data from row 13 (G13 through U13) -> columns F through T
    'G13': 'F', 'H13': 'G', 'I13': 'H', 'J13': 'I', 'K13': 'J',
    'L13': 'K', 'M13': 'L', 'N13': 'M', 'O13': 'N', 'P13': 'O',
    'Q13': 'P', 'R13': 'Q', 'S13': 'R', 'T13': 'S', 'U13': 'T',
    
    # Additional data
    'V13': 'U', 'W13': 'V', 'X13': 'W'
}

DEBUG_ENABLED = True


def get_budget_overview_row(partner_number: int) -> int:
    """Calculate Budget Overview row number from partner number."""
    return partner_number + 7


def get_partner_number_from_sheet_name(sheet_name: str) -> Optional[int]:
    """Extract partner number from sheet name."""
    if not sheet_name.startswith('P'):
        return None
    
    try:
        parts = sheet_name[1:].split(' ', 1)
        if parts and parts[0].isdigit():
            partner_num = int(parts[0])
            if 2 <= partner_num <= 20:
                return partner_num
    except (ValueError, IndexError):
        pass
    
    return None


class FixedBudgetOverviewHandler:
    """Fixed Budget Overview update handler that properly handles formulas vs values."""
    
    def __init__(self, parent_window: Optional[tk.Widget] = None):
        """Initialize the handler."""
        self.parent = parent_window
        self.budget_overview_sheet_name = "Budget Overview"
        self.debug_window = None
    
    def show_debug_window(self, title: str, content: str):
        """Show debug information in a window."""
        if self.debug_window:
            self.debug_window.destroy()
        
        # Use shared root pattern to avoid multiple Tk() instances
        if self.parent:
            self.debug_window = tk.Toplevel(self.parent)
        else:
            # Import shared root utility
            from ..utils.window_positioning import ScreenInfo
            if not ScreenInfo._shared_root:
                ScreenInfo._shared_root = tk.Tk()
                ScreenInfo._shared_root.withdraw()  # Hide the main root
            self.debug_window = tk.Toplevel(ScreenInfo._shared_root)
        self.debug_window.title(title)
        self.debug_window.geometry("900x700")
        
        # Create text widget with scrollbar
        frame = tk.Frame(self.debug_window)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        text_widget = tk.Text(frame, wrap=tk.WORD, font=("Consolas", 9))
        scrollbar = tk.Scrollbar(frame, orient=tk.VERTICAL, command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        text_widget.insert(tk.END, content)
        text_widget.config(state=tk.DISABLED)
        
        # Close button
        close_btn = tk.Button(self.debug_window, text="Close",
                              command=self.debug_window.destroy)
        close_btn.pack(pady=5)
    
    def extract_cell_value(self, cell, source_cell: str) -> Tuple[Any, str, str]:
        """
        Extract the actual value from a cell, handling formulas properly.
        
        Returns:
            Tuple[value, cell_type, debug_info]
        """
        debug_info = []
        
        if cell.value is None:
            return None, "empty", "Empty cell"
        
        # Check if this is a formula cell
        if hasattr(cell, 'data_type') and cell.data_type == 'f':
            # This is a formula cell
            formula = getattr(cell, 'formula', str(cell.value))
            calculated_value = cell.value
            
            debug_info.append(f"📋 Formula cell: {formula}")
            debug_info.append(f"📊 Calculated value: {calculated_value}")
            
            return calculated_value, "formula", "; ".join(debug_info)
        
        elif isinstance(cell.value, str) and cell.value.startswith('='):
            # String that looks like a formula
            debug_info.append(f"⚠️  String formula detected: {cell.value}")
            return cell.value, "string_formula", "; ".join(debug_info)
        
        else:
            # Regular value cell
            value = cell.value
            value_type = type(value).__name__
            debug_info.append(f"📄 Direct value ({value_type}): {value}")
            
            return value, "value", "; ".join(debug_info)
    
    def get_partner_worksheets(self, workbook) -> List[Tuple[str, int]]:
        """Get list of partner worksheets in the workbook."""
        partner_sheets = []
        debug_info = []
        
        debug_info.append("🔍 DISCOVERING PARTNER WORKSHEETS")
        debug_info.append("=" * 60)
        debug_info.append(f"Total worksheets: {len(workbook.sheetnames)}")
        debug_info.append("")
        
        for sheet_name in workbook.sheetnames:
            partner_number = get_partner_number_from_sheet_name(sheet_name)
            if partner_number:
                partner_sheets.append((sheet_name, partner_number))
                target_row = get_budget_overview_row(partner_number)
                debug_info.append(f"✅ {sheet_name} → Partner {partner_number} → Budget Row {target_row}")
            else:
                debug_info.append(f"⚪ {sheet_name} (not a partner sheet)")
        
        partner_sheets.sort(key=lambda x: x[1])
        
        debug_info.append("")
        debug_info.append(f"📊 Found {len(partner_sheets)} partner worksheets")
        
        if DEBUG_ENABLED:
            self.show_debug_window("Partner Worksheets Discovery", "\n".join(debug_info))
        
        return partner_sheets
    
    def extract_partner_data(self, worksheet, partner_number: int) -> Dict[str, Any]:
        """Extract data from a partner worksheet, handling formulas vs values."""
        data = {
            'partner_number': partner_number,
            'cell_mappings': {}
        }
        
        debug_info = []
        debug_info.append(f"🔍 EXTRACTING DATA from Partner {partner_number}")
        debug_info.append(f"Worksheet: {worksheet.title}")
        debug_info.append("=" * 90)
        debug_info.append("SOURCE | VALUE                    | TYPE      | TARGET | DETAILS")
        debug_info.append("-" * 90)
        
        for source_cell, target_col in BUDGET_OVERVIEW_CELL_MAPPINGS.items():
            try:
                cell = worksheet[source_cell]
                value, cell_type, details = self.extract_cell_value(cell, source_cell)
                
                # Store the extracted value
                data['cell_mappings'][source_cell] = {
                    'raw_value': cell.value,
                    'extracted_value': value,
                    'target_col': target_col,
                    'cell_type': cell_type,
                    'details': details
                }
                
                value_display = str(value)[:20] if value is not None else "None"
                debug_info.append(f"{source_cell:>6} | {value_display:<20} | {cell_type:<9} | {target_col:>6} | {details}")
                
            except Exception as e:
                data['cell_mappings'][source_cell] = {
                    'raw_value': None,
                    'extracted_value': None,
                    'target_col': target_col,
                    'cell_type': 'error',
                    'details': f"Error: {str(e)}",
                    'error': str(e)
                }
                debug_info.append(f"{source_cell:>6} | ERROR: {str(e):<15} | error     | {target_col:>6} | Exception occurred")
        
        debug_info.append("-" * 90)
        debug_info.append(f"✅ Extracted {len(data['cell_mappings'])} cell mappings")
        
        if DEBUG_ENABLED:
            self.show_debug_window(f"Partner {partner_number} Data Extraction", "\n".join(debug_info))
        
        return data
    
    def update_budget_row(self, budget_ws, partner_data: Dict[str, Any]) -> bool:
        """Update a specific row in the Budget Overview worksheet with values only."""
        partner_number = partner_data['partner_number']
        target_row = get_budget_overview_row(partner_number)
        cell_mappings = partner_data.get('cell_mappings', {})
        
        debug_info = []
        debug_info.append(f"🎯 UPDATING Budget Overview Row {target_row}")
        debug_info.append(f"Partner: {partner_number}")
        debug_info.append("=" * 90)
        debug_info.append("SOURCE | TARGET | VALUE                    | TYPE      | STATUS")
        debug_info.append("-" * 90)
        
        success_count = 0
        
        for source_cell, mapping_info in cell_mappings.items():
            try:
                target_col = mapping_info['target_col']
                value = mapping_info['extracted_value']
                cell_type = mapping_info['cell_type']
                target_cell_ref = f"{target_col}{target_row}"
                
                # Get the target cell object
                target_cell = budget_ws[target_cell_ref]
                
                # Set the value and ensure it's not interpreted as a formula
                if value is not None:
                    target_cell.value = value
                    
                    # Force the cell to be a value cell, not a formula
                    if hasattr(target_cell, 'data_type'):
                        if isinstance(value, (int, float)):
                            target_cell.data_type = 'n'  # numeric
                        elif isinstance(value, str) and not str(value).startswith('='):
                            target_cell.data_type = 's'  # string
                        else:
                            # For any string that might be interpreted as formula, force it to be text
                            target_cell.data_type = 'inlineStr'
                else:
                    target_cell.value = ""
                    if hasattr(target_cell, 'data_type'):
                        target_cell.data_type = 's'
                
                success_count += 1
                status = "✅ OK"
                value_display = str(value)[:20] if value is not None else "EMPTY"
                debug_info.append(f"{source_cell:>6} | {target_cell_ref:>6} | {value_display:<20} | {cell_type:<9} | {status}")
                
            except Exception as e:
                target_col = mapping_info.get('target_col', 'UNKNOWN')
                status = f"❌ {str(e)[:15]}"
                debug_info.append(f"{source_cell:>6} | {target_col}{target_row} | ERROR               | error     | {status}")
        
        debug_info.append("-" * 90)
        debug_info.append(f"✅ Updated {success_count}/{len(cell_mappings)} cells successfully")
        
        if DEBUG_ENABLED:
            self.show_debug_window(f"Budget Overview Row {target_row} Update", "\n".join(debug_info))
        
        return success_count > 0
    
    def update_budget_overview(self, workbook) -> bool:
        """Update Budget Overview worksheet with partner data."""
        try:
            # Validate Budget Overview worksheet exists
            if self.budget_overview_sheet_name not in workbook.sheetnames:
                error_msg = f"'{self.budget_overview_sheet_name}' worksheet not found in workbook.\n\nAvailable worksheets:\n" + "\n".join(workbook.sheetnames)
                if DEBUG_ENABLED:
                    self.show_debug_window("Validation Error", error_msg)
                messagebox.showerror("Error", f"'{self.budget_overview_sheet_name}' worksheet not found.")
                return False
            
            # Get partner worksheets
            partner_sheets = self.get_partner_worksheets(workbook)
            if not partner_sheets:
                messagebox.showwarning("Warning", "No partner worksheets found (P2-P20).")
                return False
            
            # Get Budget Overview worksheet
            budget_ws = workbook[self.budget_overview_sheet_name]
            
            # Update each partner
            updated_count = 0
            for sheet_name, partner_number in partner_sheets:
                try:
                    partner_ws = workbook[sheet_name]
                    partner_data = self.extract_partner_data(partner_ws, partner_number)
                    
                    if self.update_budget_row(budget_ws, partner_data):
                        updated_count += 1
                        
                except Exception as e:
                    error_msg = f"Failed to update partner {partner_number}: {str(e)}\n\nTraceback:\n{traceback.format_exc()}"
                    if DEBUG_ENABLED:
                        self.show_debug_window(f"Partner {partner_number} Update Error", error_msg)
                    continue
            
            if updated_count > 0:
                messagebox.showinfo("Success", f"Updated {updated_count} partner(s) in Budget Overview.\n\nValues (not formulas) have been copied to the target cells.")
                return True
            else:
                messagebox.showerror("Error", "No partners were updated successfully.")
                return False
                
        except Exception as e:
            error_msg = f"Critical error during Budget Overview update:\n{str(e)}\n\nTraceback:\n{traceback.format_exc()}"
            if DEBUG_ENABLED:
                self.show_debug_window("Critical Error", error_msg)
            messagebox.showerror("Error", f"Update failed: {str(e)}")
            return False


def update_budget_overview_with_progress(parent_window, workbook) -> bool:
    """Update Budget Overview with fixed formula handling."""
    try:
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Error", "openpyxl library is not available. Please install it with: pip install openpyxl")
            return False
        
        handler = FixedBudgetOverviewHandler(parent_window)
        return handler.update_budget_overview(workbook)
        
    except Exception as e:
        error_msg = f"Failed to create Budget Overview handler:\n{str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        
        # Show error in a debug window
        # Use shared root pattern to avoid multiple Tk() instances
        if parent_window:
            debug_window = tk.Toplevel(parent_window)
        else:
            # Import shared root utility
            from ..utils.window_positioning import ScreenInfo
            if not ScreenInfo._shared_root:
                ScreenInfo._shared_root = tk.Tk()
                ScreenInfo._shared_root.withdraw()  # Hide the main root
            debug_window = tk.Toplevel(ScreenInfo._shared_root)
        debug_window.title("Budget Overview Handler Error")
        debug_window.geometry("600x400")
        
        text_widget = tk.Text(debug_window, wrap=tk.WORD, font=("Consolas", 10))
        text_widget.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        text_widget.insert(tk.END, error_msg)
        text_widget.config(state=tk.DISABLED)
        
        close_btn = tk.Button(debug_window, text="Close", command=debug_window.destroy)
        close_btn.pack(pady=5)
        
        messagebox.showerror("Error", f"Handler creation failed: {str(e)}")
        return False