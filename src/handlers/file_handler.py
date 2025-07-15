"""
File management functions for working with Excel workbooks.
"""
from tkinter import filedialog
from tkinter import messagebox
import openpyxl
import os
import time
from utils.error_handler import ExceptionHandler
from utils.security_validator import SecurityValidator
from gui.progress_dialog import ProgressContext, show_progress_for_operation
from logger import get_structured_logger, LogContext


# Create exception handler instance
exception_handler = ExceptionHandler()
logger = get_structured_logger("handlers.file")


def _open_file_with_progress(progress_dialog, filepath):
    """
    Open file with progress feedback.
    
    Args:
        progress_dialog: Progress dialog instance
        filepath: Path to the file to open
        
    Returns:
        Tuple of (workbook, filepath) or (None, None) on error
    """
    try:
        with LogContext("open_file_with_progress", filepath=filepath):
            progress_dialog.update_status("Opening file...")
            progress_dialog.update_progress(10, 100)
            
            # Check if operation was cancelled
            if progress_dialog.is_cancelled():
                return None, None
            
            # Simulate reading file metadata
            file_size = os.path.getsize(filepath) if os.path.exists(filepath) else 0
            logger.info("Opening Excel file", filepath=filepath, file_size_mb=file_size/1024/1024)
            
            progress_dialog.update_status("Loading workbook...")
            progress_dialog.update_progress(30, 100)
            
            # Check cancellation again
            if progress_dialog.is_cancelled():
                return None, None
            
            # Load the workbook (this is the slow part)
            workbook = openpyxl.load_workbook(filepath)
            
            progress_dialog.update_status("Analyzing worksheets...")
            progress_dialog.update_progress(70, 100)
            
            # Simulate analysis of worksheets
            worksheet_count = len(workbook.sheetnames)
            logger.info("Workbook loaded", worksheet_count=worksheet_count)
            
            # Check cancellation
            if progress_dialog.is_cancelled():
                workbook.close()
                return None, None
            
            progress_dialog.update_status("Finalizing...")
            progress_dialog.update_progress(100, 100)
            
            logger.info("File opened successfully", filepath=filepath)
            return workbook, filepath
            
    except Exception as e:
        logger.exception("Failed to open file", filepath=filepath)
        progress_dialog.update_status(f"Error: {str(e)}")
        time.sleep(1)  # Let user see the error
        raise


def _save_file_with_progress(progress_dialog, workbook, filepath):
    """
    Save file with progress feedback.
    
    Args:
        progress_dialog: Progress dialog instance
        workbook: Workbook to save
        filepath: Path to save to
        
    Returns:
        Filepath if successful, None otherwise
    """
    try:
        with LogContext("save_file_with_progress", filepath=filepath):
            progress_dialog.update_status("Preparing file for save...")
            progress_dialog.update_progress(10, 100)
            
            # Check cancellation
            if progress_dialog.is_cancelled():
                return None
            
            # Analyze workbook size for progress estimation
            worksheet_count = len(workbook.sheetnames)
            logger.info("Saving Excel file", filepath=filepath, worksheet_count=worksheet_count)
            
            progress_dialog.update_status("Validating data...")
            progress_dialog.update_progress(30, 100)
            
            # Check cancellation
            if progress_dialog.is_cancelled():
                return None
            
            progress_dialog.update_status("Writing file...")
            progress_dialog.update_progress(50, 100)
            
            # Save the workbook (this is the slow part)
            workbook.save(filepath)
            
            progress_dialog.update_status("Finalizing save...")
            progress_dialog.update_progress(90, 100)
            
            # Check cancellation (though save is complete)
            if progress_dialog.is_cancelled():
                return None
            
            progress_dialog.update_status("Save completed!")
            progress_dialog.update_progress(100, 100)
            
            logger.info("File saved successfully", filepath=filepath)
            return filepath
            
    except Exception as e:
        logger.exception("Failed to save file", filepath=filepath)
        progress_dialog.update_status(f"Save error: {str(e)}")
        time.sleep(1)  # Let user see the error
        raise


@exception_handler.handle_exceptions(show_dialog=True, log_error=True, return_value=(None, None))
def open_file(parent_window=None):
    """Open an Excel file and return the workbook with progress feedback."""
    filepath = filedialog.askopenfilename(
        title="Open Budgetinator File",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )

    if filepath:
        try:
            # Validate file path and content
            is_valid, error_msg = SecurityValidator.validate_excel_file(filepath)
            if not is_valid:
                messagebox.showerror("Security Error", f"Cannot open file: {error_msg}")
                logger.warning(f"Security validation failed for file: {filepath}", error=error_msg)
                return None, None
            
            # Sanitize file path
            safe_path = SecurityValidator.validate_file_path(filepath)
            
            if parent_window:
                # Use progress dialog for large files
                def completion_callback(result):
                    if result and result[0]:
                        messagebox.showinfo("Success", "File opened successfully!")
                        
                show_progress_for_operation(
                    parent_window,
                    lambda progress: _open_file_with_progress(progress, safe_path),
                    title="Opening File...",
                    can_cancel=True,
                    show_eta=True,
                    completion_callback=completion_callback
                )
                return None, None  # Result handled by callback
            else:
                # Fallback to direct loading
                workbook = openpyxl.load_workbook(safe_path)
                return workbook, safe_path
                
        except ValueError as e:
            messagebox.showerror("Security Error", str(e))
            logger.warning(f"Security validation error: {str(e)}")
            return None, None
            
    return None, None


@exception_handler.handle_exceptions(show_dialog=True, log_error=True, return_value=None)
def save_file(workbook, filepath=None, parent_window=None):
    """Save the workbook to a file with progress feedback."""
    if not filepath:
        filepath = filedialog.asksaveasfilename(
            title="Save Budgetinator File As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

    if filepath:
        try:
            # Validate and sanitize file path
            safe_path = SecurityValidator.validate_file_path(filepath)
            
            # Ensure proper extension
            if not safe_path.lower().endswith('.xlsx'):
                safe_path += '.xlsx'
            
            if parent_window:
                # Use progress dialog for save operation
                def completion_callback(result):
                    if result:
                        messagebox.showinfo("Success", f"File saved successfully to: {safe_path}")
                        
                show_progress_for_operation(
                    parent_window,
                    lambda progress: _save_file_with_progress(progress, workbook, safe_path),
                    title="Saving File...",
                    can_cancel=True,
                    show_eta=True,
                    completion_callback=completion_callback
                )
                return None  # Result handled by callback
            else:
                # Fallback to direct saving
                workbook.save(safe_path)
                messagebox.showinfo("Success", f"File saved successfully to: {safe_path}")
                return safe_path
                
        except ValueError as e:
            messagebox.showerror("Security Error", str(e))
            logger.warning(f"Security validation error during save: {str(e)}")
            return None
            
    return None


# Convenience functions for direct file operations (no progress)
@exception_handler.handle_exceptions(show_dialog=True, log_error=True, return_value=(None, None))
def open_file_direct():
    """Open an Excel file directly without progress feedback."""
    filepath = filedialog.askopenfilename(
        title="Open Budgetinator File",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )

    if filepath:
        try:
            # Validate file path and content
            is_valid, error_msg = SecurityValidator.validate_excel_file(filepath)
            if not is_valid:
                messagebox.showerror("Security Error", f"Cannot open file: {error_msg}")
                logger.warning(f"Security validation failed for file: {filepath}", error=error_msg)
                return None, None
            
            # Sanitize file path
            safe_path = SecurityValidator.validate_file_path(filepath)
            
            workbook = openpyxl.load_workbook(safe_path)
            return workbook, safe_path
            
        except ValueError as e:
            messagebox.showerror("Security Error", str(e))
            logger.warning(f"Security validation error: {str(e)}")
            return None, None
    return None, None


@exception_handler.handle_exceptions(show_dialog=True, log_error=True, return_value=None)
def save_file_direct(workbook, filepath=None):
    """Save the workbook to a file directly without progress feedback."""
    if not filepath:
        filepath = filedialog.asksaveasfilename(
            title="Save Budgetinator File As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

    if filepath:
        try:
            # Validate and sanitize file path
            safe_path = SecurityValidator.validate_file_path(filepath)
            
            # Ensure proper extension
            if not safe_path.lower().endswith('.xlsx'):
                safe_path += '.xlsx'
            
            workbook.save(safe_path)
            messagebox.showinfo("Success", f"File saved successfully to: {safe_path}")
            return safe_path
            
        except ValueError as e:
            messagebox.showerror("Security Error", str(e))
            logger.warning(f"Security validation error during save: {str(e)}")
            return None
    return None


@exception_handler.handle_exceptions(show_dialog=True, log_error=True, return_value=(None, None))
def new_file():
    """Create a new Excel workbook."""
    workbook = openpyxl.Workbook()
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])
    return workbook, None
