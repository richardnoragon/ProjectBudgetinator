"""
Tools and utilities for workpackage management.
"""
import openpyxl
from tkinter import messagebox, Toplevel, Label, Entry, Button, StringVar, Frame
import tkinter as tk
from datetime import datetime
from version import full_version_string


class WorkpackageDialog(Toplevel):
    """Dialog for entering workpackage details."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Add Workpackage Details")
        self.resizable(False, False)
        self.result = None
        self.vars = {}
        
        fields = [
            ("workpackage_number", "Workpackage Number"),
            ("workpackage_title", "Workpackage Title"),
            ("lead_partner", "Lead Partner"),
            ("start_month", "Start Month"),
            ("end_month", "End Month"),
            ("objectives", "Objectives"),
            ("description", "Description"),
            ("deliverables", "Deliverables")
        ]
        
        for key, label in fields:
            frame = Frame(self)
            frame.pack(fill="x", padx=8, pady=2)
            Label(frame, text=label+":").pack(side="left")
            var = StringVar()
            entry = Entry(frame, textvariable=var, width=32)
            entry.pack(side="right", expand=True, fill="x", padx=4)
            self.vars[key] = var

        btn_frame = Frame(self)
        btn_frame.pack(fill="x", padx=8, pady=8)
        Button(
            btn_frame,
            text="Commit",
            command=self.commit
        ).pack(side="left", expand=True, padx=4)
        Button(
            btn_frame,
            text="Cancel",
            command=self.cancel
        ).pack(side="left", expand=True, padx=4)
        
        self.protocol("WM_DELETE_WINDOW", self.cancel)
        self.grab_set()
        self.wait_window()

    def commit(self):
        """Save the dialog data and close."""
        self.result = {k: v.get() for k, v in self.vars.items()}
        self.destroy()

    def cancel(self):
        """Cancel the dialog without saving."""
        self.result = None
        self.destroy()


def add_workpackage_to_workbook(workbook, workpackage_info):
    """Add a workpackage worksheet to an Excel workbook."""
    sheet_name = f"WP{workpackage_info['workpackage_number']}"
    if sheet_name in workbook.sheetnames:
        messagebox.showerror("Error", "Worksheet already exists.")
        return False

    ws = workbook.create_sheet(title=sheet_name)
    
    # Map fields to cell ranges
    field_map = {
        "workpackage_number": "B2",
        "workpackage_title": "B3",
        "lead_partner": "B4",
        "start_month": "B5",
        "end_month": "B6",
        "objectives": "B7",
        "description": "B8",
        "deliverables": "B9"
    }
    
    # Set field labels and values
    for key, cell in field_map.items():
        label_cell = f"A{cell[1]}"
        ws[label_cell] = key.replace("_", " ").title() + ":"
        ws[cell] = workpackage_info.get(key, "")

    # Update version history
    update_version_history(
        workbook,
        f"Added workpackage: {sheet_name} - {workpackage_info['workpackage_title']}"
    )
    
    return True


def update_version_history(workbook, summary):
    """Update the version history sheet with a new entry."""
    version_sheet_name = "Version History"
    version_info = full_version_string()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    if version_sheet_name not in workbook.sheetnames:
        vh_ws = workbook.create_sheet(title=version_sheet_name)
        vh_ws.append(["Timestamp", "Version Info", "Summary"])
    else:
        vh_ws = workbook[version_sheet_name]
    
    vh_ws.append([timestamp, version_info, summary])
