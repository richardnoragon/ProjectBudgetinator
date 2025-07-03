"""
Partner management functions for working with Excel workbooks.
"""
import openpyxl
from tkinter import messagebox, simpledialog, Toplevel, Label, Entry, Button, StringVar
import tkinter as tk
from openpyxl.styles import PatternFill
import datetime

class PartnerDialog(Toplevel):
    """Dialog for entering partner details"""
    def __init__(self, master, partner_number, partner_acronym):
        super().__init__(master)
        self.title("Add Partner Details")
        self.resizable(False, False)
        self.result = None
        self.partner_number = partner_number
        self.partner_acronym = partner_acronym
        self.vars = {}
        
        # Debug output to verify initialization
        print(f"\nDEBUG - Initializing dialog with:")
        print(f"  Partner Number: {partner_number}")
        print(f"  Partner Acronym: {partner_acronym}")
        
        fields = [
            ("partner_identification_code", "Partner ID Code"),
            ("name_of_beneficiary", "Name of Beneficiary"),
            ("country", "Country"),
            ("role", "Role"),
            ("wp1", "WP1"),
            ("wp2", "WP2"),
            ("wp3", "WP3"),
            ("wp4", "WP4"),
            ("wp5", "WP5"),
            ("wp6", "WP6"),
            ("wp7", "WP7"),
            ("wp8", "WP8"),
            ("wp9", "WP9"),
            ("wp10", "WP10"),
            ("wp11", "WP11"),
            ("wp12", "WP12"),
            ("wp13", "WP13"),
            ("wp14", "WP14"),
            ("wp15", "WP15"),
            ("name_subcontractor_1", "Name Subcontrator 1"),
            ("sum_subcontractor_1", "Sum Subcontractor 1"),
            ("explanation_subcontractor_1", "Explanation of Subcontractor 1"),
            ("other_explanation_self", "Other Explanation Self")
        ]
        row = 0
        Label(self, text=f"Partner Number: {partner_number}").grid(
            row=row, column=0, sticky="w", padx=8, pady=2)
        row += 1
        Label(self, text=f"Partner Acronym: {partner_acronym}").grid(
            row=row, column=0, sticky="w", padx=8, pady=2)
        row += 1
        
        # Create and store fields
        for key, label in fields:
            Label(self, text=label+":").grid(row=row, column=0, sticky="w", padx=8, pady=2)
            var = StringVar()
            entry = Entry(self, textvariable=var, width=32)
            entry.grid(row=row, column=1, padx=8, pady=2)
            self.vars[key] = var
            row += 1
            
        btn_frame = tk.Frame(self)
        btn_frame.grid(row=row, column=0, columnspan=2, pady=8)
        Button(btn_frame, text="Commit", command=self.commit).pack(side="left", padx=8)
        Button(btn_frame, text="Cancel", command=self.cancel).pack(side="left", padx=8)
        self.protocol("WM_DELETE_WINDOW", self.cancel)
        self.grab_set()
        self.wait_window()

    def commit(self):
        try:
            # First show what values we have in the dialog
            wp_values = {}
            debug_msg = ["Current dialog values:"]
            
            for key, var in self.vars.items():
                value = var.get().strip()
                if key.startswith('wp'):
                    try:
                        wp_values[key] = float(value) if value else 0.0
                        debug_msg.append(f"{key} = {wp_values[key]}")
                    except ValueError:
                        wp_values[key] = 0.0
                        debug_msg.append(f"{key} = 0.0 (converted from '{value}')")
            
            # Show the values we collected
            messagebox.showinfo("Debug - Dialog Values", "\n".join(debug_msg))
            
            # Now create the result dictionary with all values
            self.result = {
                'project_partner_number': self.partner_number,
                'partner_acronym': self.partner_acronym,
                'partner_identification_code': self.vars['partner_identification_code'].get(),
                'name_of_beneficiary': self.vars['name_of_beneficiary'].get(),
                'country': self.vars['country'].get(),
                'role': self.vars['role'].get(),
            }
            
            # Add all WP values
            self.result.update(wp_values)
            
            # Show final result that will be saved
            debug_msg = ["Values being saved:"]
            for k, v in self.result.items():
                debug_msg.append(f"{k} = {v}")
            messagebox.showinfo("Debug - Final Values", "\n".join(debug_msg))
            
            self.destroy()
        except Exception as e:
            messagebox.showerror("Error", f"Error in commit: {str(e)}")

    def cancel(self):
        self.result = None
        self.destroy()


def add_partner_to_workbook(workbook, partner_info, dev_log=None):
    """Add a partner worksheet to an Excel workbook"""
    # Debug: Print all received partner info
    print("DEBUG - Received partner info:")
    for key, value in partner_info.items():
        print(f"  {key}: {value}")
        
    sheet_name = f"P{partner_info['project_partner_number']} {partner_info['partner_acronym']}"
    if sheet_name in workbook.sheetnames:
        messagebox.showerror("Error", "Worksheet already exists.")
        return False

    ws = workbook.create_sheet(title=sheet_name)
    
    # Map fields to cell ranges for merging and value insertion
    merge_map = {
        "project_partner_number": "D2:E2",
        "partner_identification_code": "D4:E4",
        "partner_acronym": "D3:E3",
        "name_of_beneficiary": "D13",
        "country": "D5:E5",
        "role": "D6:E6"
    }

    # Merge and set values for mapped fields
    debug_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    for key, cell_range in merge_map.items():
        value = partner_info.get(key, "")
        if ":" in cell_range:
            ws.merge_cells(cell_range)
            start_cell = cell_range.split(":")[0]
            # Fill merged range for debugging
            min_col = openpyxl.utils.cell.column_index_from_string(start_cell[0])
            min_row = int(start_cell[1:])
            end_cell = cell_range.split(":")[1]
            max_col = openpyxl.utils.cell.column_index_from_string(end_cell[0])
            max_row = int(end_cell[1:])
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).fill = debug_fill
        else:
            start_cell = cell_range
        ws[start_cell] = value

    # Process WP values
    wp_row = 18  # The row where WP values should go
    wp_header_row = 17  # The row for WP headers
    
    # Show the WP values we're about to process
    debug_msg = ["WP Values to process:"]
    for key in partner_info:
        if key.startswith('wp'):
            value = partner_info[key]
            debug_msg.append(f"{key} = {value} (type: {type(value).__name__})")
    messagebox.showinfo("Debug - WP Values", "\n".join(debug_msg))
    
    # Map WP fields to Excel columns (WP1->C, WP2->D, etc)
    wp_fields = {
        'wp1': 'C18', 'wp2': 'D18', 'wp3': 'E18', 'wp4': 'F18', 'wp5': 'G18',
        'wp6': 'H18', 'wp7': 'I18', 'wp8': 'J18', 'wp9': 'K18', 'wp10': 'L18',
        'wp11': 'M18', 'wp12': 'N18', 'wp13': 'O18', 'wp14': 'P18', 'wp15': 'Q18'
    }
    
    # First write the headers (row 17)
    for wp_key, cell_ref in wp_fields.items():
        # Write header
        header_cell = f"{cell_ref[0]}{wp_header_row}"  # Use first char of cell_ref (column) + header row
        ws[header_cell] = wp_key.upper()  # wp1 -> WP1
        
        # Get and convert the value
        raw_value = partner_info.get(wp_key, "0")
        try:
            if isinstance(raw_value, (int, float)):
                value = float(raw_value)
            else:
                value = float(str(raw_value).strip() or 0)
        except (ValueError, TypeError):
            if dev_log:
                dev_log(f"Warning: Could not convert {wp_key}='{raw_value}' to number, using 0")
            value = 0.0
            
        # Write the value
        ws[cell_ref] = value
        cell = ws[cell_ref]
        cell.number_format = '#,##0.00'
        
        if dev_log:
            dev_log(f"Set {cell_ref} = {value} for {wp_key} (raw_value='{raw_value}')")

    # Update version history
    update_version_history(workbook, f"Added partner: {sheet_name}")
    
    return True

def update_version_history(workbook, summary):
    """Update the version history sheet with a new entry"""
    from version import full_version_string
    version_sheet_name = "Version History"
    version_info = full_version_string()
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    if version_sheet_name not in workbook.sheetnames:
        vh_ws = workbook.create_sheet(title=version_sheet_name)
        vh_ws.append(["Timestamp", "Version Info", "Summary"])
    else:
        vh_ws = workbook[version_sheet_name]
    
    vh_ws.append([timestamp, version_info, summary])
