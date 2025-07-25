"""
Excel file management utilities with performance optimizations and memory management.

This module provides optimized Excel file handling with lazy loading,
memory management, and comprehensive resource cleanup.
"""

import os
import gc
import psutil
import logging
import threading
import time
from typing import Optional, List, Dict, Any, Iterator, TYPE_CHECKING
from contextlib import contextmanager
from pathlib import Path

# Conditional imports for type hints
if TYPE_CHECKING:
    from openpyxl import load_workbook
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
else:
    # Runtime fallbacks when openpyxl is not available
    try:
        from openpyxl import load_workbook
        from openpyxl.workbook.workbook import Workbook
        from openpyxl.worksheet.worksheet import Worksheet
    except ImportError:
        load_workbook = None
        Workbook = None
        Worksheet = None

# Import performance monitoring
from utils.performance_monitor import monitor_file_operation
from utils.security_validator import SecurityValidator

logger = logging.getLogger(__name__)


class ExcelMemoryTracker:
    """Enhanced memory tracker with thresholds and automatic cleanup."""
    
    def __init__(self, memory_threshold_mb: float = 500.0):
        self.open_workbooks = {}  # file_path -> {'workbook': wb, 'size_mb': size, 'opened_at': timestamp}
        self.memory_threshold_mb = memory_threshold_mb
        self.peak_memory_mb = 0.0
        self.cleanup_callbacks = []
        self._lock = threading.Lock()
    
    def track_workbook(self, file_path: str, workbook: "Workbook"):
        """Track an opened workbook with enhanced metadata."""
        with self._lock:
            current_memory = self._get_memory_usage()
            file_size = self._estimate_workbook_size(file_path)
            
            self.open_workbooks[file_path] = {
                'workbook': workbook,
                'size_mb': file_size,
                'opened_at': time.time(),
                'memory_at_open': current_memory
            }
            
            # Update peak memory
            if current_memory > self.peak_memory_mb:
                self.peak_memory_mb = current_memory
            
            logger.debug(f"Opened workbook: {file_path}, Size: {file_size:.2f}MB, Memory: {current_memory:.2f}MB")
            
            # Check if we need automatic cleanup
            self._check_memory_threshold()
    
    def untrack_workbook(self, file_path: str):
        """Untrack a closed workbook."""
        with self._lock:
            if file_path in self.open_workbooks:
                workbook_info = self.open_workbooks.pop(file_path)
                memory = self._get_memory_usage()
                size_mb = workbook_info.get('size_mb', 0)
                logger.debug(f"Closed workbook: {file_path}, Size: {size_mb:.2f}MB, Memory: {memory:.2f}MB")
    
    def _estimate_workbook_size(self, file_path: str) -> float:
        """Estimate workbook memory footprint."""
        try:
            file_size = os.path.getsize(file_path) / (1024 * 1024)
            # Excel files typically expand 2-4x in memory
            return file_size * 3.0
        except Exception:
            return 0.0
    
    def _get_memory_usage(self) -> float:
        """Get current memory usage in MB."""
        try:
            process = psutil.Process()
            return process.memory_info().rss / (1024 * 1024)
        except Exception:
            return 0.0
    
    def _check_memory_threshold(self):
        """Check if memory threshold is exceeded and trigger cleanup if needed."""
        current_memory = self._get_memory_usage()
        if current_memory > self.memory_threshold_mb:
            logger.warning(f"Memory threshold exceeded: {current_memory:.2f}MB > {self.memory_threshold_mb:.2f}MB")
            self._trigger_automatic_cleanup()
    
    def _trigger_automatic_cleanup(self):
        """Trigger automatic cleanup of oldest workbooks."""
        if not self.open_workbooks:
            return
        
        # Sort by opened time (oldest first)
        sorted_workbooks = sorted(
            self.open_workbooks.items(),
            key=lambda x: x[1]['opened_at']
        )
        
        # Close oldest workbook
        oldest_path, oldest_info = sorted_workbooks[0]
        logger.info(f"Auto-cleanup: closing oldest workbook {oldest_path}")
        
        try:
            oldest_info['workbook'].close()
            self.untrack_workbook(oldest_path)
            gc.collect()
        except Exception as e:
            logger.error(f"Error during auto-cleanup of {oldest_path}: {e}")
        
        # Run cleanup callbacks
        for callback in self.cleanup_callbacks:
            try:
                callback()
            except Exception as e:
                logger.error(f"Error in cleanup callback: {e}")
    
    def register_cleanup_callback(self, callback):
        """Register a callback to be called during automatic cleanup."""
        self.cleanup_callbacks.append(callback)
    
    def get_stats(self) -> Dict[str, Any]:
        """Get comprehensive memory tracking statistics."""
        with self._lock:
            current_memory = self._get_memory_usage()
            total_estimated_size = sum(info['size_mb'] for info in self.open_workbooks.values())
            
            return {
                'open_workbooks': len(self.open_workbooks),
                'open_files': list(self.open_workbooks.keys()),
                'current_memory_mb': current_memory,
                'peak_memory_mb': self.peak_memory_mb,
                'memory_threshold_mb': self.memory_threshold_mb,
                'total_estimated_size_mb': total_estimated_size,
                'memory_efficiency': (total_estimated_size / current_memory * 100) if current_memory > 0 else 0,
                'oldest_workbook': min(self.open_workbooks.items(),
                                       key=lambda x: x[1]['opened_at'])[0] if self.open_workbooks else None
            }
    
    def force_cleanup_all(self):
        """Force cleanup of all tracked workbooks."""
        with self._lock:
            for file_path, workbook_info in list(self.open_workbooks.items()):
                try:
                    workbook_info['workbook'].close()
                    logger.debug(f"Force closed workbook: {file_path}")
                except Exception as e:
                    logger.error(f"Error force closing {file_path}: {e}")
            
            self.open_workbooks.clear()
            gc.collect()
            logger.info("Force cleanup completed for all workbooks")


# Global memory tracker
memory_tracker = ExcelMemoryTracker()


class ExcelManager:
    """
    Optimized Excel file manager with lazy loading and comprehensive memory management.
    
    Features:
    - Lazy loading of workbooks to reduce memory usage
    - Read-only mode for large files
    - Automatic resource cleanup with context managers
    - Memory usage tracking and monitoring
    - Force garbage collection on close
    """
    
    def __init__(self, file_path: str):
        """
        Initialize ExcelManager with file path.
        
        Args:
            file_path: Path to the Excel file
        """
        # Validate file path
        try:
            self.file_path = SecurityValidator.validate_file_path(file_path)
        except ValueError as e:
            raise ValueError(f"Invalid file path: {str(e)}")
        
        self._workbook: Optional["Workbook"] = None
        self._active_sheet: Optional["Worksheet"] = None
        self._is_open = False
        self._memory_usage = 0
        
        # Validate Excel file
        is_valid, error_msg = SecurityValidator.validate_excel_file(self.file_path)
        if not is_valid:
            raise ValueError(f"Invalid Excel file: {error_msg}")
    
    @property
    def workbook(self) -> "Workbook":
        """
        Lazy-loaded workbook property with memory tracking.
        
        Returns:
            Workbook: The Excel workbook instance
        """
        if self._workbook is None:
            self._load_workbook()
        if self._workbook is None:
            raise RuntimeError("Failed to load workbook")
        return self._workbook
    
    @property
    def active_sheet(self) -> Optional["Worksheet"]:
        """
        Get the active worksheet.
        
        Returns:
            Worksheet: The active worksheet or None if not available
        """
        if self._active_sheet is None and self.workbook:
            self._active_sheet = self.workbook.active
        return self._active_sheet
    
    def _load_workbook(self) -> None:
        """
        Load workbook with adaptive settings and memory tracking.
        
        Uses adaptive loading strategy based on file size and system memory.
        """
        try:
            logger.info(f"Loading Excel file: {self.file_path}")
            
            # Get file size and system memory to determine loading strategy
            file_size = os.path.getsize(self.file_path)
            file_size_mb = file_size / (1024 * 1024)
            
            # Get available system memory
            available_memory = psutil.virtual_memory().available / (1024 * 1024)
            
            # Adaptive loading strategy
            use_read_only = self._should_use_read_only(file_size_mb, available_memory)
            use_data_only = self._should_use_data_only(file_size_mb)
            
            logger.info(f"Loading strategy: file_size={file_size_mb:.1f}MB, "
                        f"available_memory={available_memory:.1f}MB, "
                        f"read_only={use_read_only}, data_only={use_data_only}")
            
            # Check if we need to free memory before loading
            if file_size_mb > 50:  # For files larger than 50MB
                self._prepare_memory_for_large_file()
            
            self._workbook = load_workbook(
                self.file_path,
                read_only=use_read_only,
                data_only=use_data_only,
                keep_links=False,  # Don't load external links
                keep_vba=False     # Don't load VBA macros to save memory
            )
            
            self._is_open = True
            memory_tracker.track_workbook(self.file_path, self._workbook)
            
            # Log memory usage after loading
            current_memory = memory_tracker._get_memory_usage()
            logger.info(f"Successfully loaded workbook: {self.file_path}, "
                        f"Memory usage: {current_memory:.2f}MB")
            
        except Exception as e:
            logger.error(f"Failed to load Excel file {self.file_path}: {str(e)}")
            raise
    
    def _should_use_read_only(self, file_size_mb: float, available_memory_mb: float) -> bool:
        """Determine if read-only mode should be used based on file size and available memory."""
        # Use read-only for files larger than 10MB or when available memory is low
        return file_size_mb > 10 or available_memory_mb < 500
    
    def _should_use_data_only(self, file_size_mb: float) -> bool:
        """Determine if data_only mode should be used."""
        # Use data_only for files larger than 5MB to avoid formula calculation overhead
        return file_size_mb > 5
    
    def _prepare_memory_for_large_file(self) -> None:
        """Prepare system memory for loading a large file."""
        logger.info("Preparing memory for large file loading...")
        
        # Force garbage collection
        gc.collect()
        
        # Check if memory tracker needs cleanup
        current_memory = memory_tracker._get_memory_usage()
        if current_memory > memory_tracker.memory_threshold_mb * 0.8:
            logger.info("Triggering preventive memory cleanup before large file load")
            memory_tracker._trigger_automatic_cleanup()
    
    def close(self) -> None:
        """
        Close the workbook and release resources with memory cleanup.
        
        This method ensures proper cleanup of Excel resources to prevent
        memory leaks, including garbage collection and memory tracking.
        """
        if self._workbook is not None:
            try:
                # Close the workbook
                self._workbook.close()
                memory_tracker.untrack_workbook(self.file_path)
                
                # Force garbage collection
                gc.collect()
                
                logger.info(f"Closed workbook: {self.file_path}")
                
            except Exception as e:
                logger.warning(f"Error closing workbook: {str(e)}")
            finally:
                self._workbook = None
                self._active_sheet = None
                self._is_open = False
                
                # Additional cleanup
                gc.collect()
    
    def force_close(self) -> None:
        """
        Force close workbook and perform aggressive memory cleanup.
        
        Use this method when you need to ensure complete cleanup,
        such as before processing large files or when memory is constrained.
        """
        self.close()
        
        # Additional memory cleanup
        gc.collect()
        
        # Log memory usage
        memory = memory_tracker._get_memory_usage()
        logger.info(f"Memory after force close: {memory:.2f}MB")
    
    def get_memory_usage(self) -> Dict[str, Any]:
        """
        Get current memory usage information.
        
        Returns:
            Dict[str, Any]: Memory usage statistics
        """
        return {
            'file_path': self.file_path,
            'is_open': self._is_open,
            'memory_tracker': memory_tracker.get_stats()
        }


@contextmanager
def excel_context(file_path: str, **kwargs) -> Iterator["Workbook"]:
    """
    Context manager for safe Excel file handling with automatic cleanup.
    
    Args:
        file_path: Path to the Excel file
        **kwargs: Additional arguments for load_workbook
    
    Yields:
        Workbook: The Excel workbook instance
    
    Example:
        >>> with excel_context('file.xlsx') as wb:
        ...     sheet = wb.active
        ...     # Process workbook
        ...     # Auto-cleanup on exit
    """
    workbook = None
    try:
        # Apply optimized settings
        kwargs.setdefault('data_only', True)
        kwargs.setdefault('keep_links', False)
        
        # Use read-only for large files
        file_size = os.path.getsize(file_path)
        if file_size > 10 * 1024 * 1024:
            kwargs.setdefault('read_only', True)
        
        # Monitor file operation
        with monitor_file_operation('open', file_path):
            workbook = load_workbook(file_path, **kwargs)
            memory_tracker.track_workbook(file_path, workbook)
        
        logger.debug(f"Opened workbook in context: {file_path}")
        yield workbook
        
    except Exception as e:
        logger.error(f"Error in excel_context: {str(e)}")
        raise
    finally:
        if workbook:
            try:
                with monitor_file_operation('close', file_path):
                    workbook.close()
                    memory_tracker.untrack_workbook(file_path)
                    gc.collect()
                logger.debug(f"Closed workbook in context: {file_path}")
            except Exception as e:
                logger.warning(f"Error closing workbook in context: {str(e)}")


@contextmanager
def excel_safe_context(file_path: str, **kwargs) -> Iterator[ExcelManager]:
    """
    Context manager for safe ExcelManager usage with full cleanup.
    
    Args:
        file_path: Path to the Excel file
        **kwargs: Additional arguments for ExcelManager
    
    Yields:
        ExcelManager: The Excel manager instance
    
    Example:
        >>> with excel_safe_context('file.xlsx') as excel:
        ...     sheet = excel.active_sheet
        ...     # Process using ExcelManager
        ...     # Full cleanup on exit
    """
    excel_manager = None
    try:
        excel_manager = ExcelManager(file_path)
        yield excel_manager
    except Exception as e:
        logger.error(f"Error in excel_safe_context: {str(e)}")
        raise
    finally:
        if excel_manager:
            excel_manager.force_close()


class ResourceManager:
    """
    Centralized resource management for Excel operations.
    
    Tracks and manages all Excel resources to prevent memory leaks.
    """
    
    def __init__(self):
        self.active_managers: Dict[str, ExcelManager] = {}
        self.cleanup_callbacks = []
    
    def get_manager(self, file_path: str) -> ExcelManager:
        """Get or create ExcelManager for file."""
        if file_path not in self.active_managers:
            self.active_managers[file_path] = ExcelManager(file_path)
        return self.active_managers[file_path]
    
    def release_manager(self, file_path: str) -> None:
        """Release and cleanup ExcelManager for file."""
        if file_path in self.active_managers:
            self.active_managers[file_path].force_close()
            del self.active_managers[file_path]
    
    def cleanup_all(self) -> None:
        """Force cleanup of all managed resources."""
        for file_path in list(self.active_managers.keys()):
            self.release_manager(file_path)
        
        # Force garbage collection
        gc.collect()
        
        # Log final memory state
        memory = memory_tracker._get_memory_usage()
        logger.info(f"Final memory after cleanup: {memory:.2f}MB")
    
    def register_cleanup_callback(self, callback):
        """Register callback for cleanup events."""
        self.cleanup_callbacks.append(callback)
    
    def get_resource_stats(self) -> Dict[str, Any]:
        """Get resource management statistics."""
        return {
            'active_managers': len(self.active_managers),
            'active_files': list(self.active_managers.keys()),
            'memory_tracker': memory_tracker.get_stats()
        }


# Global resource manager
resource_manager = ResourceManager()


# Enhanced context managers for specific use cases

@contextmanager
def read_only_excel_context(file_path: str) -> Iterator["Workbook"]:
    """Context manager for read-only Excel access."""
    with excel_context(file_path, read_only=True, data_only=True) as wb:
        yield wb


@contextmanager
def write_excel_context(file_path: str) -> Iterator["Workbook"]:
    """Context manager for write operations with full features."""
    with excel_context(file_path, read_only=False, data_only=False) as wb:
        yield wb


@contextmanager
def memory_safe_bulk_operation(file_paths: List[str]) -> Iterator[Dict[str, "Workbook"]]:
    """
    Context manager for bulk Excel operations with memory management.
    
    Args:
        file_paths: List of Excel file paths to process
    
    Yields:
        Dict[str, Workbook]: Dictionary mapping file paths to workbooks
    
    Example:
        >>> files = ['file1.xlsx', 'file2.xlsx']
        >>> with memory_safe_bulk_operation(files) as workbooks:
        ...     for path, wb in workbooks.items():
        ...         # Process each workbook
        ...         # Auto-cleanup all workbooks on exit
    """
    workbooks = {}
    try:
        for file_path in file_paths:
            if os.path.exists(file_path):
                wb = load_workbook(file_path, read_only=True, data_only=True)
                workbooks[file_path] = wb
                memory_tracker.track_workbook(file_path, wb)
        
        logger.info(f"Opened {len(workbooks)} workbooks for bulk operation")
        yield workbooks
        
    except Exception as e:
        logger.error(f"Error in bulk operation: {str(e)}")
        raise
    finally:
        # Close all workbooks
        for file_path, wb in workbooks.items():
            try:
                wb.close()
                memory_tracker.untrack_workbook(file_path)
            except Exception as e:
                logger.warning(f"Error closing {file_path}: {str(e)}")
        
        # Force garbage collection
        gc.collect()
        logger.info("Completed bulk operation cleanup")


# Utility functions for memory management

def get_memory_usage() -> Dict[str, float]:
    """Get comprehensive memory usage information."""
    try:
        process = psutil.Process()
        memory_info = process.memory_info()
        
        return {
            'rss_mb': memory_info.rss / (1024 * 1024),
            'vms_mb': memory_info.vms / (1024 * 1024),
            'percent': process.memory_percent(),
            'open_workbooks': len(memory_tracker.open_workbooks)
        }
    except Exception as e:
        logger.error(f"Error getting memory usage: {str(e)}")
        return {'rss_mb': 0, 'vms_mb': 0, 'percent': 0, 'open_workbooks': 0}


def force_memory_cleanup() -> None:
    """Force comprehensive memory cleanup."""
    logger.info("Starting comprehensive memory cleanup...")
    
    # Close all managed resources
    resource_manager.cleanup_all()
    
    # Force garbage collection multiple times
    for i in range(3):
        collected = gc.collect()
        logger.debug(f"GC cycle {i + 1}: collected {collected} objects")
    
    # Log final state
    memory = get_memory_usage()
    logger.info(f"Memory cleanup completed: {memory['rss_mb']:.2f}MB used")


def monitor_memory_usage(func):
    """Decorator to monitor memory usage of Excel operations."""
    def wrapper(*args, **kwargs):
        start_memory = get_memory_usage()
        logger.info(f"Memory before {func.__name__}: {start_memory['rss_mb']:.2f}MB")
        
        try:
            result = func(*args, **kwargs)
            return result
        finally:
            end_memory = get_memory_usage()
            delta = end_memory['rss_mb'] - start_memory['rss_mb']
            logger.info(f"Memory after {func.__name__}: {end_memory['rss_mb']:.2f}MB (Δ{delta:+.2f}MB)")
    
    return wrapper
