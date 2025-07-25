"""
Workbook utility functions for ProjectBudgetinator.

This module provides centralized workbook loading, validation, and management
functionality to eliminate code duplication across the application.
"""

import os
import logging
from typing import Optional, Tuple, Any
from tkinter import messagebox, filedialog

from utils.security_validator import SecurityValidator
from logger import get_structured_logger, LogContext

# Type hints for openpyxl (using forward references)
try:
    from openpyxl import load_workbook
    from openpyxl.workbook.workbook import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    load_workbook = None
    Workbook = None
    OPENPYXL_AVAILABLE = False

# Constants
EXCEL_FILETYPES = [("Excel files", "*.xlsx;*.xls"), ("All files", "*.*")]

logger = get_structured_logger("ProjectBudgetinator.workbook_utils")


class WorkbookLoadResult:
    """Result object for workbook loading operations."""
    
    def __init__(self, success: bool, workbook: Optional[Any] = None, 
                 file_path: Optional[str] = None, error_message: Optional[str] = None):
        self.success = success
        self.workbook = workbook
        self.file_path = file_path
        self.error_message = error_message


def validate_and_load_workbook(file_path: str) -> WorkbookLoadResult:
    """
    Validate and load an Excel workbook with comprehensive error handling.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        WorkbookLoadResult: Result object containing success status, workbook, and error info
    """
    try:
        # Validate file path and content
        is_valid, error_msg = SecurityValidator.validate_excel_file(file_path)
        if not is_valid:
            logger.warning("Security validation failed for file",
                           file_path=file_path, error=error_msg)
            return WorkbookLoadResult(
                success=False,
                error_message=f"Security validation failed: {error_msg}"
            )
        
        # Sanitize file path
        safe_path = SecurityValidator.validate_file_path(file_path)
        
        # Check if openpyxl is available
        if not OPENPYXL_AVAILABLE or load_workbook is None:
            return WorkbookLoadResult(
                success=False,
                error_message="openpyxl library is not available. Please install it with: pip install openpyxl"
            )
        
        # Load workbook
        workbook = load_workbook(safe_path)
        logger.info("Workbook loaded successfully", file_path=safe_path)
        
        return WorkbookLoadResult(
            success=True,
            workbook=workbook,
            file_path=safe_path
        )
        
    except ValueError as e:
        error_msg = f"Security validation error: {str(e)}"
        logger.warning("Security validation error", error=str(e))
        return WorkbookLoadResult(success=False, error_message=error_msg)
    except Exception as e:
        error_msg = f"Failed to load workbook: {str(e)}"
        logger.error("Failed to load workbook", file_path=file_path, error=str(e))
        return WorkbookLoadResult(success=False, error_message=error_msg)


def prompt_for_workbook_if_needed(current_workbook: Optional[Any], 
                                  operation_name: str = "operation") -> WorkbookLoadResult:
    """
    Prompt user to open a workbook if none is currently loaded.
    
    Args:
        current_workbook: Currently loaded workbook (None if not loaded)
        operation_name: Name of the operation requiring a workbook
        
    Returns:
        WorkbookLoadResult: Result object with loaded workbook or error
    """
    if current_workbook is not None:
        return WorkbookLoadResult(
            success=True,
            workbook=current_workbook,
            file_path="current_workbook"
        )
    
    logger.warning(f"No workbook open for {operation_name}, prompting user to open one")
    response = messagebox.askyesno(
        "No Workbook Open",
        f"No workbook is currently open for {operation_name}. Would you like to open one now?"
    )
    
    if not response:
        logger.info(f"User declined to open workbook for {operation_name}")
        return WorkbookLoadResult(
            success=False,
            error_message="User declined to open workbook"
        )
    
    file_path = filedialog.askopenfilename(
        title="Open Excel Workbook",
        filetypes=EXCEL_FILETYPES
    )
    
    if not file_path:
        logger.info(f"User cancelled workbook selection for {operation_name}")
        return WorkbookLoadResult(
            success=False,
            error_message="User cancelled workbook selection"
        )
    
    return validate_and_load_workbook(file_path)


def save_workbook_with_dialog(workbook: Any, title: str = "Save Workbook",
                              success_message: str = "Workbook saved successfully") -> Tuple[bool, Optional[str]]:
    """
    Save workbook with user dialog for file selection.
    
    Args:
        workbook: Workbook object to save
        title: Dialog title
        success_message: Message to show on successful save
        
    Returns:
        Tuple[bool, Optional[str]]: (success, file_path)
    """
    try:
        save_path = filedialog.asksaveasfilename(
            title=title,
            defaultextension=".xlsx",
            filetypes=EXCEL_FILETYPES
        )
        
        if not save_path:
            logger.info("User cancelled save operation")
            messagebox.showwarning(
                "Warning",
                "Operation completed but workbook not saved!"
            )
            return False, None
        
        workbook.save(save_path)
        messagebox.showinfo("Success", f"{success_message}\n\nWorkbook saved to: {save_path}")
        logger.info("Workbook saved successfully", file_path=save_path)
        return True, save_path
        
    except Exception as e:
        error_msg = f"Failed to save workbook: {str(e)}"
        logger.error("Failed to save workbook", error=str(e))
        messagebox.showerror("Error", error_msg)
        return False, None


def ensure_workbook_loaded(current_workbook: Optional[Any],
                           operation_name: str) -> Tuple[bool, Optional[Any], Optional[str]]:
    """
    Ensure a workbook is loaded for an operation, prompting user if needed.
    
    Args:
        current_workbook: Currently loaded workbook
        operation_name: Name of the operation requiring a workbook
        
    Returns:
        Tuple[bool, Optional[Any], Optional[str]]: (success, workbook, error_message)
    """
    result = prompt_for_workbook_if_needed(current_workbook, operation_name)
    
    if not result.success:
        if result.error_message and "declined" not in result.error_message and "cancelled" not in result.error_message:
            messagebox.showerror("Error", result.error_message)
        return False, None, result.error_message
    
    return True, result.workbook, None


def validate_worksheet_exists(workbook: Any, worksheet_name: str) -> bool:
    """
    Validate that a specific worksheet exists in the workbook.
    
    Args:
        workbook: Workbook object to check
        worksheet_name: Name of the worksheet to validate
        
    Returns:
        bool: True if worksheet exists, False otherwise
    """
    try:
        if worksheet_name not in workbook.sheetnames:
            logger.warning(f"Worksheet '{worksheet_name}' not found in workbook")
            messagebox.showerror(
                "Worksheet Not Found",
                f"The '{worksheet_name}' worksheet was not found in the workbook.\n\n"
                f"Available worksheets: {', '.join(workbook.sheetnames)}"
            )
            return False
        return True
    except Exception as e:
        logger.error(f"Error validating worksheet existence: {e}")
        messagebox.showerror("Error", f"Failed to validate worksheet: {str(e)}")
        return False


class WorkbookOperationContext:
    """Context manager for workbook operations with automatic error handling."""
    
    def __init__(self, operation_name: str, current_workbook: Optional[Any] = None):
        self.operation_name = operation_name
        self.current_workbook = current_workbook
        self.workbook = None
        self.success = False
        self.error_message = None
        
    def __enter__(self):
        """Enter the context and ensure workbook is loaded."""
        with LogContext(f"workbook_operation_{self.operation_name.lower().replace(' ', '_')}"):
            logger.info(f"Starting {self.operation_name}")
            
            success, workbook, error_message = ensure_workbook_loaded(
                self.current_workbook, self.operation_name
            )
            
            if not success:
                self.error_message = error_message
                return self
            
            self.workbook = workbook
            self.success = True
            return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Exit the context and handle any exceptions."""
        if exc_type is not None:
            logger.exception(f"Exception during {self.operation_name}")
            messagebox.showerror(
                "Error",
                f"An error occurred during {self.operation_name}:\n{str(exc_val)}"
            )
        return False  # Don't suppress exceptions


def create_workbook_operation_context(operation_name: str,
                                      current_workbook: Optional[Any] = None) -> WorkbookOperationContext:
    """
    Create a context manager for workbook operations.
    
    Args:
        operation_name: Name of the operation
        current_workbook: Currently loaded workbook
        
    Returns:
        WorkbookOperationContext: Context manager for the operation
    """
    return WorkbookOperationContext(operation_name, current_workbook)


# Utility functions for common workbook operations

def get_partner_worksheets(workbook: Any) -> list:
    """
    Get list of partner worksheets (P2-P20) from workbook.
    
    Args:
        workbook: Workbook object to analyze
        
    Returns:
        list: List of partner worksheet names
    """
    partner_sheets = []
    for sheet_name in workbook.sheetnames:
        if is_partner_worksheet(sheet_name):
            partner_sheets.append(sheet_name)
    return partner_sheets


def is_partner_worksheet(sheet_name: str) -> bool:
    """
    Check if a sheet name represents a partner worksheet (P2 through P20).
    
    Args:
        sheet_name: Name of the worksheet
        
    Returns:
        bool: True if it's a partner worksheet, False otherwise
    """
    if not sheet_name.startswith('P'):
        return False
    
    # Extract the part after 'P'
    rest = sheet_name[1:]
    if not rest:
        return False
    
    # Get the number part (before any hyphen or other characters)
    parts = rest.split('-')
    if not parts:
        return False
    
    number_part = parts[0]
    
    # Check if it's a valid number between 2 and 20
    try:
        partner_number = int(number_part)
        return 2 <= partner_number <= 20
    except ValueError:
        return False


def extract_partner_number_from_sheet_name(sheet_name: str) -> Optional[int]:
    """
    Extract partner number from sheet name.
    
    Args:
        sheet_name: Sheet name like "P2-ACME" or "P3-University"
        
    Returns:
        Optional[int]: Partner number or None if invalid
    """
    if not sheet_name.startswith('P'):
        return None
    
    try:
        # Extract the part after 'P' and before the first hyphen
        parts = sheet_name[1:].split('-', 1)
        if parts and parts[0].isdigit():
            partner_num = int(parts[0])
            # Only accept partners 2-20
            if 2 <= partner_num <= 20:
                return partner_num
    except (ValueError, IndexError):
        pass
    
    return None