"""
Advanced Caching utilities for ProjectBudgetinator.

This module provides intelligent caching for Excel operations, file validation,
and frequently accessed data to improve application performance with advanced
caching strategies, memory-efficient operations, and comprehensive performance tracking.
"""

import os
import hashlib
import json
import time
import threading
import weakref
import pickle
import gzip
from functools import lru_cache, wraps
from typing import Dict, List, Any, Optional, Tuple, Union, Callable
from collections import OrderedDict, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


@dataclass
class CacheEntry:
    """Enhanced cache entry with metadata and performance tracking."""
    data: Any
    timestamp: float
    ttl: float
    access_count: int = 0
    last_accessed: float = field(default_factory=time.time)
    size_bytes: int = 0
    compression_ratio: float = 1.0
    creation_time: float = field(default_factory=time.time)
    
    def is_expired(self) -> bool:
        """Check if cache entry is expired."""
        return time.time() - self.timestamp > self.ttl
    
    def access(self) -> Any:
        """Access the cached data and update access statistics."""
        self.access_count += 1
        self.last_accessed = time.time()
        return self.data
    
    def get_age_seconds(self) -> float:
        """Get age of cache entry in seconds."""
        return time.time() - self.timestamp


@dataclass
class CacheStats:
    """Comprehensive cache statistics."""
    hits: int = 0
    misses: int = 0
    evictions: int = 0
    total_size_bytes: int = 0
    entry_count: int = 0
    avg_access_time_ms: float = 0.0
    hit_rate: float = 0.0
    memory_efficiency: float = 0.0
    compression_savings: float = 0.0


class AdvancedCacheManager:
    """
    Advanced centralized caching system for ProjectBudgetinator.
    
    Features:
    - Multi-level caching (memory + disk)
    - LRU and LFU eviction policies
    - Compression for large data
    - Thread-safe operations
    - Performance monitoring
    - Cache warming strategies
    - Intelligent prefetching
    """
    
    def __init__(self, cache_dir: Optional[str] = None,
                 max_memory_size: int = 100 * 1024 * 1024,  # 100MB
                 max_disk_size: int = 500 * 1024 * 1024,    # 500MB
                 compression_threshold: int = 1024,          # 1KB
                 enable_compression: bool = True):
        """
        Initialize Advanced CacheManager.
        
        Args:
            cache_dir: Directory for persistent cache storage
            max_memory_size: Maximum memory cache size in bytes
            max_disk_size: Maximum disk cache size in bytes
            compression_threshold: Minimum size for compression
            enable_compression: Whether to enable data compression
        """
        self.cache_dir = cache_dir or os.path.join(
            str(Path.home()), "ProjectBudgetinator", "cache"
        )
        os.makedirs(self.cache_dir, exist_ok=True)
        
        # Configuration
        self.max_memory_size = max_memory_size
        self.max_disk_size = max_disk_size
        self.compression_threshold = compression_threshold
        self.enable_compression = enable_compression
        
        # Multi-level cache storage
        self.memory_cache: OrderedDict[str, CacheEntry] = OrderedDict()
        self.disk_cache_index: Dict[str, Dict[str, Any]] = {}
        
        # Thread safety
        self._lock = threading.RLock()
        
        # Statistics and monitoring
        self.stats = CacheStats()
        self.access_times: Dict[str, List[float]] = defaultdict(list)
        self.prefetch_queue: List[str] = []
        
        # Performance tracking
        self.operation_times: Dict[str, float] = {}
        self.cache_efficiency: Dict[str, float] = {}
        
        # Load existing disk cache index
        self._load_disk_cache_index()
        
        # Start background maintenance
        self._start_maintenance_thread()
    
    def _start_maintenance_thread(self):
        """Start background thread for cache maintenance."""
        def maintenance_worker():
            while True:
                try:
                    time.sleep(60)  # Run every minute
                    self._cleanup_expired_entries()
                    self._optimize_cache_layout()
                    self._update_statistics()
                except Exception as e:
                    logger.error(f"Cache maintenance error: {e}")
        
        maintenance_thread = threading.Thread(
            target=maintenance_worker,
            daemon=True,
            name="CacheMaintenanceWorker"
        )
        maintenance_thread.start()
    
    def _load_disk_cache_index(self):
        """Load disk cache index from persistent storage."""
        index_file = os.path.join(self.cache_dir, "cache_index.json")
        try:
            if os.path.exists(index_file):
                with open(index_file, 'r') as f:
                    self.disk_cache_index = json.load(f)
                logger.info(f"Loaded {len(self.disk_cache_index)} disk cache entries")
        except Exception as e:
            logger.warning(f"Failed to load cache index: {e}")
            self.disk_cache_index = {}
    
    def _save_disk_cache_index(self):
        """Save disk cache index to persistent storage."""
        index_file = os.path.join(self.cache_dir, "cache_index.json")
        try:
            with self._lock:
                with open(index_file, 'w') as f:
                    json.dump(self.disk_cache_index, f, indent=2)
        except Exception as e:
            logger.error(f"Failed to save cache index: {e}")
    
    def _compress_data(self, data: Any) -> Tuple[bytes, float]:
        """Compress data and return compressed bytes with compression ratio."""
        try:
            # Serialize data
            serialized = pickle.dumps(data)
            original_size = len(serialized)
            
            # Compress if above threshold
            if original_size >= self.compression_threshold and self.enable_compression:
                compressed = gzip.compress(serialized)
                compression_ratio = len(compressed) / original_size
                return compressed, compression_ratio
            else:
                return serialized, 1.0
                
        except Exception as e:
            logger.error(f"Compression failed: {e}")
            return pickle.dumps(data), 1.0
    
    def _decompress_data(self, compressed_data: bytes, is_compressed: bool = True) -> Any:
        """Decompress data and return original object."""
        try:
            if is_compressed and self.enable_compression:
                decompressed = gzip.decompress(compressed_data)
                return pickle.loads(decompressed)
            else:
                return pickle.loads(compressed_data)
        except Exception as e:
            logger.error(f"Decompression failed: {e}")
            raise
    
    def _calculate_size(self, data: Any) -> int:
        """Calculate approximate size of data in bytes."""
        try:
            return len(pickle.dumps(data))
        except Exception:
            return 0
    
    def _evict_memory_cache(self, required_space: int = 0):
        """Evict entries from memory cache using LRU policy."""
        with self._lock:
            current_size = sum(entry.size_bytes for entry in self.memory_cache.values())
            target_size = self.max_memory_size - required_space
            
            while current_size > target_size and self.memory_cache:
                # Remove least recently used item
                key, entry = self.memory_cache.popitem(last=False)
                current_size -= entry.size_bytes
                self.stats.evictions += 1
                
                # Move to disk cache if valuable
                if entry.access_count > 1:
                    self._store_to_disk(key, entry)
                
                logger.debug(f"Evicted cache entry: {key}")
    
    def _store_to_disk(self, key: str, entry: CacheEntry):
        """Store cache entry to disk."""
        try:
            disk_file = os.path.join(self.cache_dir, f"{hashlib.md5(key.encode()).hexdigest()}.cache")
            compressed_data, compression_ratio = self._compress_data(entry.data)
            
            with open(disk_file, 'wb') as f:
                f.write(compressed_data)
            
            # Update disk cache index
            self.disk_cache_index[key] = {
                'file': disk_file,
                'timestamp': entry.timestamp,
                'ttl': entry.ttl,
                'size': len(compressed_data),
                'compressed': compression_ratio < 1.0,
                'access_count': entry.access_count
            }
            
            self._save_disk_cache_index()
            
        except Exception as e:
            logger.error(f"Failed to store to disk cache: {e}")
    
    def _cleanup_expired_entries(self):
        """Remove expired entries from both memory and disk cache."""
        current_time = time.time()
        
        with self._lock:
            # Clean memory cache
            expired_keys = [
                key for key, entry in self.memory_cache.items()
                if entry.is_expired()
            ]
            
            for key in expired_keys:
                del self.memory_cache[key]
                self.stats.evictions += 1
            
            # Clean disk cache
            expired_disk_keys = [
                key for key, info in self.disk_cache_index.items()
                if current_time - info['timestamp'] > info['ttl']
            ]
            
            for key in expired_disk_keys:
                try:
                    disk_file = self.disk_cache_index[key]['file']
                    if os.path.exists(disk_file):
                        os.remove(disk_file)
                    del self.disk_cache_index[key]
                except Exception as e:
                    logger.error(f"Failed to remove expired disk cache: {e}")
            
            if expired_keys or expired_disk_keys:
                logger.info(f"Cleaned up {len(expired_keys)} memory and {len(expired_disk_keys)} disk cache entries")
    
    def _optimize_cache_layout(self):
        """Optimize cache layout for better performance."""
        with self._lock:
            # Move frequently accessed items to memory cache
            if len(self.memory_cache) < self.max_memory_size // (10 * 1024):  # Keep some space
                for key, info in sorted(
                    self.disk_cache_index.items(),
                    key=lambda x: x[1]['access_count'],
                    reverse=True
                )[:10]:  # Top 10 most accessed
                    if key not in self.memory_cache:
                        try:
                            data = self._load_from_disk(key)
                            if data is not None:
                                entry = CacheEntry(
                                    data=data,
                                    timestamp=info['timestamp'],
                                    ttl=info['ttl'],
                                    access_count=info['access_count'],
                                    size_bytes=self._calculate_size(data)
                                )
                                self.memory_cache[key] = entry
                                # Remove from disk to avoid duplication
                                self._remove_from_disk(key)
                        except Exception as e:
                            logger.error(f"Failed to promote cache entry: {e}")
    
    def _update_statistics(self):
        """Update cache statistics."""
        with self._lock:
            total_requests = self.stats.hits + self.stats.misses
            if total_requests > 0:
                self.stats.hit_rate = (self.stats.hits / total_requests) * 100
            
            self.stats.entry_count = len(self.memory_cache) + len(self.disk_cache_index)
            self.stats.total_size_bytes = sum(entry.size_bytes for entry in self.memory_cache.values())
            
            # Calculate average access time
            all_times = []
            for times in self.access_times.values():
                all_times.extend(times[-10:])  # Last 10 access times per key
            
            if all_times:
                self.stats.avg_access_time_ms = sum(all_times) / len(all_times) * 1000
    
    def _load_from_disk(self, key: str) -> Optional[Any]:
        """Load data from disk cache."""
        try:
            if key not in self.disk_cache_index:
                return None
            
            info = self.disk_cache_index[key]
            disk_file = info['file']
            
            if not os.path.exists(disk_file):
                # Clean up stale index entry
                del self.disk_cache_index[key]
                return None
            
            with open(disk_file, 'rb') as f:
                compressed_data = f.read()
            
            return self._decompress_data(compressed_data, info.get('compressed', False))
            
        except Exception as e:
            logger.error(f"Failed to load from disk cache: {e}")
            return None
    
    def _remove_from_disk(self, key: str):
        """Remove entry from disk cache."""
        try:
            if key in self.disk_cache_index:
                info = self.disk_cache_index[key]
                disk_file = info['file']
                if os.path.exists(disk_file):
                    os.remove(disk_file)
                del self.disk_cache_index[key]
        except Exception as e:
            logger.error(f"Failed to remove from disk cache: {e}")
    
    @staticmethod
    @lru_cache(maxsize=128)
    def get_file_hash(file_path: str) -> str:
        """
        Get MD5 hash of file for change detection.
        
        Args:
            file_path: Path to the file
            
        Returns:
            str: MD5 hash of file content
        """
        try:
            with open(file_path, 'rb') as f:
                return hashlib.md5(f.read()).hexdigest()
        except (IOError, OSError):
            return ""
    
    @staticmethod
    @lru_cache(maxsize=64)
    def get_file_metadata(file_path: str) -> Dict[str, Any]:
        """
        Cache file metadata for quick access.
        
        Args:
            file_path: Path to the file
            
        Returns:
            Dict[str, Any]: File metadata including size, mtime, etc.
        """
        try:
            stat = os.stat(file_path)
            return {
                'size': stat.st_size,
                'modified': stat.st_mtime,
                'created': stat.st_ctime,
                'hash': AdvancedCacheManager.get_file_hash(file_path)
            }
        except (IOError, OSError):
            return {}
    
    def get(self, key: str, default: Any = None) -> Any:
        """
        Get cached data with performance tracking.
        
        Args:
            key: Cache key
            default: Default value if not found
            
        Returns:
            Cached data or default value
        """
        start_time = time.perf_counter()
        
        try:
            with self._lock:
                # Check memory cache first
                if key in self.memory_cache:
                    entry = self.memory_cache[key]
                    if not entry.is_expired():
                        # Move to end (LRU)
                        self.memory_cache.move_to_end(key)
                        result = entry.access()
                        self.stats.hits += 1
                        
                        # Track access time
                        access_time = time.perf_counter() - start_time
                        self.access_times[key].append(access_time)
                        if len(self.access_times[key]) > 100:
                            self.access_times[key] = self.access_times[key][-50:]
                        
                        return result
                    else:
                        # Remove expired entry
                        del self.memory_cache[key]
                
                # Check disk cache
                if key in self.disk_cache_index:
                    data = self._load_from_disk(key)
                    if data is not None:
                        # Promote to memory cache
                        entry = CacheEntry(
                            data=data,
                            timestamp=time.time(),
                            ttl=self.disk_cache_index[key]['ttl'],
                            size_bytes=self._calculate_size(data)
                        )
                        
                        # Ensure space in memory cache
                        required_space = entry.size_bytes
                        if required_space > self.max_memory_size * 0.1:  # Don't cache very large items
                            self.stats.misses += 1
                            return data
                        
                        self._evict_memory_cache(required_space)
                        self.memory_cache[key] = entry
                        
                        # Remove from disk to avoid duplication
                        self._remove_from_disk(key)
                        
                        self.stats.hits += 1
                        return data
                
                # Cache miss
                self.stats.misses += 1
                return default
                
        except Exception as e:
            logger.error(f"Cache get error for key {key}: {e}")
            self.stats.misses += 1
            return default
    
    def set(self, key: str, value: Any, ttl: float = 300.0) -> bool:
        """
        Set cached data with intelligent storage strategy.
        
        Args:
            key: Cache key
            value: Data to cache
            ttl: Time to live in seconds
            
        Returns:
            bool: True if successfully cached
        """
        try:
            with self._lock:
                # Calculate size and determine storage strategy
                size_bytes = self._calculate_size(value)
                
                # Create cache entry
                entry = CacheEntry(
                    data=value,
                    timestamp=time.time(),
                    ttl=ttl,
                    size_bytes=size_bytes
                )
                
                # Decide storage location based on size and access patterns
                if size_bytes <= self.max_memory_size * 0.1:  # Small items go to memory
                    self._evict_memory_cache(size_bytes)
                    self.memory_cache[key] = entry
                else:
                    # Large items go directly to disk
                    self._store_to_disk(key, entry)
                
                return True
                
        except Exception as e:
            logger.error(f"Cache set error for key {key}: {e}")
            return False
    
    def delete(self, key: str) -> bool:
        """
        Delete cached data.
        
        Args:
            key: Cache key to delete
            
        Returns:
            bool: True if successfully deleted
        """
        try:
            with self._lock:
                deleted = False
                
                # Remove from memory cache
                if key in self.memory_cache:
                    del self.memory_cache[key]
                    deleted = True
                
                # Remove from disk cache
                if key in self.disk_cache_index:
                    self._remove_from_disk(key)
                    deleted = True
                
                # Clean up access times
                if key in self.access_times:
                    del self.access_times[key]
                
                return deleted
                
        except Exception as e:
            logger.error(f"Cache delete error for key {key}: {e}")
            return False
    
    def clear(self) -> int:
        """
        Clear all cached data.
        
        Returns:
            int: Number of entries cleared
        """
        try:
            with self._lock:
                count = len(self.memory_cache) + len(self.disk_cache_index)
                
                # Clear memory cache
                self.memory_cache.clear()
                
                # Clear disk cache
                for key in list(self.disk_cache_index.keys()):
                    self._remove_from_disk(key)
                
                # Clear access times
                self.access_times.clear()
                
                # Reset statistics
                self.stats = CacheStats()
                
                return count
                
        except Exception as e:
            logger.error(f"Cache clear error: {e}")
            return 0
    
    def warm_cache(self, keys_and_loaders: List[Tuple[str, Callable[[], Any], float]]):
        """
        Warm cache with frequently accessed data.
        
        Args:
            keys_and_loaders: List of (key, loader_function, ttl) tuples
        """
        logger.info(f"Warming cache with {len(keys_and_loaders)} entries")
        
        for key, loader, ttl in keys_and_loaders:
            try:
                if self.get(key) is None:  # Only load if not already cached
                    data = loader()
                    self.set(key, data, ttl)
                    logger.debug(f"Warmed cache entry: {key}")
            except Exception as e:
                logger.error(f"Failed to warm cache for key {key}: {e}")
    
    @lru_cache(maxsize=32)
    def is_file_modified(self, file_path: str, last_known_hash: str) -> bool:
        """
        Check if file has been modified since last known hash.
        
        Args:
            file_path: Path to check
            last_known_hash: Previously cached hash
            
        Returns:
            bool: True if file has been modified
        """
        current_hash = self.get_file_hash(file_path)
        return current_hash != last_known_hash
    
    @lru_cache(maxsize=256)
    def validate_excel_file(self, file_path: str) -> Tuple[bool, str]:
        """
        Cache Excel file validation results.
        
        Args:
            file_path: Path to validate
            
        Returns:
            Tuple[bool, str]: (is_valid, error_message)
        """
        try:
            if not os.path.exists(file_path):
                return False, "File does not exist"
            
            if not file_path.lower().endswith(('.xlsx', '.xls')):
                return False, "Invalid file extension"
            
            # Check file size (max 100MB)
            file_size = os.path.getsize(file_path)
            if file_size > 100 * 1024 * 1024:
                return False, "File too large (>100MB)"
            
            # Try to open with openpyxl
            from openpyxl import load_workbook
            load_workbook(file_path, read_only=True, data_only=True).close()
            
            return True, ""
            
        except Exception as e:
            return False, str(e)
    
    @lru_cache(maxsize=128)
    def get_excel_sheet_names(self, file_path: str) -> List[str]:
        """
        Cache Excel sheet names for quick access.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            List[str]: List of sheet names
        """
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True)
            try:
                return wb.sheetnames
            finally:
                wb.close()
        except Exception:
            return []
    
    @lru_cache(maxsize=64)
    def get_partner_data(self, file_path: str, partner_number: str) -> Dict[str, Any]:
        """
        Cache partner data extraction from Excel files.
        
        Args:
            file_path: Path to Excel file
            partner_number: Partner identifier
            
        Returns:
            Dict[str, Any]: Partner data
        """
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            try:
                sheet_name = f"P{partner_number}"
                if sheet_name not in wb.sheetnames:
                    return {}
                
                sheet = wb[sheet_name]
                
                # Extract basic partner info
                data = {
                    'partner_number': partner_number,
                    'sheet_name': sheet_name,
                    'exists': True,
                    'data_extracted': False
                }
                
                # Try to extract common fields
                try:
                    data.update({
                        'partner_id': sheet['D4'].value,
                        'beneficiary_name': sheet['D5'].value,
                        'country': sheet['D6'].value,
                        'role': sheet['D7'].value,
                        'data_extracted': True
                    })
                except (KeyError, AttributeError):
                    pass
                
                return data
            finally:
                wb.close()
                
        except Exception as e:
            logger.error(f"Error extracting partner data: {e}")
            return {}
    
    @lru_cache(maxsize=64)
    def get_workpackage_data(self, file_path: str, workpackage_id: str) -> Dict[str, Any]:
        """
        Cache workpackage data extraction.
        
        Args:
            file_path: Path to Excel file
            workpackage_id: Workpackage identifier
            
        Returns:
            Dict[str, Any]: Workpackage data
        """
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            try:
                if "PM Summary" not in wb.sheetnames:
                    return {}
                
                sheet = wb["PM Summary"]
                
                # Find workpackage row (simplified logic)
                data = {
                    'workpackage_id': workpackage_id,
                    'exists': False,
                    'row_data': None
                }
                
                # Scan for workpackage ID in column A
                for row in range(2, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row, column=1).value
                    if str(cell_value) == str(workpackage_id):
                        data.update({
                            'exists': True,
                            'row': row,
                            'row_data': [
                                sheet.cell(row=row, column=col).value
                                for col in range(1, min(sheet.max_column + 1, 10))
                            ]
                        })
                        break
                
                return data
            finally:
                wb.close()
                
        except Exception as e:
            logger.error(f"Error extracting workpackage data: {e}")
            return {}
    
    def cache_with_ttl(self, key: str, data: Any, ttl_seconds: int = 300) -> None:
        """
        Cache data with time-to-live.
        
        Args:
            key: Cache key
            data: Data to cache
            ttl_seconds: Time to live in seconds
        """
        cache_file = os.path.join(self.cache_dir, f"{key}.json")
        cache_data = {
            'data': data,
            'timestamp': time.time(),
            'ttl': ttl_seconds
        }
        
        try:
            with open(cache_file, 'w') as f:
                json.dump(cache_data, f, default=str)
        except Exception as e:
            logger.error(f"Failed to cache data: {e}")
    
    def get_cached_with_ttl(self, key: str) -> Optional[Any]:
        """
        Get cached data with TTL check.
        
        Args:
            key: Cache key
            
        Returns:
            Optional[Any]: Cached data or None if expired/missing
        """
        cache_file = os.path.join(self.cache_dir, f"{key}.json")
        
        try:
            if not os.path.exists(cache_file):
                return None
            
            with open(cache_file, 'r') as f:
                cache_data = json.load(f)
            
            # Check if cache is expired
            if time.time() - cache_data['timestamp'] > cache_data['ttl']:
                os.remove(cache_file)
                return None
            
            return cache_data['data']
            
        except Exception as e:
            logger.error(f"Failed to get cached data: {e}")
            return None
    
    def clear_cache(self, pattern: str = "*") -> int:
        """
        Clear cache files matching pattern.
        
        Args:
            pattern: File pattern to match (default: all)
            
        Returns:
            int: Number of files removed
        """
        import glob
        
        removed_count = 0
        cache_pattern = os.path.join(self.cache_dir, pattern)
        
        try:
            for cache_file in glob.glob(cache_pattern):
                if cache_file.endswith('.json'):
                    os.remove(cache_file)
                    removed_count += 1
        except Exception as e:
            logger.error(f"Failed to clear cache: {e}")
        
        return removed_count
    
    def get_cache_stats(self) -> Dict[str, Any]:
        """
        Get comprehensive cache statistics.
        
        Returns:
            Dict[str, Any]: Cache statistics including performance metrics
        """
        with self._lock:
            return {
                'memory_cache_size': len(self.memory_cache),
                'disk_cache_size': len(self.disk_cache_index),
                'total_entries': len(self.memory_cache) + len(self.disk_cache_index),
                'memory_usage_bytes': sum(entry.size_bytes for entry in self.memory_cache.values()),
                'max_memory_size': self.max_memory_size,
                'max_disk_size': self.max_disk_size,
                'hit_rate': self.stats.hit_rate,
                'hits': self.stats.hits,
                'misses': self.stats.misses,
                'evictions': self.stats.evictions,
                'avg_access_time_ms': self.stats.avg_access_time_ms,
                'compression_enabled': self.enable_compression,
                'compression_threshold': self.compression_threshold,
                'file_hash_cache_size': self.get_file_hash.cache_info().currsize if hasattr(self.get_file_hash, 'cache_info') else 0,
                'file_hash_cache_max': self.get_file_hash.cache_info().maxsize if hasattr(self.get_file_hash, 'cache_info') else 0,
                'metadata_cache_size': self.get_file_metadata.cache_info().currsize if hasattr(self.get_file_metadata, 'cache_info') else 0,
                'metadata_cache_max': self.get_file_metadata.cache_info().maxsize if hasattr(self.get_file_metadata, 'cache_info') else 0,
            }
    
    def invalidate_cache_for_file(self, file_path: str) -> None:
        """
        Invalidate all cached data for a specific file.
        
        Args:
            file_path: Path to invalidate cache for
        """
        # Clear LRU cache entries for this file
        self.get_file_hash.cache_clear()
        self.get_file_metadata.cache_clear()
        self.get_excel_sheet_names.cache_clear()
        self.validate_excel_file.cache_clear()
        self.get_partner_data.cache_clear()
        self.get_workpackage_data.cache_clear()
        
        logger.info(f"Invalidated cache for file: {file_path}")


def cached_with_invalidation(cache_manager: AdvancedCacheManager, ttl: int = 300):
    """
    Decorator for caching function results with TTL and invalidation.
    
    Args:
        cache_manager: CacheManager instance
        ttl: Time to live in seconds
    """
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            # Create cache key from function name and arguments
            key_parts = [func.__name__]
            key_parts.extend(str(arg) for arg in args)
            key_parts.extend(f"{k}:{v}" for k, v in sorted(kwargs.items()))
            key = "_".join(key_parts)
            
            # Try to get from cache
            cached = cache_manager.get_cached_with_ttl(key)
            if cached is not None:
                logger.debug(f"Cache hit for {key}")
                return cached
            
            # Compute and cache
            result = func(*args, **kwargs)
            cache_manager.cache_with_ttl(key, result, ttl)
            logger.debug(f"Cache miss for {key}, cached result")
            
            return result
        return wrapper
    return decorator


class CacheAwareExcelManager:
    """
    ExcelManager wrapper with caching capabilities.
    
    Provides caching layer on top of Excel operations to avoid
    repeated expensive operations.
    """
    
    def __init__(self, cache_manager: AdvancedCacheManager):
        """Initialize with cache manager."""
        self.cache = cache_manager
    
    def get_cached_sheet_names(self, file_path: str) -> List[str]:
        """Get sheet names with caching."""
        return self.cache.get_excel_sheet_names(file_path)
    
    def get_cached_partner_data(self, file_path: str, partner_number: str) -> Dict[str, Any]:
        """Get partner data with caching."""
        return self.cache.get_partner_data(file_path, partner_number)
    
    def get_cached_workpackage_data(self, file_path: str, workpackage_id: str) -> Dict[str, Any]:
        """Get workpackage data with caching."""
        return self.cache.get_workpackage_data(file_path, workpackage_id)
    
    def is_file_valid(self, file_path: str) -> Tuple[bool, str]:
        """Check file validity with caching."""
        return self.cache.validate_excel_file(file_path)
    
    def has_file_changed(self, file_path: str, last_hash: str) -> bool:
        """Check if file has changed since last known hash."""
        return self.cache.is_file_modified(file_path, last_hash)


# Global cache manager instance
# Global cache manager instance
cache_manager = AdvancedCacheManager()

# Backward compatibility alias
CacheManager = AdvancedCacheManager
