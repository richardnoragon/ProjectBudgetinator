"""
Streaming data processor for large Excel files.

This module provides memory-efficient processing of large Excel files
through chunked reading, streaming operations, and progressive data handling.
"""

import logging
import gc
import time
from typing import Iterator, List, Dict, Any, Optional, Callable, Union
from contextlib import contextmanager
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    load_workbook = None
    Worksheet = None

from utils.resource_cleanup import excel_tracker
from utils.performance_monitor import monitor_performance

logger = logging.getLogger(__name__)


class StreamingExcelProcessor:
    """
    Memory-efficient Excel file processor with streaming capabilities.
    
    Processes large Excel files in chunks to minimize memory usage
    and prevent memory exhaustion on large datasets.
    """
    
    def __init__(self, file_path: str, chunk_size: int = 1000, 
                 memory_limit_mb: float = 200.0):
        """
        Initialize streaming processor.
        
        Args:
            file_path: Path to Excel file
            chunk_size: Number of rows to process per chunk
            memory_limit_mb: Memory limit in MB before triggering cleanup
        """
        self.file_path = Path(file_path)
        self.chunk_size = chunk_size
        self.memory_limit_mb = memory_limit_mb
        self.processed_rows = 0
        self.total_chunks = 0
        self.operation_id = f"stream_{int(time.time())}"
        
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for streaming Excel processing")
    
    @monitor_performance(include_memory=True)
    def process_worksheet_in_chunks(self, worksheet_name: str,
                                    processor_func: Callable[[List[List[Any]]], None],
                                    start_row: int = 1, end_row: Optional[int] = None) -> Dict[str, Any]:
        """
        Process worksheet data in memory-efficient chunks.
        
        Args:
            worksheet_name: Name of worksheet to process
            processor_func: Function to process each chunk of data
            start_row: Starting row number (1-based)
            end_row: Ending row number (None for all rows)
            
        Returns:
            Dict[str, Any]: Processing statistics
        """
        logger.info(f"Starting chunked processing of {worksheet_name} in {self.file_path}")
        
        # Start tracking streaming operation
        excel_tracker.start_streaming_operation(
            self.operation_id, str(self.file_path), self.chunk_size
        )
        
        try:
            with self._get_worksheet_context(worksheet_name) as worksheet:
                if not worksheet:
                    raise ValueError(f"Worksheet '{worksheet_name}' not found")
                
                # Determine row range
                max_row = end_row or worksheet.max_row
                total_rows = max_row - start_row + 1
                
                logger.info(f"Processing {total_rows} rows in chunks of {self.chunk_size}")
                
                current_row = start_row
                chunk_count = 0
                
                while current_row <= max_row:
                    # Calculate chunk end
                    chunk_end = min(current_row + self.chunk_size - 1, max_row)
                    
                    # Extract chunk data
                    chunk_data = self._extract_chunk_data(worksheet, current_row, chunk_end)
                    
                    # Process chunk
                    processor_func(chunk_data)
                    
                    # Update progress
                    chunk_count += 1
                    self.processed_rows += len(chunk_data)
                    excel_tracker.update_streaming_progress(self.operation_id, chunk_count)
                    
                    # Check memory usage and cleanup if needed
                    self._check_memory_and_cleanup()
                    
                    # Move to next chunk
                    current_row = chunk_end + 1
                    
                    logger.debug(f"Processed chunk {chunk_count}: rows {current_row - len(chunk_data)} to {chunk_end}")
                
                self.total_chunks = chunk_count
                
                return {
                    'total_rows_processed': self.processed_rows,
                    'total_chunks': self.total_chunks,
                    'chunk_size': self.chunk_size,
                    'file_path': str(self.file_path),
                    'worksheet_name': worksheet_name
                }
                
        except Exception as e:
            logger.error(f"Error during chunked processing: {e}")
            raise
        finally:
            excel_tracker.finish_streaming_operation(self.operation_id)
    
    @contextmanager
    def _get_worksheet_context(self, worksheet_name: str):
        """Get worksheet with proper resource management."""
        workbook = None
        try:
            # Use read-only mode for streaming
            if not OPENPYXL_AVAILABLE or load_workbook is None:
                raise ImportError("openpyxl is not available")
            
            workbook = load_workbook(
                str(self.file_path),
                read_only=True,
                data_only=True,
                keep_links=False
            )
            
            if worksheet_name not in workbook.sheetnames:
                yield None
                return
            
            worksheet = workbook[worksheet_name]
            yield worksheet
            
        except Exception as e:
            logger.error(f"Error accessing worksheet {worksheet_name}: {e}")
            raise
        finally:
            if workbook:
                try:
                    workbook.close()
                    gc.collect()
                except Exception as e:
                    logger.warning(f"Error closing workbook: {e}")
    
    def _extract_chunk_data(self, worksheet, start_row: int, end_row: int) -> List[List[Any]]:
        """Extract data from worksheet chunk."""
        chunk_data = []
        
        try:
            for row_num in range(start_row, end_row + 1):
                row_data = []
                for col_num in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=row_num, column=col_num).value
                    row_data.append(cell_value)
                chunk_data.append(row_data)
                
        except Exception as e:
            logger.error(f"Error extracting chunk data from rows {start_row}-{end_row}: {e}")
            raise
        
        return chunk_data
    
    def _check_memory_and_cleanup(self):
        """Check memory usage and trigger cleanup if needed."""
        try:
            import psutil
            current_memory = psutil.Process().memory_info().rss / (1024 * 1024)
            
            if current_memory > self.memory_limit_mb:
                logger.warning(f"Memory limit exceeded: {current_memory:.1f}MB > {self.memory_limit_mb:.1f}MB")
                logger.info("Triggering garbage collection...")
                gc.collect()
                
                # Check memory again after cleanup
                new_memory = psutil.Process().memory_info().rss / (1024 * 1024)
                logger.info(f"Memory after cleanup: {new_memory:.1f}MB (freed {current_memory - new_memory:.1f}MB)")
                
        except ImportError:
            # psutil not available, skip memory check
            pass
        except Exception as e:
            logger.warning(f"Error checking memory usage: {e}")


class BatchExcelProcessor:
    """
    Process multiple Excel files with memory management.
    
    Handles batch processing of Excel files with automatic memory cleanup
    and progress tracking.
    """
    
    def __init__(self, memory_limit_mb: float = 500.0):
        """
        Initialize batch processor.
        
        Args:
            memory_limit_mb: Memory limit before triggering cleanup
        """
        self.memory_limit_mb = memory_limit_mb
        self.processed_files = 0
        self.failed_files = []
        self.processing_stats = {}
    
    @monitor_performance(include_memory=True)
    def process_files(self, file_paths: List[str],
                      processor_func: Callable[[str], Dict[str, Any]]) -> Dict[str, Any]:
        """
        Process multiple Excel files with memory management.
        
        Args:
            file_paths: List of file paths to process
            processor_func: Function to process each file
            
        Returns:
            Dict[str, Any]: Batch processing results
        """
        logger.info(f"Starting batch processing of {len(file_paths)} files")
        
        start_time = time.time()
        
        for i, file_path in enumerate(file_paths):
            try:
                logger.info(f"Processing file {i + 1}/{len(file_paths)}: {file_path}")
                
                # Process individual file
                file_stats = processor_func(file_path)
                self.processing_stats[file_path] = file_stats
                self.processed_files += 1
                
                # Check memory and cleanup if needed
                self._check_memory_and_cleanup()
                
                logger.info(f"Completed file {i + 1}/{len(file_paths)}")
                
            except Exception as e:
                logger.error(f"Failed to process file {file_path}: {e}")
                self.failed_files.append({
                    'file_path': file_path,
                    'error': str(e)
                })
        
        duration = time.time() - start_time
        
        return {
            'total_files': len(file_paths),
            'processed_files': self.processed_files,
            'failed_files': len(self.failed_files),
            'failure_details': self.failed_files,
            'processing_time_seconds': duration,
            'files_per_second': len(file_paths) / duration if duration > 0 else 0,
            'individual_stats': self.processing_stats
        }
    
    def _check_memory_and_cleanup(self):
        """Check memory usage and trigger cleanup if needed."""
        try:
            import psutil
            current_memory = psutil.Process().memory_info().rss / (1024 * 1024)
            
            if current_memory > self.memory_limit_mb:
                logger.warning(f"Memory limit exceeded during batch processing: "
                               f"{current_memory:.1f}MB > {self.memory_limit_mb:.1f}MB")
                
                # Force comprehensive cleanup
                excel_tracker.cleanup_temp_files()
                gc.collect()
                
                new_memory = psutil.Process().memory_info().rss / (1024 * 1024)
                logger.info(f"Memory after batch cleanup: {new_memory:.1f}MB "
                            f"(freed {current_memory - new_memory:.1f}MB)")
                
        except ImportError:
            pass
        except Exception as e:
            logger.warning(f"Error during batch memory cleanup: {e}")


# Utility functions for common streaming operations

def stream_process_large_file(file_path: str, worksheet_name: str,
                              processor_func: Callable[[List[List[Any]]], None],
                              chunk_size: int = 1000) -> Dict[str, Any]:
    """
    Convenience function for streaming processing of large Excel files.
    
    Args:
        file_path: Path to Excel file
        worksheet_name: Name of worksheet to process
        processor_func: Function to process each chunk
        chunk_size: Number of rows per chunk
        
    Returns:
        Dict[str, Any]: Processing statistics
    """
    processor = StreamingExcelProcessor(file_path, chunk_size)
    return processor.process_worksheet_in_chunks(worksheet_name, processor_func)


def batch_process_excel_files(file_paths: List[str],
                              processor_func: Callable[[str], Dict[str, Any]]) -> Dict[str, Any]:
    """
    Convenience function for batch processing Excel files.
    
    Args:
        file_paths: List of file paths to process
        processor_func: Function to process each file
        
    Returns:
        Dict[str, Any]: Batch processing results
    """
    processor = BatchExcelProcessor()
    return processor.process_files(file_paths, processor_func)


# Example usage functions

def example_chunk_processor(chunk_data: List[List[Any]]) -> None:
    """
    Example chunk processor function.
    
    Args:
        chunk_data: List of rows, each row is a list of cell values
    """
    # Process the chunk data
    for row_idx, row in enumerate(chunk_data):
        # Example: print first column of each row
        if row and row[0] is not None:
            logger.debug(f"Processing row {row_idx}: {row[0]}")


def example_file_processor(file_path: str) -> Dict[str, Any]:
    """
    Example file processor function.
    
    Args:
        file_path: Path to Excel file
        
    Returns:
        Dict[str, Any]: Processing results
    """
    try:
        # Example: count rows in first worksheet
        if not OPENPYXL_AVAILABLE or load_workbook is None:
            raise ImportError("openpyxl is not available")
        
        wb = load_workbook(file_path, read_only=True)
        try:
            first_sheet = wb.worksheets[0]
            row_count = first_sheet.max_row
        finally:
            wb.close()
            
        return {
            'file_path': file_path,
            'row_count': row_count,
            'status': 'success'
        }
    except Exception as e:
        return {
            'file_path': file_path,
            'error': str(e),
            'status': 'failed'
        }


if __name__ == "__main__":
    # Example usage
    logger.info("Streaming processor module loaded successfully")