"""
Resource cleanup utilities for ProjectBudgetinator.

This module provides comprehensive resource management including:
- Excel workbook lifecycle management
- Memory leak prevention
- Automatic cleanup on application exit
- Memory usage monitoring
"""

import os
import gc
import atexit
import weakref
import threading
import time
from typing import Dict, List, Any, Optional, Iterator
from contextlib import contextmanager
import logging

logger = logging.getLogger(__name__)


class ResourceCleanupManager:
    """
    Centralized resource cleanup manager for Excel operations.
    
    Ensures proper cleanup of Excel resources, prevents memory leaks,
    and provides monitoring capabilities.
    """
    
    def __init__(self):
        self._open_resources = {}
        self._cleanup_callbacks = []
        self._lock = threading.Lock()
        self._setup_cleanup_handlers()
    
    def _setup_cleanup_handlers(self):
        """Setup automatic cleanup handlers."""
        # Register cleanup on application exit
        atexit.register(self._cleanup_all_resources)
        
        # Register cleanup callbacks
        self._register_cleanup_callbacks()
    
    def register_resource(self, resource_id: str, resource, cleanup_func):
        """Register a resource for automatic cleanup."""
        with self._lock:
            self._open_resources[resource_id] = {
                'resource': resource,
                'cleanup_func': cleanup_func,
                'created_at': threading.current_thread().ident
            }
    
    def unregister_resource(self, resource_id: str):
        """Unregister a resource from cleanup tracking."""
        with self._lock:
            if resource_id in self._open_resources:
                del self._open_resources[resource_id]
    
    def cleanup_resource(self, resource_id: str) -> bool:
        """Cleanup a specific resource."""
        with self._lock:
            if resource_id in self._open_resources:
                try:
                    resource_info = self._open_resources[resource_id]
                    resource_info['cleanup_func'](resource_info['resource'])
                    del self._open_resources[resource_id]
                    logger.debug(f"Cleaned up resource: {resource_id}")
                    return True
                except Exception as e:
                    logger.error(f"Error cleaning up resource {resource_id}: {e}")
                    return False
        return False
    
    def _cleanup_all_resources(self):
        """Cleanup all registered resources."""
        logger.info("Starting comprehensive resource cleanup...")
        
        with self._lock:
            resources_to_cleanup = list(self._open_resources.items())
            
            for resource_id, resource_info in resources_to_cleanup:
                try:
                    resource_info['cleanup_func'](resource_info['resource'])
                    logger.debug(f"Auto-cleanup: {resource_id}")
                except Exception as e:
                    logger.error(f"Auto-cleanup error for {resource_id}: {e}")
            
            self._open_resources.clear()
        
        # Run garbage collection
        gc.collect()
        logger.info("Resource cleanup completed")
    
    def get_resource_stats(self) -> Dict[str, Any]:
        """Get resource management statistics."""
        with self._lock:
            return {
                'active_resources': len(self._open_resources),
                'resource_ids': list(self._open_resources.keys()),
                'cleanup_callbacks': len(self._cleanup_callbacks)
            }
    
    def _register_cleanup_callbacks(self):
        """Register various cleanup callbacks."""
        # Excel workbook cleanup
        self.add_cleanup_callback(self._cleanup_excel_resources)
        
        # Temporary file cleanup
        self.add_cleanup_callback(self._cleanup_temp_files)
        
        # Memory cleanup
        self.add_cleanup_callback(self._cleanup_memory)
    
    def _cleanup_excel_resources(self):
        """Cleanup Excel-specific resources."""
        try:
            import openpyxl
            # Force close any remaining workbooks
            gc.collect()
        except ImportError:
            pass
    
    def _cleanup_temp_files(self):
        """Cleanup temporary files."""
        try:
            temp_dir = os.path.join(os.path.expanduser("~"), "ProjectBudgetinator", "temp")
            if os.path.exists(temp_dir):
                import shutil
                shutil.rmtree(temp_dir, ignore_errors=True)
        except Exception as e:
            logger.warning(f"Error cleaning temp files: {e}")
    
    def _cleanup_memory(self):
        """Force memory cleanup."""
        try:
            gc.collect()
            logger.debug("Forced garbage collection")
        except Exception as e:
            logger.warning(f"Error during memory cleanup: {e}")
    
    def add_cleanup_callback(self, callback):
        """Add a custom cleanup callback."""
        self._cleanup_callbacks.append(callback)


# Global resource cleanup manager
cleanup_manager = ResourceCleanupManager()


class ExcelResourceTracker:
    """
    Enhanced Excel resource tracker with streaming and chunked processing support.
    """
    
    def __init__(self):
        self._workbooks = {}
        self._temp_files = set()
        self._streaming_operations = {}
        self._memory_usage_history = []
        self._lock = threading.Lock()
    
    def track_workbook(self, file_path: str, workbook):
        """Track an Excel workbook for cleanup with enhanced metadata."""
        def cleanup_func(wb):
            try:
                wb.close()
                logger.debug(f"Closed workbook: {file_path}")
            except Exception as e:
                logger.error(f"Error closing workbook {file_path}: {e}")
        
        with self._lock:
            cleanup_manager.register_resource(f"workbook_{file_path}", workbook, cleanup_func)
            self._workbooks[file_path] = {
                'workbook': workbook,
                'opened_at': time.time(),
                'memory_at_open': self._get_current_memory()
            }
    
    def untrack_workbook(self, file_path: str):
        """Untrack an Excel workbook."""
        with self._lock:
            cleanup_manager.unregister_resource(f"workbook_{file_path}")
            if file_path in self._workbooks:
                workbook_info = self._workbooks.pop(file_path)
                # Record memory usage history
                self._record_memory_usage(file_path, workbook_info)
    
    def track_temp_file(self, file_path: str):
        """Track a temporary file for cleanup."""
        with self._lock:
            self._temp_files.add(file_path)
    
    def cleanup_temp_files(self):
        """Cleanup tracked temporary files."""
        with self._lock:
            for file_path in list(self._temp_files):
                try:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                        logger.debug(f"Removed temp file: {file_path}")
                    self._temp_files.discard(file_path)
                except Exception as e:
                    logger.error(f"Error removing temp file {file_path}: {e}")
    
    def start_streaming_operation(self, operation_id: str, file_path: str, chunk_size: int = 1000):
        """Start tracking a streaming operation for large Excel files."""
        with self._lock:
            self._streaming_operations[operation_id] = {
                'file_path': file_path,
                'chunk_size': chunk_size,
                'started_at': time.time(),
                'processed_chunks': 0,
                'memory_snapshots': []
            }
            logger.info(f"Started streaming operation {operation_id} for {file_path}")
    
    def update_streaming_progress(self, operation_id: str, chunks_processed: int):
        """Update progress of a streaming operation."""
        with self._lock:
            if operation_id in self._streaming_operations:
                op = self._streaming_operations[operation_id]
                op['processed_chunks'] = chunks_processed
                
                # Take memory snapshot every 10 chunks
                if chunks_processed % 10 == 0:
                    memory_mb = self._get_current_memory()
                    op['memory_snapshots'].append({
                        'chunk': chunks_processed,
                        'memory_mb': memory_mb,
                        'timestamp': time.time()
                    })
                    
                    # Check for memory growth
                    if len(op['memory_snapshots']) > 1:
                        prev_memory = op['memory_snapshots'][-2]['memory_mb']
                        if memory_mb > prev_memory * 1.2:  # 20% increase
                            logger.warning(f"Memory growth detected in streaming operation {operation_id}: "
                                           f"{prev_memory:.1f}MB -> {memory_mb:.1f}MB")
    
    def finish_streaming_operation(self, operation_id: str):
        """Finish and cleanup a streaming operation."""
        with self._lock:
            if operation_id in self._streaming_operations:
                op = self._streaming_operations.pop(operation_id)
                duration = time.time() - op['started_at']
                logger.info(f"Completed streaming operation {operation_id}: "
                            f"{op['processed_chunks']} chunks in {duration:.2f}s")
                
                # Force cleanup after streaming operation
                gc.collect()
    
    def _get_current_memory(self) -> float:
        """Get current memory usage in MB."""
        try:
            import psutil
            return psutil.Process().memory_info().rss / (1024 * 1024)
        except Exception:
            return 0.0
    
    def _record_memory_usage(self, file_path: str, workbook_info: Dict[str, Any]):
        """Record memory usage history for analysis."""
        current_memory = self._get_current_memory()
        memory_delta = current_memory - workbook_info.get('memory_at_open', 0)
        
        self._memory_usage_history.append({
            'file_path': file_path,
            'opened_at': workbook_info.get('opened_at', 0),
            'closed_at': time.time(),
            'memory_delta_mb': memory_delta,
            'peak_memory_mb': current_memory
        })
        
        # Keep only last 100 entries
        if len(self._memory_usage_history) > 100:
            self._memory_usage_history = self._memory_usage_history[-100:]
    
    def get_stats(self) -> Dict[str, Any]:
        """Get comprehensive resource tracking statistics."""
        with self._lock:
            avg_memory_delta = 0
            if self._memory_usage_history:
                avg_memory_delta = sum(h['memory_delta_mb'] for h in self._memory_usage_history) / len(self._memory_usage_history)
            
            return {
                'tracked_workbooks': len(self._workbooks),
                'tracked_temp_files': len(self._temp_files),
                'active_files': list(self._workbooks.keys()),
                'streaming_operations': len(self._streaming_operations),
                'memory_history_entries': len(self._memory_usage_history),
                'avg_memory_delta_mb': avg_memory_delta,
                'current_memory_mb': self._get_current_memory()
            }
    
    def get_memory_analysis(self) -> Dict[str, Any]:
        """Get detailed memory usage analysis."""
        with self._lock:
            if not self._memory_usage_history:
                return {'analysis': 'No memory history available'}
            
            # Analyze memory patterns
            memory_deltas = [h['memory_delta_mb'] for h in self._memory_usage_history]
            peak_memories = [h['peak_memory_mb'] for h in self._memory_usage_history]
            
            return {
                'total_operations': len(self._memory_usage_history),
                'avg_memory_delta_mb': sum(memory_deltas) / len(memory_deltas),
                'max_memory_delta_mb': max(memory_deltas),
                'min_memory_delta_mb': min(memory_deltas),
                'avg_peak_memory_mb': sum(peak_memories) / len(peak_memories),
                'max_peak_memory_mb': max(peak_memories),
                'memory_leaks_detected': len([d for d in memory_deltas if d > 50]),  # >50MB delta
                'recent_operations': self._memory_usage_history[-10:]
            }


# Global Excel resource tracker
excel_tracker = ExcelResourceTracker()


@contextmanager
def safe_excel_context(file_path: str, **kwargs) -> Iterator:
    """
    Enhanced context manager for safe Excel file handling with resource tracking.
    
    Args:
        file_path: Path to the Excel file
        **kwargs: Additional arguments for load_workbook
    
    Yields:
        Workbook: The Excel workbook instance
    """
    from openpyxl import load_workbook
    
    workbook = None
    try:
        # Apply optimized settings
        kwargs.setdefault('data_only', True)
        kwargs.setdefault('keep_links', False)
        
        # Use read-only for large files
        file_size = os.path.getsize(file_path)
        if file_size > 10 * 1024 * 1024:
            kwargs.setdefault('read_only', True)
        
        workbook = load_workbook(file_path, **kwargs)
        excel_tracker.track_workbook(file_path, workbook)
        
        logger.debug(f"Opened workbook with tracking: {file_path}")
        yield workbook
        
    except Exception as e:
        logger.error(f"Error in safe_excel_context: {str(e)}")
        raise
    finally:
        if workbook:
            try:
                workbook.close()
                excel_tracker.untrack_workbook(file_path)
                gc.collect()
                logger.debug(f"Closed workbook with tracking: {file_path}")
            except Exception as e:
                logger.warning(f"Error closing workbook: {str(e)}")


@contextmanager
def memory_monitored_context(file_path: str, **kwargs) -> Iterator:
    """
    Context manager with memory monitoring for Excel operations.
    
    Args:
        file_path: Path to the Excel file
        **kwargs: Additional arguments for load_workbook
    
    Yields:
        Workbook: The Excel workbook instance
    """
    from openpyxl import load_workbook
    
    # Get initial memory usage
    import psutil
    process = psutil.Process()
    initial_memory = process.memory_info().rss / (1024 * 1024)
    
    workbook = None
    try:
        workbook = load_workbook(file_path, **kwargs)
        
        current_memory = process.memory_info().rss / (1024 * 1024)
        logger.info(f"Memory after loading {file_path}: {current_memory:.2f}MB "
                    f"(Δ{current_memory - initial_memory:+.2f}MB)")
        
        yield workbook
        
    finally:
        if workbook:
            try:
                workbook.close()
                gc.collect()
                
                final_memory = process.memory_info().rss / (1024 * 1024)
                logger.info(f"Memory after closing {file_path}: {final_memory:.2f}MB "
                            f"(Δ{final_memory - initial_memory:+.2f}MB)")
            except Exception as e:
                logger.warning(f"Error during memory cleanup: {str(e)}")


class MemoryMonitor:
    """
    Monitor and report memory usage for Excel operations.
    """
    
    def __init__(self):
        self.operation_stats = {}
    
    def monitor_operation(self, operation_name: str):
        """Decorator to monitor memory usage of operations."""
        def decorator(func):
            def wrapper(*args, **kwargs):
                import psutil
                process = psutil.Process()
                
                start_memory = process.memory_info().rss / (1024 * 1024)
                start_time = time.time()
                
                try:
                    result = func(*args, **kwargs)
                    return result
                finally:
                    end_memory = process.memory_info().rss / (1024 * 1024)
                    duration = time.time() - start_time
                    
                    self.operation_stats[operation_name] = {
                        'memory_delta_mb': end_memory - start_memory,
                        'duration_seconds': duration,
                        'peak_memory_mb': max(start_memory, end_memory)
                    }
                    
                    logger.info(f"Operation {operation_name}: "
                                f"{duration:.2f}s, "
                                f"Δ{end_memory - start_memory:+.2f}MB")
            
            return wrapper
        return decorator
    
    def get_stats(self) -> Dict[str, Any]:
        """Get memory monitoring statistics."""
        return {
            'operations': self.operation_stats,
            'total_operations': len(self.operation_stats)
        }


# Global memory monitor
memory_monitor = MemoryMonitor()


def cleanup_on_exception(func):
    """Decorator to ensure cleanup on exceptions."""
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            logger.error(f"Exception in {func.__name__}: {str(e)}")
            # Force cleanup on exception
            gc.collect()
            raise
    return wrapper


def get_system_memory_info() -> Dict[str, Any]:
    """Get comprehensive system memory information."""
    try:
        import psutil
        
        memory = psutil.virtual_memory()
        swap = psutil.swap_memory()
        
        return {
            'total_memory_gb': memory.total / (1024**3),
            'available_memory_gb': memory.available / (1024**3),
            'used_memory_gb': memory.used / (1024**3),
            'memory_percent': memory.percent,
            'swap_total_gb': swap.total / (1024**3),
            'swap_used_gb': swap.used / (1024**3),
            'swap_percent': swap.percent
        }
    except ImportError:
        return {'error': 'psutil not available'}


def force_system_cleanup():
    """Force comprehensive system cleanup."""
    logger.info("Starting system-wide resource cleanup...")
    
    # Cleanup all managed resources
    cleanup_manager._cleanup_all_resources()
    
    # Cleanup temporary files
    excel_tracker.cleanup_temp_files()
    
    # Force garbage collection
    gc.collect()
    
    # Log final state
    memory_info = get_system_memory_info()
    logger.info(f"System cleanup completed: {memory_info}")


# Register cleanup on module import
atexit.register(force_system_cleanup)
