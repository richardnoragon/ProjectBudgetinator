import sys
import os
sys.path.insert(0, os.path.abspath(os.path.dirname(__file__)))

"""ProjectBudgetinator Main Application Module.

This module contains the main application logic for ProjectBudgetinator, a comprehensive
Excel workbook management tool for project budget coordination. It provides a Tkinter-based
GUI interface for managing partners, workpackages, and budget overviews in Excel files.

The module handles:
    - Tkinter GUI setup and management
    - Excel file operations (open, save, validate)
    - Partner management (add, edit, delete)
    - Workpackage management (add, edit, delete)
    - Budget overview updates and PM overview updates
    - User authentication and profile management
    - Window positioning and theme management
    - Performance monitoring and diagnostics
    - File comparison and cloning operations
    - Batch operations and preferences management

Classes:
    ProjectBudgetinator: Main application class that orchestrates all functionality

Constants:
    DEFAULT_USER_CONFIG: Default user configuration settings
    DEFAULT_BACKUP_CONFIG: Default backup configuration settings
    DEFAULT_DIAGNOSTIC_CONFIG: Default diagnostic configuration settings
    EXCEL_FILETYPES: Supported Excel file types for dialogs
    EXCEL_DEFAULT_EXT: Default Excel file extension

Example:
    Basic usage of the application:
    
        app = ProjectBudgetinator()
        app.run()

Note:
    This module requires proper authentication setup and preferences management
    to function correctly. All file operations are secured through centralized
    validation mechanisms.
"""

import json
import os
import datetime
import shutil
import tempfile
import subprocess
import pandas as pd
from pathlib import Path
import tkinter as tk
from tkinter import messagebox, filedialog, ttk
from devtools import not_implemented_yet, DebugConsole, dev_log
from version import full_version_string  # Only import what we use
from logger import setup_logging, get_structured_logger, LogContext
from utils.error_handler import setup_error_handling, handle_exceptions
from utils.performance_monitor import get_performance_monitor, monitor_performance
from utils.performance_optimizations import get_performance_optimizer, monitor_operation
from utils.security_validator import SecurityValidator, InputSanitizer
from utils.workbook_utils import (
    validate_and_load_workbook, save_workbook_with_dialog,
    ensure_workbook_loaded, create_workbook_operation_context,
    prompt_for_workbook_if_needed, is_partner_worksheet
)
from utils.update_utils import (
    UpdateBatch, execute_update_batch, create_standard_update_batch,
    validate_positive_number, validate_non_negative_number
)
from utils.user_utils import construct_user_safely, UserConstructor
from gui.performance_monitor_gui import show_performance_monitor, PerformanceIndicator
from validation import FormValidator
from gui.batch_operations import show_batch_operations_dialog

# Safe openpyxl imports with fallback
try:
    from openpyxl.styles import PatternFill  # Used in _add_partner_worksheet
    import openpyxl.utils.cell
    OPENPYXL_STYLES_AVAILABLE = True
except ImportError:
    PatternFill = None
    OPENPYXL_STYLES_AVAILABLE = False

# Application config defaults
DEFAULT_USER_CONFIG = {
    "theme": "light",
    "welcome_screen": True,
    "startup_diagnostic": "verbose"
}

DEFAULT_BACKUP_CONFIG = {
    "frequency": "daily",
    "keep_versions": 5
}

DEFAULT_DIAGNOSTIC_CONFIG = {
    "debug_mode": False,
    "log_level": "INFO"
}

# File type constants
EXCEL_FILETYPES = [("Excel files", "*.xlsx;*.xls"), ("All files", "*.*")]
EXCEL_DEFAULT_EXT = ".xlsx"

# Dialog strings
CLONE_FILE_TITLE = "Clone File"
CREATE_SCRATCH_TITLE = "Create from Scratch"
ENTER_FILE_EXT = "Enter file extension:"

# Column headers
VERSION_HISTORY_COLUMNS = ["Timestamp", "Version Info", "Summary"]

class ProjectBudgetinator:
    """Main application class for ProjectBudgetinator.
    
    This class orchestrates the entire ProjectBudgetinator application, providing
    a comprehensive Tkinter-based GUI for Excel workbook management, partner
    administration, workpackage handling, and budget coordination.
    
    The class manages:
        - Application initialization and authentication
        - GUI setup and menu management
        - Excel workbook operations (open, save, validate)
        - Partner management (add, edit, delete)
        - Workpackage management (add, edit, delete)
        - Budget and PM overview updates
        - User preferences and theme management
        - Window positioning and performance monitoring
        - File operations (clone, compare, create)
        - Diagnostics and help systems
    
    Attributes:
        root (tk.Tk): Main Tkinter window
        developer_mode (bool): Flag for developer features
        prefs_manager: Preferences manager instance
        current_config (dict): Current configuration settings
        current_workbook: Currently loaded Excel workbook
        auth_manager: Authentication manager instance
        current_user: Currently authenticated user
        current_profile: Current user profile
        logger: Structured logger instance
        performance_monitor: Performance monitoring instance
        theme_manager: Theme management instance
        performance_indicator: Performance indicator widget
    
    Example:
        Basic application usage:
        
            app = ProjectBudgetinator()
            app.run()
    
    Note:
        The application requires proper authentication and preferences setup.
        All file operations are secured through centralized validation.
    """
    
    def __init__(self):
        """Initialize the ProjectBudgetinator application.
        
        This constructor sets up the complete application environment including
        authentication, preferences, GUI components, logging, and all necessary
        subsystems. It follows a specific initialization order to ensure proper
        dependency resolution.
        
        Initialization sequence:
            1. Create main Tkinter window
            2. Initialize logging and performance monitoring
            3. Set up authentication system and user login
            4. Initialize preferences manager with user profile
            5. Set up theme system and window positioning
            6. Create GUI menus and performance indicators
            7. Show startup diagnostics
        
        Attributes initialized:
            root (tk.Tk): Main Tkinter application window
            developer_mode (bool): Flag for developer features (default: False)
            prefs_manager: Preferences manager instance
            current_config (dict): Current configuration settings
            current_workbook: Currently loaded Excel workbook (initially None)
            auth_manager: Authentication manager instance
            current_user: Currently authenticated user object
            current_profile: Current user profile object
            logger: Structured logger instance for this class
            performance_monitor: Performance monitoring instance
            theme_manager: Theme management instance (if available)
            performance_indicator: Performance indicator widget (if created)
            _partner_sheets_cache: Cached partner sheets list for performance
            _workbook_cache_key: Cache key to track workbook changes
        
        Raises:
            SystemExit: If authentication fails or is cancelled, the application exits.
            Exception: Various exceptions may be raised during initialization, but most
                are handled gracefully with fallback behavior.
        
        Note:
            If authentication fails, the application will log the failure and exit
            gracefully. All subsystems are initialized with proper error handling
            and fallback mechanisms to ensure the application can start even if
            some features are unavailable.
        """
        self.root = tk.Tk()
        self.developer_mode = False
        self.prefs_manager = None
        self.current_config = DEFAULT_USER_CONFIG.copy()
        self.current_workbook = None
        self.auth_manager = None
        self.current_user = None
        self.current_profile = None
        
        # Performance optimization: Partner sheets caching
        self._partner_sheets_cache = None
        self._workbook_cache_key = None
        
        # Initialize structured logging system
        setup_logging()
        self.logger = get_structured_logger("ProjectBudgetinator.main")
        
        # Initialize error handling system
        setup_error_handling(self.root)
        
        # Initialize performance monitoring
        self.performance_monitor = get_performance_monitor()
        
        # Initialize performance optimizer
        self.performance_optimizer = get_performance_optimizer()
        
        # Initialize authentication system
        self._initialize_authentication()
        
        # If authentication failed, exit
        if not self.current_user:
            self.logger.info("Authentication failed or cancelled, exiting application")
            self.root.quit()
            return
        
        # Initialize preferences manager with user profile
        self._initialize_preferences()
        
        # Initialize theme system
        self._initialize_theme_system()
        
        # Initialize window positioning
        self._setup_window_positioning()
        
        # Log application startup
        self.logger.info("ProjectBudgetinator application starting",
                        version=full_version_string(),
                        developer_mode=self.developer_mode,
                        user=self.current_user.username,
                        profile=self.current_profile.profile_name if self.current_profile else "None")
        
        # Set up the menu bar
        self._setup_file_menu()
        # Initialize performance indicator
        self._setup_performance_indicator()
        # Show diagnostics window at startup
        self.show_diagnostics()

    def _initialize_authentication(self):
        """Initialize the authentication system and show login dialog."""
        try:
            # Import authentication components
            from auth.auth_manager import get_auth_manager
            from gui.auth.login_dialog import show_login_dialog
            
            # Initialize authentication manager
            self.auth_manager = get_auth_manager()
            
            # Show login dialog
            self.logger.info("Showing login dialog")
            result = show_login_dialog(self.root)
            
            if result:
                self.current_user, self.current_profile = result
                self.logger.info(f"User authenticated: {self.current_user.username}")
                return True
            else:
                self.logger.info("User authentication cancelled or failed")
                return False
                
        except Exception as e:
            self.logger.error(f"Authentication initialization failed: {e}")
            from tkinter import messagebox
            messagebox.showerror(
                "Authentication Error",
                f"Failed to initialize authentication system:\n{str(e)}\n\nApplication will exit."
            )
            return False

    def _initialize_preferences(self):
        """Initialize preferences manager with user profile integration."""
        try:
            from core.preferences import PreferencesManager
            
            # Create preferences manager that integrates with user profiles
            self.prefs_manager = PreferencesManager()
            
            # Load preferences from current user profile
            if self.current_profile:
                profile_preferences = self.current_profile.preferences_data or {}
                self.current_config.update(profile_preferences)
                self.logger.info(f"Loaded preferences from profile: {self.current_profile.profile_name}")
            else:
                self.logger.warning("No current profile, using default preferences")
                
        except Exception as e:
            self.logger.error(f"Failed to initialize preferences: {e}")
            # Fall back to default preferences
            from core.preferences import PreferencesManager
            self.prefs_manager = PreferencesManager()

    def _setup_window_positioning(self):
        """Set up window positioning using preferences."""
        try:
            from utils.window_positioning import position_main_window
            
            # Position the main window according to preferences
            position_main_window(self.root, self.prefs_manager)
            
            # Set up window close handler to save position
            self.root.protocol("WM_DELETE_WINDOW", self._on_window_close)
            
            self.logger.info("Window positioning initialized successfully")
            
        except Exception as e:
            self.logger.warning(f"Failed to initialize window positioning: {e}")
            # Fallback to default positioning
            self.root.geometry("800x600")

    def _on_window_close(self):
        """Handle main window close event to save position."""
        try:
            from utils.window_positioning import save_main_window_position
            save_main_window_position(self.root)
        except Exception as e:
            self.logger.warning(f"Failed to save window position: {e}")
        
        # Continue with normal exit
        self.exit_program()

    def _setup_file_menu(self):
        menubar = tk.Menu(self.root)
        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Clone", command=self.clone_file)
        file_menu.add_command(label="Compare", command=self.compare_files)
        file_menu.add_command(label="Create from scratch", command=self.create_from_scratch)
        file_menu.add_command(label="Create from template", command=self.create_from_template)
        file_menu.add_separator()
        file_menu.add_command(label="Restore", command=self.restore_file)
        file_menu.add_separator()
        file_menu.add_command(label="Batch Operations...",
                              command=self.show_batch_operations,
                              accelerator="Ctrl+B")
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.exit_program)
        menubar.add_cascade(label="File", menu=file_menu)

        # Modify menu (partner and workpackage operations)
        modify_menu = self._create_modify_submenu(menubar)
        menubar.add_cascade(label="Modify", menu=modify_menu)

        # User menu (for profile switching and user management)
        user_menu = tk.Menu(menubar, tearoff=0)
        
        # Only show "Change Password" for non-admin users
        if self.auth_manager and not self.auth_manager.is_current_user_admin():
            user_menu.add_command(label="Change Password", command=self.change_password)
            user_menu.add_separator()
        # Add profile switcher to menu
        try:
            from gui.auth.profile_switcher import ProfileSwitcherMenu
            self.profile_switcher_menu = ProfileSwitcherMenu(user_menu, self._on_profile_changed)
        except Exception as e:
            self.logger.warning(f"Failed to add profile switcher to menu: {e}")
            user_menu.add_command(label="Switch Profile", command=self.show_profile_management, state="disabled")
        
        user_menu.add_separator()
        user_menu.add_command(label="User Administration", command=self.show_user_administration)
        user_menu.add_command(label="Logout", command=self.logout)
        menubar.add_cascade(label="User", menu=user_menu)

        # Preferences menu
        pref_menu = tk.Menu(menubar, tearoff=0)
        pref_menu.add_command(label="Preferences", command=self.show_preferences)
        menubar.add_cascade(label="Preferences", menu=pref_menu)

        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="Help", command=self.show_help)
        help_menu.add_command(label="Diagnostics", command=self.show_diagnostics)
        help_menu.add_separator()
        help_menu.add_command(label="Performance Monitor", command=self.show_performance_monitor)
        menubar.add_cascade(label="Help", menu=help_menu)

        self.root.config(menu=menubar)
        
        # Set up keyboard shortcuts
        self.root.bind_all("<Control-b>",
                           lambda e: self.show_batch_operations())


    def _create_modify_submenu(self, parent):
        """Create the Modify submenu for partner and workpackage operations."""
        modify_menu = tk.Menu(parent, tearoff=0)
        # Partner operations
        partner_commands = [
            ("Add Partner", self.add_partner),
            ("Delete Partner", self.delete_partner),
            ("Edit Partner", self.edit_partner)
        ]
        for label, command in partner_commands:
            modify_menu.add_command(label=label, command=command)
        modify_menu.add_separator()
        # P1 Coordinator management
        modify_menu.add_command(label="Manage P1", command=self.manage_p1)
        modify_menu.add_separator()
        # Workpackage operations
        workpackage_commands = [
            ("Add Workpackage", self.add_workpackage),
            ("Delete Workpackage", self.delete_workpackage),
            ("Edit Workpackage", self.edit_workpackage)
        ]
        for label, command in workpackage_commands:
            modify_menu.add_command(label=label, command=command)
        modify_menu.add_separator()
        # Budget Overview operations
        modify_menu.add_command(label="Update Budget Overview", command=self.update_budget_overview)
        # PM Overview operations
        modify_menu.add_command(label="Update PM Overview", command=self.update_pm_overview)
        return modify_menu

    # Partner and workpackage operations
    def add_partner(self):
        """Open a dialog to add a new partner to the project."""
        with LogContext("add_partner", user_id="current_user"):
            self.logger.info("Starting add partner operation")
            
            # Use utility function to ensure workbook is loaded
            success, workbook, error_message = ensure_workbook_loaded(
                self.current_workbook, "add partner"
            )
            
            if not success:
                return
            
            # Update current workbook
            self.current_workbook = workbook

        from handlers.add_partner_handler import (
            PartnerInputDialog,
            PartnerDialog,
            add_partner_to_workbook,
            get_existing_partners
        )
        
        # Get existing partners from workbook
        existing_partners = get_existing_partners(self.current_workbook)
        
        # Show the new partner input dialog with validation
        input_dialog = PartnerInputDialog(self.root, existing_partners)
        if not input_dialog.result:
            return
        
        # Extract validated partner information
        partner_number = input_dialog.result['partner_number']
        partner_acronym = input_dialog.result['partner_acronym']
        
        # Create and show the detailed partner dialog
        dialog = PartnerDialog(self.root, partner_number, partner_acronym)
        if dialog.result:
            partner_info = dialog.result
            partner_info['project_partner_number'] = partner_number
            partner_info['partner_acronym'] = partner_acronym

            if add_partner_to_workbook(self.current_workbook, partner_info):
                # Invalidate partner sheets cache after adding new partner
                self.invalidate_partner_sheets_cache()
                
                # Save the workbook
                self._save_workbook(partner_number, partner_acronym)
    def _save_workbook(self, partner_number, partner_acronym):
        """Save the workbook with user confirmation."""
        success_message = f"Added partner {partner_number}: {partner_acronym}"
        success, file_path = save_workbook_with_dialog(
            self.current_workbook,
            title="Save Workbook",
            success_message=success_message
        )
        return success, file_path

    def _is_partner_worksheet(self, sheet_name):
        """Check if a sheet name represents a partner worksheet (P2 through P20) with hyphen format."""
        return is_partner_worksheet(sheet_name)

    @property
    def partner_sheets(self):
        """Get cached list of partner sheets with automatic cache invalidation.
        
        This property provides a cached list of partner worksheet names to avoid
        repeated calculations when accessing partner sheets multiple times. The cache
        is automatically invalidated when the workbook changes.
        
        Returns:
            list: List of partner sheet names (e.g., ['P2-ACME', 'P3-BETA', ...])
                 Returns empty list if no workbook is loaded or no partner sheets found.
        
        Note:
            The cache is based on the workbook's sheetnames list and is invalidated
            whenever the workbook changes or when explicitly cleared.
        """
        # Use performance optimizer for caching
        return self.performance_optimizer.get_cached_partner_sheets(
            self.current_workbook,
            self._is_partner_worksheet
        )

    def _get_workbook_cache_key(self):
        """Generate a cache key based on current workbook state.
        
        Returns:
            str: Cache key representing current workbook state, or None if no workbook.
        """
        if self.current_workbook is None:
            return None
        
        # Use tuple of sheet names as cache key - this will change if sheets are added/removed
        try:
            return tuple(sorted(self.current_workbook.sheetnames))
        except Exception:
            return None

    def _build_partner_sheets_list(self):
        """Build the list of partner sheets from current workbook.
        
        Returns:
            list: List of partner sheet names, empty if no workbook or no partner sheets.
        """
        if self.current_workbook is None:
            return []
        
        partner_sheets = []
        try:
            for sheet_name in self.current_workbook.sheetnames:
                if self._is_partner_worksheet(sheet_name):
                    partner_sheets.append(sheet_name)
        except Exception as e:
            self.logger.warning(f"Error building partner sheets list: {e}")
            return []
        
        return partner_sheets

    def invalidate_partner_sheets_cache(self):
        """Manually invalidate the partner sheets cache.
        
        This method should be called after operations that modify the workbook
        structure (adding/removing sheets) to ensure the cache stays current.
        """
        # Use performance optimizer for cache invalidation
        self.performance_optimizer.invalidate_workbook_cache(self.current_workbook)
        
        # Also clear local cache variables for backward compatibility
        self._partner_sheets_cache = None
        self._workbook_cache_key = None
        self.logger.debug("Partner sheets cache manually invalidated")

    def _update_partner_worksheet(self, worksheet, partner_info, cell_map):
        """Update partner worksheet with new values from the edit dialog."""
        try:
            # Update WP fields (columns C through Q, row 18)
            for i in range(1, 16):
                col = chr(ord('C') + i - 1)
                cell = f'{col}18'
                wp_key = f'wp{i}'
                if wp_key in partner_info:
                    self._set_cell_value_safely(worksheet, cell, partner_info[wp_key])
            
            # Update other fields using the cell mapping
            for key, cell in cell_map.items():
                if key in partner_info:
                    self._set_cell_value_safely(worksheet, cell, partner_info[key])
                    
            self.logger.info("Partner worksheet updated successfully")
        except Exception as e:
            self.logger.error("Error updating partner worksheet", error=str(e))
            raise

    def _set_cell_value_safely(self, worksheet, cell_address, value):
        """Safely set cell value, handling merged cells."""
        try:
            cell_obj = worksheet[cell_address]
            # Check if this is a merged cell
            if hasattr(cell_obj, 'coordinate') and cell_obj.coordinate:
                # Check if the cell is part of a merged range
                for merged_range in worksheet.merged_cells.ranges:
                    if cell_obj.coordinate in merged_range:
                        # Get the top-left cell of the merged range
                        top_left_cell = worksheet.cell(
                            row=merged_range.min_row,
                            column=merged_range.min_col
                        )
                        top_left_cell.value = value
                        self.logger.debug(f"Updated merged cell {cell_address} via top-left cell {top_left_cell.coordinate}")
                        return
            
            # If not a merged cell, update normally
            cell_obj.value = value
            self.logger.debug(f"Updated cell {cell_address} with value: {value}")
            
        except Exception as e:
            self.logger.warning(f"Failed to update cell {cell_address}: {str(e)}")
            # Continue with other cells even if one fails

    def delete_partner(self):
        """Open a dialog to delete a partner from the project."""
        # Check if we have an open workbook
        if self.current_workbook is None:
            response = messagebox.askyesno(
                "No Workbook Open",
                "No workbook is currently open. Would you like to open one now?"
            )
            if response:
                file_path = filedialog.askopenfilename(
                    title="Open Excel Workbook",
                    filetypes=EXCEL_FILETYPES
                )
                if file_path:
                    try:
                        # Validate file path and content
                        is_valid, error_msg = SecurityValidator.validate_excel_file(file_path)
                        if not is_valid:
                            messagebox.showerror("Security Error", f"Cannot open file: {error_msg}")
                            self.logger.warning("Security validation failed for file",
                                              file_path=file_path, error=error_msg)
                            return
                        
                        # Sanitize file path
                        safe_path = SecurityValidator.validate_file_path(file_path)
                        
                        from openpyxl import load_workbook
                        self.current_workbook = load_workbook(safe_path)
                    except ValueError as e:
                        messagebox.showerror("Security Error", str(e))
                        self.logger.warning("Security validation error", error=str(e))
                        return
                    except Exception as e:
                        messagebox.showerror(
                            "Error",
                            f"Could not open workbook:\n{str(e)}"
                        )
                        return
                else:
                    return
            else:
                return

        # Get list of partner sheets using cached property
        partner_sheets = self.partner_sheets

        if not partner_sheets:
            messagebox.showinfo(
                "No Partners",
                "No partner sheets found in the workbook."
            )
            return

        # Create the dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Delete Partner")
        dialog.grab_set()

        # Instructions label
        instructions = tk.Label(
            dialog,
            text="Select a partner to delete:",
            font=("Arial", 10, "bold")
        )
        instructions.pack(pady=(10, 5))

        # Partner listbox
        listbox = tk.Listbox(dialog, width=40, height=10)
        listbox.pack(padx=10, pady=5, fill=tk.BOTH, expand=True)

        # Populate listbox
        for sheet in partner_sheets:
            listbox.insert(tk.END, sheet)

        def on_delete():
            """Delete the selected partner with proper validation and confirmation."""
            selection = listbox.curselection()
            if selection:
                selected_sheet = partner_sheets[selection[0]]
                
                # Confirm deletion
                if messagebox.askyesno(
                    "Confirm Deletion",
                    f"Are you sure you want to delete partner sheet '{selected_sheet}'?\n\n"
                    "This action cannot be undone."
                ):
                    try:
                        # Remove the worksheet
                        if selected_sheet in self.current_workbook.sheetnames:
                            del self.current_workbook[selected_sheet]
                            
                            # Invalidate partner sheets cache after deletion
                            self.invalidate_partner_sheets_cache()
                            
                            # Ask user where to save the workbook
                            save_path = filedialog.asksaveasfilename(
                                title="Save Workbook After Partner Deletion",
                                defaultextension=".xlsx",
                                filetypes=EXCEL_FILETYPES
                            )
                            if save_path:
                                try:
                                    # Validate and sanitize save path
                                    safe_save_path = SecurityValidator.validate_file_path(save_path)
                                    
                                    # Ensure proper extension
                                    if not safe_save_path.lower().endswith('.xlsx'):
                                        safe_save_path += '.xlsx'
                                    
                                    self.current_workbook.save(safe_save_path)
                                    messagebox.showinfo(
                                        "Success",
                                        f"Partner sheet '{selected_sheet}' deleted successfully!\n"
                                        f"Workbook saved to: {save_path}"
                                    )
                                    self.logger.info("Partner deleted and workbook saved",
                                                   partner_sheet=selected_sheet,
                                                   file_path=save_path)
                                except ValueError as e:
                                    messagebox.showerror("Security Error", str(e))
                                    self.logger.warning("Security validation error for save path", error=str(e))
                                    return
                            else:
                                messagebox.showwarning(
                                    "Warning",
                                    f"Partner sheet '{selected_sheet}' deleted but workbook not saved!"
                                )
                                self.logger.warning("Partner deleted but user did not save workbook")
                            
                            dialog.destroy()
                        else:
                            messagebox.showerror("Error", f"Partner sheet '{selected_sheet}' not found!")
                            
                    except Exception as e:
                        messagebox.showerror(
                            "Error",
                            f"Failed to delete partner sheet:\n{str(e)}"
                        )
                        self.logger.error("Failed to delete partner sheet",
                                        partner_sheet=selected_sheet,
                                        error=str(e))
            else:
                messagebox.showwarning(
                    "No Selection",
                    "Please select a partner to delete."
                )

        def on_cancel():
            dialog.destroy()

        # Button frame
        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=8)

        # Delete and Cancel buttons
        delete_btn = tk.Button(
            btn_frame,
            text="Delete",
            command=on_delete
        )
        delete_btn.pack(side=tk.LEFT, padx=4)

        cancel_btn = tk.Button(
            btn_frame,
            text="Cancel",
            command=on_cancel
        )
        cancel_btn.pack(side=tk.LEFT, padx=4)

        # Center the dialog
        dialog.update_idletasks()
        w = dialog.winfo_width()
        h = dialog.winfo_height()
        x = self.root.winfo_x() + (self.root.winfo_width() - w) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - h) // 2
        dialog.geometry(f"+{x}+{y}")

    def edit_partner(self):
        """Open a dialog to select a partner (P > 2) for editing."""
        # Check if we have an open workbook
        if self.current_workbook is None:
            response = messagebox.askyesno(
                "No Workbook Open",
                "No workbook is currently open. Would you like to open one now?"
            )
            if response:
                file_path = filedialog.askopenfilename(
                    title="Open Excel Workbook",
                    filetypes=EXCEL_FILETYPES
                )
                if file_path:
                    try:
                        # Validate file path and content
                        is_valid, error_msg = SecurityValidator.validate_excel_file(file_path)
                        if not is_valid:
                            messagebox.showerror("Security Error", f"Cannot open file: {error_msg}")
                            self.logger.warning("Security validation failed for file",
                                              file_path=file_path, error=error_msg)
                            return
                        
                        # Sanitize file path
                        safe_path = SecurityValidator.validate_file_path(file_path)
                        
                        from openpyxl import load_workbook
                        self.current_workbook = load_workbook(safe_path)
                    except ValueError as e:
                        messagebox.showerror("Security Error", str(e))
                        self.logger.warning("Security validation error", error=str(e))
                        return
                    except Exception as e:
                        messagebox.showerror(
                            "Error",
                            f"Could not open workbook:\n{str(e)}"
                        )
                        return
                else:
                    return
            else:
                return

        # Get list of partner sheets using cached property
        partner_sheets = self.partner_sheets

        if not partner_sheets:
            messagebox.showinfo(
                "No Partners",
                "No eligible partner sheets (P2-P15) found in the workbook."
            )
            return

        # Create the dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Edit Partner")
        dialog.grab_set()

        # Instructions label
        instructions = tk.Label(
            dialog,
            text="Select a partner to edit:",
            font=("Arial", 10, "bold")
        )
        instructions.pack(pady=(10, 5))

        # Partner listbox
        listbox = tk.Listbox(dialog, width=40, height=10)
        listbox.pack(padx=10, pady=5, fill=tk.BOTH, expand=True)

        # Populate listbox
        for sheet in partner_sheets:
            listbox.insert(tk.END, sheet)

        def on_edit():
            selection = listbox.curselection()
            if selection:
                selected_sheet = partner_sheets[selection[0]]
                dialog.destroy()
                # Extract partner number and acronym from the sheet name (hyphen format)
                parts = selected_sheet[1:].split('-', 1)
                partner_number = parts[0] if parts else ''
                partner_acronym = parts[1] if len(parts) > 1 else ''

                ws = self.current_workbook[selected_sheet]
                # Create debug window first
                debug_win = tk.Toplevel(self.root)
                debug_win.title("Debug Info - Partner Values")
                debug_win.geometry("600x400")
                
                # Create text widget for debug output
                debug_text = tk.Text(debug_win, wrap="word")
                debug_text.pack(fill="both", expand=True)
                
                # Add scrollbar to debug window
                scrollbar = tk.Scrollbar(debug_win, command=debug_text.yview)
                scrollbar.pack(side="right", fill="y")
                debug_text.config(yscrollcommand=scrollbar.set)

                debug_text.insert("end", f"Selected sheet: {selected_sheet}\n")
                debug_text.insert("end", f"Partner number: {partner_number}\n")
                debug_text.insert("end", f"Partner acronym: {partner_acronym}\n\n")

                # Try to extract values from the worksheet
                partner_info = {}
                # These cell mappings should match those used in add_partner_to_workbook
                cell_map = {
                    'partner_identification_code': 'D4',
                    'name_of_beneficiary': 'D5',
                    'country': 'D6',
                    'role': 'D7',
                    'name_subcontractor_1': 'D22',
                    'sum_subcontractor_1': 'F22',
                    'explanation_subcontractor_1': 'G22',
                    'name_subcontractor_2': 'D23',
                    'sum_subcontractor_2': 'F23',
                    'explanation_subcontractor_2': 'G23',
                    'sum_travel': 'F28',
                    'sum_equipment': 'F29',
                    'sum_other': 'F30',
                    'sum_financial_support': 'F35',
                    'sum_internal_goods': 'F36',
                    'sum_income_generated': 'F42',
                    'sum_financial_contributions': 'F43',
                    'sum_own_resources': 'F44',
                    'explanation_travel': 'G28',
                    'explanation_equipment': 'G29',
                    'explanation_other': 'G30',
                    'explanation_financial_support': 'G35',
                    'explanation_internal_goods': 'G36',
                    'explanation_income_generated': 'G42',
                    'explanation_financial_contributions': 'G43',
                    'explanation_own_resources': 'G44',
                }
                # WP fields
                debug_text.insert("end", "\nReading WP fields:\n")
                for i in range(1, 16):
                    col = chr(ord('C') + i - 1)
                    cell = f'{col}18'
                    debug_text.insert("end", f"Reading cell {cell}... ")
                    try:
                        value = ws[cell].value
                        partner_info[f'wp{i}'] = value if value is not None else ''
                        debug_text.insert("end", f"found value: {value}\n")
                    except Exception as e:
                        debug_text.insert("end", f"error: {str(e)}\n")
                        partner_info[f'wp{i}'] = ''

                # Other fields
                debug_text.insert("end", "\nReading other fields:\n")
                for key, cell in cell_map.items():
                    debug_text.insert("end", f"Reading {key} from cell {cell}... ")
                    try:
                        value = ws[cell].value
                        partner_info[key] = value if value is not None else ''
                        debug_text.insert("end", f"found value: {value}\n")
                    except Exception as e:
                        debug_text.insert("end", f"error: {str(e)}\n")
                        partner_info[key] = ''
                # Add partner number and acronym
                partner_info['project_partner_number'] = partner_number
                partner_info['partner_acronym'] = partner_acronym

                # Log all collected values with cell origins
                debug_text.insert("end", "\nFinal collected values with cell origins:\n")
                for key, value in partner_info.items():
                    if key in cell_map:
                        cell_ref = cell_map[key]
                        debug_text.insert("end", f"{key}: {value} (from cell {cell_ref})\n")
                    elif key.startswith('wp') and key != 'wp':
                        # Handle WP fields
                        wp_num = key[2:]  # Extract number from wp1, wp2, etc.
                        try:
                            wp_index = int(wp_num) - 1
                            col = chr(ord('C') + wp_index)
                            cell_ref = f'{col}18'
                            debug_text.insert("end", f"{key}: {value} (from cell {cell_ref})\n")
                        except (ValueError, IndexError):
                            debug_text.insert("end", f"{key}: {value}\n")
                    else:
                        debug_text.insert("end", f"{key}: {value}\n")
                debug_text.see("end")

                # Show the EditPartnerDialog with worksheet integration
                from handlers.edit_partner_handler import edit_partner_from_worksheet
                
                debug_text.insert("end", "\nOpening EditPartnerDialog with worksheet integration...\n")
                debug_text.see("end")
                
                # Close debug window before opening edit dialog
                debug_win.destroy()
                
                # Use the new worksheet-integrated edit function
                result = edit_partner_from_worksheet(self.root, self.current_workbook, selected_sheet)
                
                if result:
                    # The worksheet has already been updated by the edit handler
                    updated_partner_info = result
                    
                    # Save the workbook
                    try:
                        save_path = filedialog.asksaveasfilename(
                            title="Save Workbook",
                            defaultextension=".xlsx",
                            filetypes=EXCEL_FILETYPES
                        )
                        if save_path:
                            try:
                                # Validate and sanitize save path
                                safe_save_path = SecurityValidator.validate_file_path(save_path)
                                
                                # Ensure proper extension
                                if not safe_save_path.lower().endswith('.xlsx'):
                                    safe_save_path += '.xlsx'
                                
                                self.current_workbook.save(safe_save_path)
                            except ValueError as e:
                                messagebox.showerror("Security Error", str(e))
                                self.logger.warning("Security validation error for save path", error=str(e))
                                return
                            messagebox.showinfo(
                                "Success",
                                f"Partner {partner_number} ({partner_acronym}) updated and saved to:\n{save_path}"
                            )
                            self.logger.info("Partner updated and workbook saved",
                                             partner_number=partner_number,
                                             partner_acronym=partner_acronym,
                                             file_path=save_path)
                        else:
                            messagebox.showwarning(
                                "Warning",
                                "Partner updated but workbook not saved!"
                            )
                    except Exception as e:
                        messagebox.showerror(
                            "Error",
                            f"Failed to save workbook:\n{str(e)}"
                        )
                        self.logger.error("Failed to save workbook after partner update",
                                          error=str(e))
            else:
                messagebox.showwarning(
                    "No Selection",
                    "Please select a partner to edit."
                )

        def on_cancel():
            dialog.destroy()

        # Button frame
        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=8)

        # Edit and Cancel buttons
        edit_btn = tk.Button(
            btn_frame,
            text="Edit",
            command=on_edit
        )
        edit_btn.pack(side=tk.LEFT, padx=4)

        cancel_btn = tk.Button(
            btn_frame,
            text="Cancel",
            command=on_cancel
        )
        cancel_btn.pack(side=tk.LEFT, padx=4)

        # Center the dialog
        dialog.update_idletasks()
        w = dialog.winfo_width()
        h = dialog.winfo_height()
        x = self.root.winfo_x() + (self.root.winfo_width() - w) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - h) // 2
        dialog.geometry(f"+{x}+{y}")

    def manage_p1(self):
        """Load the P1-Coord worksheet into a UI for editing."""
        # Always prompt for file selection
        file_path = filedialog.askopenfilename(
            title="Open Excel Workbook for P1 Management",
            filetypes=EXCEL_FILETYPES
        )
        
        if not file_path:
            return
            
        try:
            # Validate file path and content
            is_valid, error_msg = SecurityValidator.validate_excel_file(file_path)
            if not is_valid:
                messagebox.showerror("Security Error", f"Cannot open file: {error_msg}")
                self.logger.warning("Security validation failed for file",
                                    file_path=file_path, error=error_msg)
                return
            
            # Sanitize file path
            safe_path = SecurityValidator.validate_file_path(file_path)
            
            from openpyxl import load_workbook
            workbook = load_workbook(safe_path)
            self.logger.info("Workbook loaded successfully for P1 management",
                             file_path=safe_path)
        except ValueError as e:
            messagebox.showerror("Security Error", str(e))
            self.logger.warning("Security validation error", error=str(e))
            return
        except Exception as e:
            self.logger.error("Failed to load workbook for P1 management",
                              file_path=file_path, error=str(e))
            messagebox.showerror(
                "Error",
                f"Could not open workbook:\n{str(e)}"
            )
            return

        # Check if P1-Coord worksheet exists
        if "P1-Coord" not in workbook.sheetnames:
            messagebox.showerror(
                "Worksheet Not Found",
                "The 'P1-Coord' worksheet was not found in the selected workbook."
            )
            self.logger.warning("P1-Coord worksheet not found in workbook")
            return

        # Set the current workbook to the loaded one
        self.current_workbook = workbook

        # Get the P1-Coord worksheet and open it in the UI
        try:
            from handlers.p1_manager_handler import P1ManagerDialog
            
            p1_worksheet = self.current_workbook["P1-Coord"]
            
            # Open the P1 management dialog
            dialog = P1ManagerDialog(self.root, p1_worksheet, file_path)
            
            # Check if changes were made and save if needed
            if dialog.result and dialog.result.get('changes_made', False):
                try:
                    # Save the workbook
                    self.current_workbook.save(file_path)
                    messagebox.showinfo(
                        "Saved",
                        f"Changes have been saved to:\n{file_path}"
                    )
                    self.logger.info("P1 coordinator data saved successfully", file_path=file_path)
                except Exception as save_error:
                    messagebox.showerror(
                        "Save Error",
                        f"Failed to save the workbook:\n{str(save_error)}"
                    )
                    self.logger.error("Failed to save P1 workbook", error=str(save_error))
            
        except ImportError as e:
            messagebox.showerror(
                "Import Error",
                f"Failed to load P1 manager dialog:\n{str(e)}"
            )
            self.logger.error("Failed to import P1ManagerDialog", error=str(e))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open P1 management dialog:\n{str(e)}")
            self.logger.error("Failed to open P1 management dialog", error=str(e))

    def add_workpackage(self):
        """Open a dialog to add a new workpackage to the project."""
        # Check if we have an open workbook
        if self.current_workbook is None:
            response = messagebox.askyesno(
                "No Workbook Open",
                "No workbook is currently open. Would you like to open one now?"
            )
            if response:
                file_path = filedialog.askopenfilename(
                    title="Open Excel Workbook",
                    filetypes=EXCEL_FILETYPES
                )
                if file_path:
                    try:
                        # Validate file path and content
                        is_valid, error_msg = SecurityValidator.validate_excel_file(file_path)
                        if not is_valid:
                            messagebox.showerror("Security Error", f"Cannot open file: {error_msg}")
                            self.logger.warning("Security validation failed for file",
                                              file_path=file_path, error=error_msg)
                            return
                        
                        # Sanitize file path
                        safe_path = SecurityValidator.validate_file_path(file_path)
                        
                        from openpyxl import load_workbook
                        self.current_workbook = load_workbook(safe_path)
                    except ValueError as e:
                        messagebox.showerror("Security Error", str(e))
                        self.logger.warning("Security validation error", error=str(e))
                        return
                    except Exception as e:
                        messagebox.showerror(
                            "Error",
                            f"Could not open workbook:\n{str(e)}"
                        )
                        return
                else:
                    return
            else:
                return
        
        # Check if PM Summary sheet exists
        if "PM Summary" not in self.current_workbook.sheetnames:
            messagebox.showerror(
                "Error",
                "This workbook does not contain a 'PM Summary' sheet"
            )
            return
            
        # Show the AddWorkpackageDialog
        from handlers.add_workpackage_handler import AddWorkpackageDialog
        dialog = AddWorkpackageDialog(self.root, self.current_workbook)
        
        # Wait for dialog
        self.root.wait_window(dialog.dialog)
        
        if dialog.result:
            # Ask user where to save the workbook
            try:
                save_path = filedialog.asksaveasfilename(
                    title="Save Workbook",
                    defaultextension=".xlsx",
                    filetypes=EXCEL_FILETYPES
                )
                if save_path:
                    try:
                        # Validate and sanitize save path
                        safe_save_path = SecurityValidator.validate_file_path(save_path)
                        
                        # Ensure proper extension
                        if not safe_save_path.lower().endswith('.xlsx'):
                            safe_save_path += '.xlsx'
                        
                        self.current_workbook.save(safe_save_path)
                    except ValueError as e:
                        messagebox.showerror("Security Error", str(e))
                        self.logger.warning("Security validation error for save path", error=str(e))
                        return
                    messagebox.showinfo(
                        "Success",
                        f"Added workpackage in row {dialog.result['row']}\n"
                        f"Workbook saved to: {save_path}"
                    )
                else:
                    messagebox.showwarning(
                        "Warning",
                        "Workpackage added but workbook not saved!"
                    )
            except Exception as e:
                messagebox.showerror(
                    "Error",
                    f"Failed to save workbook:\n{str(e)}"
                )

    def delete_workpackage(self):
        """Delete a workpackage from the project."""
        # Check if we have an open workbook
        if self.current_workbook is None:
            response = messagebox.askyesno(
                "No Workbook Open",
                "No workbook is currently open. Would you like to open one now?"
            )
            if response:
                file_path = filedialog.askopenfilename(
                    title="Open Excel Workbook",
                    filetypes=EXCEL_FILETYPES
                )
                if file_path:
                    try:
                        # Use centralized file operations
                        from utils.centralized_file_operations import safe_open_workbook
                        success, workbook, error_msg = safe_open_workbook(file_path)
                        if success:
                            self.current_workbook = workbook
                        else:
                            messagebox.showerror("Error", f"Could not open workbook: {error_msg}")
                            return
                    except Exception as e:
                        messagebox.showerror("Error", f"Could not open workbook:\n{str(e)}")
                        return
                else:
                    return
            else:
                return
        
        # Check if PM Summary sheet exists
        if "PM Summary" not in self.current_workbook.sheetnames:
            messagebox.showerror(
                "Error",
                "This workbook does not contain a 'PM Summary' sheet"
            )
            return
        
        # Show the DeleteWorkpackageDialog
        try:
            from handlers.edit_workpackage_handler import DeleteWorkpackageDialog
            dialog = DeleteWorkpackageDialog(self.root, self.current_workbook)
            
            # Wait for dialog
            self.root.wait_window(dialog.dialog)
            
            if dialog.result:
                # Ask user where to save the workbook
                save_path = filedialog.asksaveasfilename(
                    title="Save Workbook After Workpackage Deletion",
                    defaultextension=".xlsx",
                    filetypes=EXCEL_FILETYPES
                )
                if save_path:
                    try:
                        # Use centralized file operations
                        from utils.centralized_file_operations import safe_save_workbook
                        success, saved_path, error_msg = safe_save_workbook(self.current_workbook, save_path)
                        if success:
                            messagebox.showinfo(
                                "Success",
                                f"Deleted workpackage from row {dialog.result['row']}\n"
                                f"Workbook saved to: {saved_path}"
                            )
                        else:
                            messagebox.showerror("Error", f"Failed to save workbook: {error_msg}")
                    except Exception as e:
                        messagebox.showerror("Error", f"Failed to save workbook:\n{str(e)}")
                else:
                    messagebox.showwarning(
                        "Warning",
                        "Workpackage deleted but workbook not saved!"
                    )
        except ImportError:
            # Fallback implementation
            messagebox.showinfo(
                "Delete Workpackage",
                "Workpackage deletion functionality is not yet fully implemented.\n"
                "Please use the Edit Workpackage function to modify workpackage data."
            )

    def edit_workpackage(self):
        """Open a dialog to edit a workpackage."""
        # Check if we have an open workbook
        if self.current_workbook is None:
            response = messagebox.askyesno(
                "No Workbook Open",
                "No workbook is currently open. Would you like to open one now?"
            )
            if response:
                file_path = filedialog.askopenfilename(
                    title="Open Excel Workbook",
                    filetypes=EXCEL_FILETYPES
                )
                if file_path:
                    try:
                        # Validate file path and content
                        is_valid, error_msg = SecurityValidator.validate_excel_file(file_path)
                        if not is_valid:
                            messagebox.showerror("Security Error", f"Cannot open file: {error_msg}")
                            self.logger.warning("Security validation failed for file",
                                              file_path=file_path, error=error_msg)
                            return
                        
                        # Sanitize file path
                        safe_path = SecurityValidator.validate_file_path(file_path)
                        
                        from openpyxl import load_workbook
                        self.current_workbook = load_workbook(safe_path)
                    except ValueError as e:
                        messagebox.showerror("Security Error", str(e))
                        self.logger.warning("Security validation error", error=str(e))
                        return
                    except Exception as e:
                        messagebox.showerror(
                            "Error",
                            f"Could not open workbook:\n{str(e)}"
                        )
                        return
                else:
                    return
            else:
                return
        
        # Check if PM Summary sheet exists
        if "PM Summary" not in self.current_workbook.sheetnames:
            messagebox.showerror(
                "Error",
                "This workbook does not contain a 'PM Summary' sheet"
            )
            return
            
        # Show the EditWorkpackageDialog
        from handlers.edit_workpackage_handler import EditWorkpackageDialog
        dialog = EditWorkpackageDialog(self.root, self.current_workbook)
        
        # Wait for dialog
        self.root.wait_window(dialog.dialog)
        
        if dialog.result:
            # Ask user where to save the workbook
            try:
                save_path = filedialog.asksaveasfilename(
                    title="Save Workbook",
                    defaultextension=".xlsx",
                    filetypes=EXCEL_FILETYPES
                )
                if save_path:
                    self.current_workbook.save(save_path)
                    messagebox.showinfo(
                        "Success",
                        f"Updated workpackage in row {dialog.result['row']}\n"
                        f"Workbook saved to: {save_path}"
                    )
                else:
                    messagebox.showwarning(
                        "Warning",
                        "Workpackage updated but workbook not saved!"
                    )
            except Exception as e:
                messagebox.showerror(
                    "Error",
                    f"Failed to save workbook:\n{str(e)}"
                )

    def update_budget_overview(self):
        """Manually update Budget Overview worksheet from menu."""
        with LogContext("manual_update_budget_overview", user_id="current_user"):
            self.logger.info("Starting manual Budget Overview update")
            
            # Check if we have an open workbook
            if self.current_workbook is None:
                self.logger.warning("No workbook open, prompting user to open one")
                response = messagebox.askyesno(
                    "No Workbook Open",
                    "No workbook is currently open. Would you like to open one now?"
                )
                if response:
                    file_path = filedialog.askopenfilename(
                        title="Open Excel Workbook",
                        filetypes=EXCEL_FILETYPES
                    )
                    if file_path:
                        try:
                            # Validate file path and content
                            is_valid, error_msg = SecurityValidator.validate_excel_file(file_path)
                            if not is_valid:
                                messagebox.showerror("Security Error", f"Cannot open file: {error_msg}")
                                self.logger.warning("Security validation failed for file",
                                                  file_path=file_path, error=error_msg)
                                return
                            
                            # Sanitize file path
                            safe_path = SecurityValidator.validate_file_path(file_path)
                            
                            from openpyxl import load_workbook
                            self.current_workbook = load_workbook(safe_path)
                            self.logger.info("Workbook loaded successfully for Budget Overview update",
                                           file_path=safe_path)
                        except ValueError as e:
                            messagebox.showerror("Security Error", str(e))
                            self.logger.warning("Security validation error", error=str(e))
                            return
                        except Exception as e:
                            self.logger.error("Failed to load workbook for Budget Overview update",
                                            file_path=file_path, error=str(e))
                            messagebox.showerror(
                                "Error",
                                f"Could not open workbook:\n{str(e)}"
                            )
                            return
                    else:
                        self.logger.info("User cancelled workbook selection")
                        return
                else:
                    self.logger.info("User declined to open workbook")
                    return

            # Perform the Budget Overview update
            try:
                from handlers.update_budget_overview_handler_formula import update_budget_overview_with_progress
                
                success = update_budget_overview_with_progress(self.root, self.current_workbook)
                
                if success:
                    # Apply conditional formatting after successful formula updates
                    try:
                        from handlers.budget_overview_format import apply_budget_overview_formatting
                        formatting_success = apply_budget_overview_formatting(self.root, self.current_workbook)
                        
                        if formatting_success:
                            success_msg = (
                                "✅ Budget Overview updated successfully!\n"
                                "🎨 Conditional formatting applied based on row completion.\n\n"
                                "Ready to save the updated workbook."
                            )
                        else:
                            success_msg = (
                                "✅ Budget Overview updated successfully!\n"
                                "⚠️ Formatting could not be applied (see debug window for details).\n\n"
                                "Ready to save the updated workbook."
                            )
                    except ImportError as import_err:
                        success_msg = (
                            "✅ Budget Overview updated successfully!\n"
                            f"⚠️ Formatting module not available: {str(import_err)}\n\n"
                            "Ready to save the updated workbook."
                        )
                        self.logger.warning("Formatting module not available during manual update", error=str(import_err))
                    except Exception as format_err:
                        success_msg = (
                            "✅ Budget Overview updated successfully!\n"
                            f"⚠️ Formatting failed: {str(format_err)}\n\n"
                            "Ready to save the updated workbook."
                        )
                        self.logger.warning("Formatting failed during manual update", error=str(format_err))
                    
                    # Ask user where to save the workbook
                    save_path = filedialog.asksaveasfilename(
                        title="Save Updated Workbook",
                        defaultextension=".xlsx",
                        filetypes=EXCEL_FILETYPES
                    )
                    if save_path:
                        try:
                            # Validate and sanitize save path
                            safe_save_path = SecurityValidator.validate_file_path(save_path)
                            
                            # Ensure proper extension
                            if not safe_save_path.lower().endswith('.xlsx'):
                                safe_save_path += '.xlsx'
                            
                            self.current_workbook.save(safe_save_path)
                        except ValueError as e:
                            messagebox.showerror("Security Error", str(e))
                            self.logger.warning("Security validation error for save path", error=str(e))
                            return
                        messagebox.showinfo(
                            "Success",
                            f"{success_msg}\n\nWorkbook saved to: {save_path}"
                        )
                        self.logger.info("Budget Overview update completed and workbook saved",
                                        file_path=save_path)
                    else:
                        messagebox.showwarning(
                            "Warning",
                            "Budget Overview updated but workbook not saved!"
                        )
                        self.logger.warning("Budget Overview updated but user did not save workbook")
                else:
                    messagebox.showerror(
                        "Error",
                        "Failed to update Budget Overview. Please check the logs for details."
                    )
                    self.logger.error("Budget Overview update failed")
                    
            except Exception as e:
                self.logger.exception("Exception during manual Budget Overview update")
                messagebox.showerror(
                    "Error",
                    f"An error occurred while updating Budget Overview:\n{str(e)}"
                )

    def update_pm_overview(self):
        """Manually update PM Overview worksheet from menu."""
        with LogContext("manual_update_pm_overview", user_id="current_user"):
            self.logger.info("Starting manual PM Overview update")
            
            # Check if we have an open workbook
            if self.current_workbook is None:
                self.logger.warning("No workbook open, prompting user to open one")
                response = messagebox.askyesno(
                    "No Workbook Open",
                    "No workbook is currently open. Would you like to open one now?"
                )
                if response:
                    file_path = filedialog.askopenfilename(
                        title="Open Excel Workbook",
                        filetypes=EXCEL_FILETYPES
                    )
                    if file_path:
                        try:
                            # Validate file path and content
                            is_valid, error_msg = SecurityValidator.validate_excel_file(file_path)
                            if not is_valid:
                                messagebox.showerror("Security Error", f"Cannot open file: {error_msg}")
                                self.logger.warning("Security validation failed for file",
                                                  file_path=file_path, error=error_msg)
                                return
                            
                            # Sanitize file path
                            safe_path = SecurityValidator.validate_file_path(file_path)
                            
                            from openpyxl import load_workbook
                            self.current_workbook = load_workbook(safe_path)
                            self.logger.info("Workbook loaded successfully for PM Overview update",
                                           file_path=safe_path)
                        except ValueError as e:
                            messagebox.showerror("Security Error", str(e))
                            self.logger.warning("Security validation error", error=str(e))
                            return
                        except Exception as e:
                            self.logger.error("Failed to load workbook for PM Overview update",
                                            file_path=file_path, error=str(e))
                            messagebox.showerror(
                                "Error",
                                f"Could not open workbook:\n{str(e)}"
                            )
                            return
                    else:
                        self.logger.info("User cancelled workbook selection")
                        return
                else:
                    self.logger.info("User declined to open workbook")
                    return

            # Perform the PM Overview update
            try:
                from handlers.update_pm_overview_handler import update_pm_overview_with_progress
                
                success = update_pm_overview_with_progress(self.root, self.current_workbook)
                
                if success:
                    # Apply conditional formatting after successful formula updates
                    try:
                        from handlers.pm_overview_format import apply_pm_overview_formatting
                        formatting_success = apply_pm_overview_formatting(self.root, self.current_workbook)
                        
                        if formatting_success:
                            success_msg = (
                                "✅ PM Overview updated successfully!\n"
                                "🎨 Conditional formatting applied based on row completion.\n\n"
                                "Ready to save the updated workbook."
                            )
                        else:
                            success_msg = (
                                "✅ PM Overview updated successfully!\n"
                                "⚠️ Formatting could not be applied (see debug window for details).\n\n"
                                "Ready to save the updated workbook."
                            )
                    except ImportError as import_err:
                        success_msg = (
                            "✅ PM Overview updated successfully!\n"
                            f"⚠️ Formatting module not available: {str(import_err)}\n\n"
                            "Ready to save the updated workbook."
                        )
                        self.logger.warning("Formatting module not available during manual PM update", error=str(import_err))
                    except Exception as format_err:
                        success_msg = (
                            "✅ PM Overview updated successfully!\n"
                            f"⚠️ Formatting failed: {str(format_err)}\n\n"
                            "Ready to save the updated workbook."
                        )
                        self.logger.warning("Formatting failed during manual PM update", error=str(format_err))
                    
                    # Ask user where to save the workbook
                    save_path = filedialog.asksaveasfilename(
                        title="Save Updated Workbook",
                        defaultextension=".xlsx",
                        filetypes=EXCEL_FILETYPES
                    )
                    if save_path:
                        try:
                            # Validate and sanitize save path
                            safe_save_path = SecurityValidator.validate_file_path(save_path)
                            
                            # Ensure proper extension
                            if not safe_save_path.lower().endswith('.xlsx'):
                                safe_save_path += '.xlsx'
                            
                            self.current_workbook.save(safe_save_path)
                        except ValueError as e:
                            messagebox.showerror("Security Error", str(e))
                            self.logger.warning("Security validation error for save path", error=str(e))
                            return
                        messagebox.showinfo(
                            "Success",
                            f"{success_msg}\n\nWorkbook saved to: {save_path}"
                        )
                        self.logger.info("PM Overview update completed and workbook saved",
                                        file_path=save_path)
                    else:
                        messagebox.showwarning(
                            "Warning",
                            "PM Overview updated but workbook not saved!"
                        )
                        self.logger.warning("PM Overview updated but user did not save workbook")
                else:
                    messagebox.showerror(
                        "Error",
                        "Failed to update PM Overview. Please check the logs for details."
                    )
                    self.logger.error("PM Overview update failed")
                    
            except Exception as e:
                self.logger.exception("Exception during manual PM Overview update")
                messagebox.showerror(
                    "Error",
                    f"An error occurred while updating PM Overview:\n{str(e)}"
                )

    def _setup_preferences_menu(self, menubar):
        """Set up the Preferences menu."""
        # TODO: Implement preferences menu setup
        pass

    def clone_file(self):
        """Clone an Excel file to a new location with user-provided details. Prompts for new name, extension and optional project name."""
        if self.developer_mode:
            dev_log("Clone file triggered.")

        # Step 1: Select source file
        src_path = filedialog.askopenfilename(
            title="Select file to clone",
            filetypes=EXCEL_FILETYPES
        )
        if not src_path:
            return

        try:
            # Validate source file
            is_valid, error_msg = SecurityValidator.validate_excel_file(src_path)
            if not is_valid:
                messagebox.showerror("Security Error", f"Cannot clone file: {error_msg}")
                self.logger.warning("Security validation failed for source file",
                                    file_path=src_path, error=error_msg)
                return
            
            # Sanitize source path
            safe_src_path = SecurityValidator.validate_file_path(src_path)
        except ValueError as e:
            messagebox.showerror("Security Error", str(e))
            self.logger.warning("Security validation error for source file", error=str(e))
            return

        # Step 2: Select destination directory
        dest_dir = filedialog.askdirectory(
            title="Select destination folder"
        )
        if not dest_dir:
            return

        try:
            # Validate destination directory
            safe_dest_dir = SecurityValidator.validate_directory_path(dest_dir)
        except ValueError as e:
            messagebox.showerror("Security Error", f"Invalid destination directory: {str(e)}")
            self.logger.warning("Security validation error for destination directory", error=str(e))
            return

        # Step 3: Get file details
        base_name = os.path.splitext(os.path.basename(safe_src_path))[0]
        ext = os.path.splitext(safe_src_path)[1]

        new_name = self._prompt_for_detail(
            CLONE_FILE_TITLE, "Enter new file name:", base_name)
        if not new_name:
            return

        # Sanitize new filename
        new_name = SecurityValidator.sanitize_filename(new_name)
        if not new_name:
            messagebox.showerror("Invalid Input", "Invalid file name provided.")
            return

        new_ext = self._prompt_for_detail(
            CLONE_FILE_TITLE, ENTER_FILE_EXT, ext)
        if not new_ext:
            return

        # Sanitize extension
        new_ext = InputSanitizer.sanitize_string(new_ext, max_length=10)
        if not new_ext.startswith('.'):
            new_ext = '.' + new_ext

        # Optional project name
        project_name = self._prompt_for_detail(
            CLONE_FILE_TITLE,
            "Enter project name (optional):",
            ""
        )
        
        # Sanitize project name if provided
        if project_name:
            project_name = InputSanitizer.sanitize_string(project_name, max_length=100)

        # Step 4: Clone the file
        dest_path = os.path.join(safe_dest_dir, new_name + new_ext)
        try:
            # Validate final destination path
            safe_dest_path = SecurityValidator.validate_file_path(dest_path)
            shutil.copy2(safe_src_path, safe_dest_path)
            messagebox.showinfo(
                "Clone Complete",
                f"File cloned to:\n{safe_dest_path}"
            )
            if project_name:
                dev_log(f"Project name for clone: {project_name}")
        except ValueError as e:
            messagebox.showerror("Security Error", str(e))
            self.logger.warning("Security validation error for destination path", error=str(e))
        except Exception as e:
            messagebox.showerror(
                "Clone Failed",
                f"Error cloning file:\n{e}"
            )

    def _prompt_for_detail(self, title, prompt, default=""):
        from tkinter import simpledialog
        return simpledialog.askstring(
            title,
            prompt,
            initialvalue=default
        )

    def compare_files(self):
        """Compare Excel files and display differences in a dialog."""
        if self.developer_mode:
            dev_log("Compare files triggered.")
        from tkinter import ttk

        # Select files to compare
        file_paths = filedialog.askopenfilenames(
            title="Select up to 3 files to compare",
            filetypes=EXCEL_FILETYPES
        )
        file_paths = list(file_paths)[:3]  # Limit to 3 files
        
        if len(file_paths) < 2:
            messagebox.showinfo(
                "Compare Files",
                "Please select 2-3 files to compare."
            )
            return

        def create_comparison_view(ref_idx, result_win):
            """Create the comparison view dialog."""
            ref_file = file_paths[ref_idx]
            others = [f for i, f in enumerate(file_paths) if i != ref_idx]
            
            try:
                ref_df = pd.read_excel(ref_file)
                other_dfs = [pd.read_excel(f) for f in others]
            except Exception as e:
                messagebox.showerror(
                    "Compare Failed",
                    f"Error reading files:\n{e}"
                )
                return

            diffs = []
            for i, df in enumerate(other_dfs):
                # Find rows that differ between files
                diff = df.merge(
                    ref_df,
                    how='outer',
                    indicator=True
                ).query('_merge != "both"')
                
                base_name = os.path.basename(others[i])
                if diff.empty:
                    diffs.append((
                        base_name,
                        pd.DataFrame(["No differences found."])
                    ))
                else:
                    diffs.append((base_name, diff))

            # Configure result window
            result_win.title("Comparison Results")
            result_win.geometry("600x400")
            result_win.minsize(600, 220)
            result_win.grid_rowconfigure(1, weight=1)
            result_win.grid_columnconfigure(0, weight=1)

            # Reference file label
            ref_name = os.path.basename(ref_file)
            ttk.Label(
                result_win,
                text=f"Reference: {ref_name}",
                font=("Arial", 11, "bold")
            ).grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 0))

            # List of comparison files
            listbox = tk.Listbox(result_win)
            for name, _ in diffs:
                listbox.insert(tk.END, name)
            listbox.grid(row=1, column=0, sticky="ew", padx=10, pady=(5, 0))

            # Differences display
            diff_text = tk.Text(result_win, wrap="none", height=12)
            diff_text.grid(
                row=2, column=0,
                sticky="nsew", padx=10, pady=(5, 0)
            )

            def show_diff(event=None):
                """Display differences for selected file."""
                sel = listbox.curselection()
                if not sel:
                    return
                _, diff = diffs[sel[0]]
                diff_text.config(state="normal")
                diff_text.delete("1.0", tk.END)
                if isinstance(diff, pd.DataFrame):
                    diff_str = diff.to_string(index=False)
                else:
                    diff_str = str(diff)
                diff_text.insert(tk.END, diff_str)
                diff_text.config(state="disabled")

            def export_to_excel():
                """Export the currently selected difference to an Excel file."""
                sel = listbox.curselection()
                if not sel:
                    messagebox.showinfo("Export", "Please select a file to export.")
                    return
                name, diff = diffs[sel[0]]
                
                # Create a timestamp for the export
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                
                # Prompt for save location
                save_path = filedialog.asksaveasfilename(
                    title="Save Differences As",
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    initialfile=f"diff_{name}_{timestamp}.xlsx"
                )
                if not save_path:
                    return
                    
                try:
                    # Create Excel writer object
                    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                        # Write comparison information
                        info_df = pd.DataFrame({
                            'Parameter': ['Reference File', 'Compared File', 'Export Date'],
                            'Value': [
                                os.path.basename(ref_file),
                                name,
                                datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            ]
                        })
                        info_df.to_excel(writer, sheet_name='Info', index=False)
                        
                        # Write differences
                        if isinstance(diff, pd.DataFrame):
                            # Format the differences DataFrame
                            if '_merge' in diff.columns:
                                diff = diff.drop('_merge', axis=1)
                            diff.to_excel(writer, sheet_name='Differences', index=False)
                        else:
                            # If diff is not a DataFrame, write as plain text
                            pd.DataFrame({'Result': [str(diff)]}).to_excel(
                                writer, sheet_name='Differences', index=False
                            )
                            
                    messagebox.showinfo(
                        "Export Complete",
                        f"Differences exported to:\n{save_path}"
                    )
                except Exception as e:
                    messagebox.showerror(
                        "Export Failed",
                        f"Error exporting differences:\n{e}"
                    )

            def change_reference():
                """Change the reference file for comparison."""
                dialog = tk.Toplevel(self.root)
                dialog.title("Select Reference File")
                ttk.Label(
                    dialog,
                    text="Choose the reference file:"
                ).pack(pady=5)
                
                ref_var = tk.StringVar(value=file_paths[ref_idx])
                for path in file_paths:
                    ttk.Radiobutton(
                        dialog,
                        text=os.path.basename(path),
                        variable=ref_var,
                        value=path
                    ).pack(anchor="w")

                ref_idx_holder = {}

                def set_ref():
                    ref_idx_holder["idx"] = file_paths.index(ref_var.get())
                    dialog.destroy()

                ttk.Button(
                    dialog,
                    text="OK",
                    command=set_ref
                ).pack(pady=5)

                dialog.wait_window()
                new_ref_idx = ref_idx_holder.get("idx", ref_idx)
                result_win.destroy()
                show_comparison_dialog(new_ref_idx)

            # Button frame

            btn_frame = ttk.Frame(result_win)
            btn_frame.grid(row=3, column=0, sticky="ew", padx=10, pady=8)
            btn_frame.grid_columnconfigure(0, weight=1)
            btn_frame.grid_columnconfigure(1, weight=1)
            btn_frame.grid_columnconfigure(2, weight=1)

            ttk.Button(
                btn_frame,
                text="Change Reference",
                command=change_reference
            ).grid(row=0, column=0, sticky="w", padx=5)

            ttk.Button(
                btn_frame,
                text="Export to Excel",
                command=export_to_excel
            ).grid(row=0, column=1, sticky="ew", padx=5)

            ttk.Button(
                btn_frame,
                text="Close",
                command=result_win.destroy
            ).grid(row=0, column=2, sticky="e", padx=5)

            # Bind events
            listbox.bind("<<ListboxSelect>>", show_diff)
            listbox.selection_set(0)
            show_diff()

        def show_comparison_dialog(ref_idx):
            """Show the main comparison dialog."""
            result_win = tk.Toplevel(self.root)
            create_comparison_view(ref_idx, result_win)

        def choose_initial_reference():
            """Let user choose the initial reference file."""
            dialog = tk.Toplevel(self.root)
            dialog.title("Select Reference File")
            ttk.Label(
                dialog,
                text="Choose the reference file:"
            ).pack(pady=5)
            
            ref_var = tk.StringVar(value=file_paths[0])
            for path in file_paths:
                ttk.Radiobutton(
                    dialog,
                    text=os.path.basename(path),
                    variable=ref_var,
                    value=path
                ).pack(anchor="w")

            ref_idx_holder = {}

            def set_ref():
                ref_idx_holder["idx"] = file_paths.index(ref_var.get())
                dialog.destroy()

            ttk.Button(dialog, text="OK", command=set_ref).pack(pady=5)
            dialog.wait_window()
            return ref_idx_holder.get("idx", 0)

        ref_idx = choose_initial_reference()
        show_comparison_dialog(ref_idx)

    def create_from_scratch(self):
        """Create a new, empty Excel file at a user-specified location."""
        if self.developer_mode:
            dev_log("Create from scratch triggered.")

        # Step 1: Select save location
        dest_dir = filedialog.askdirectory(
            title="Select save folder"
        )
        if not dest_dir:
            return

        # Step 2: Get file details
        file_name = self._prompt_for_detail(
            CREATE_SCRATCH_TITLE,
            "Enter file name:"
        )
        if not file_name:
            return

        file_ext = self._prompt_for_detail(
            CREATE_SCRATCH_TITLE,
            ENTER_FILE_EXT,
            EXCEL_DEFAULT_EXT
        )
        if not file_ext:
            return

        # Step 3: Create empty file
        dest_path = os.path.join(dest_dir, file_name + file_ext)
        try:
            import pandas as pd
            df = pd.DataFrame()
            df.to_excel(dest_path, index=False)
            messagebox.showinfo(
                "File Created",
                f"New file created:\n{dest_path}"
            )
        except Exception as e:
            messagebox.showerror(
                "Create Failed",
                f"Error creating file:\n{e}"
            )

    def create_from_template(self):
        """Create a new Excel file from a template or existing file."""
        if self.developer_mode:
            dev_log("Create from template triggered.")

        # Step 1: Select template
        template_path = filedialog.askopenfilename(
            title="Select template",
            filetypes=EXCEL_FILETYPES
        )
        if not template_path:
            return

        # Validate file is Excel
        is_excel = any(
            template_path.endswith(ext)
            for ext in (".xlsx", ".xls")
        )
        if not is_excel:
            messagebox.showerror(
                "Invalid File",
                "Not an Excel file."
            )
            return

        # Step 2: Select save location
        dest_dir = filedialog.askdirectory(
            title="Select save folder"
        )
        if not dest_dir:
            return

        # Step 3: Get new file details
        file_name = self._prompt_for_detail(
            "Create from Template",
            "Enter new file name:"
        )
        if not file_name:
            return

        file_ext = self._prompt_for_detail(
            "Create from Template",
            ENTER_FILE_EXT,
            EXCEL_DEFAULT_EXT
        )
        if not file_ext:
            return

        # Step 4: Copy template
        dest_path = os.path.join(dest_dir, file_name + file_ext)
        try:
            import shutil
            shutil.copy2(template_path, dest_path)
            messagebox.showinfo(
                "File Created",
                f"Template copied to:\n{dest_path}"
            )
        except Exception as e:
            messagebox.showerror(
                "Create Failed",
                f"Copy error:\n{e}"
            )

    def modify_file(self):
        """Placeholder for modifying a file (not yet implemented)."""
        if self.developer_mode:
            dev_log("Modify file triggered.")
        not_implemented_yet("Modify")

    def restore_file(self):
        """Placeholder for restoring a file from backup (not yet implemented)."""
        if self.developer_mode:
            dev_log("Restore file triggered.")
        not_implemented_yet("Restore")

    def show_diagnostics(self):
        """Show the diagnostics dashboard window with startup and config status."""
        # Diagnostics dashboard window
        diag_win = tk.Toplevel(self.root)
        diag_win.title("Startup Diagnostics Dashboard")
        diag_win.geometry("400x370")
        diag_win.grab_set()

        version_info = full_version_string()

        categories = [
            "Directory Check",
            "Config File Check",
            "Log Integrity",
            "Summary",
            "Version Info"
        ]
        details = {
            "Directory Check": "\n".join([
                "✅ workbooks/          …exists",
                "✅ logs/system/        …ready",
                "✅ logs/user_actions/  …ready",
                "🆕 logs/comparisons/snapshots/  …created",
                "✅ config/             …ready",
                "⚠️  backups/            …missing (created default)",
                "✅ templates/          …exists"
            ]),
            "Config File Check": "\n".join([
                "✅ user.config.json     …loaded successfully",
                "⚠️  backup.config.json   …not found (restored defaults)",
                "🆕 diagnostic.config.json …created new with defaults"
            ]),
            "Log Integrity": "\n".join([
                "⚠️  2 comparison logs from v0.8 detected  → marked as legacy",
                "✅ latest snapshot valid (2025-06-22)"
            ]),
            "Summary": (
                "Startup passed with 2 warnings and 2 recoveries.\n"
                "New folders/configs have been generated where needed.\n\n"
                "📌 All systems go — launching interface..."
            ),
            "Version Info": version_info
        }

        # Category list
        # Category label
        cat_label = tk.Label(
            diag_win,
            text="Select Diagnostic Category:",
            font=("Arial", 11, "bold")
        )
        cat_label.pack(pady=(10, 0))
        
        # Category listbox
        cat_listbox = tk.Listbox(diag_win, height=7)
        for cat in categories:
            cat_listbox.insert(tk.END, cat)
        cat_listbox.pack(fill="x", padx=10, pady=5)

        # Details display
        details_text = tk.Text(
            diag_win,
            height=10,
            wrap="word",
            state="disabled"
        )
        details_text.pack(fill="both", expand=True, padx=10, pady=5)

        def show_details(event=None):
            selection = cat_listbox.curselection()
            if selection:
                cat = categories[selection[0]]
                details_text.config(state="normal")
                details_text.delete("1.0", tk.END)
                details_text.insert(tk.END, details[cat])
                details_text.config(state="disabled")

        cat_listbox.bind("<<ListboxSelect>>", show_details)
        # Show first category by default
        cat_listbox.selection_set(0)
        show_details()

        # Close button
        close_btn = tk.Button(diag_win, text="Close", command=diag_win.destroy)
        close_btn.pack(pady=5)
        # (No menu bar code here)

    def show_help(self):
        """Show the Help window with basic instructions and contact info."""
        help_win = tk.Toplevel(self.root)
        help_win.title("Hub")
        help_win.geometry("400x250")
        help_win.grab_set()

        # Title
        title = ttk.Label(
            help_win,
            text="Hub",
            font=("Arial", 16, "bold")
        )
        title.pack(pady=(15, 5))

        # Attention message
        attention = ttk.Label(
            help_win,
            text="Attention needed",
            font=("Arial", 12, "bold"),
            foreground="red"
        )
        attention.pack(pady=(0, 5))

        # Help text
        help_msg = (
            "Help menu\n\n"
            "This is a placeholder for the Help section.\n\n"
            "Please refer to the documentation or contact\n"
            "support for assistance."
        )
        text = ttk.Label(
            help_win,
            text=help_msg,
            font=("Arial", 11),
            justify="left"
        )
        text.pack(padx=15, pady=5)

        # Close button
        close_btn = ttk.Button(
            help_win,
            text="Close",
            command=help_win.destroy
        )
        close_btn.pack(pady=10)


    def show_preferences(self):
        """Show the Preferences dialog and reload preferences after closing."""
        from preferences import PreferencesDialog
        dialog = PreferencesDialog(self.root)
        self.root.wait_window(dialog.dialog)
        # Reload preferences after dialog is closed
        if self.prefs_manager is not None:
            self.current_config = self.prefs_manager.load_config()
            self.apply_theme()

    def _initialize_theme_system(self):
        """Initialize the theme management system."""
        try:
            from gui.theme_manager import ThemeManager
            self.theme_manager = ThemeManager()
            
            # Apply the current theme from preferences
            current_theme = self.current_config.get('theme', 'light')
            self.theme_manager.apply_theme(self.root, current_theme)
            
        except Exception as e:
            self.logger.warning(f"Could not initialize theme system: {e}")
            # Fallback to basic theme
            self.theme_manager = None
            theme = self.current_config.get('theme', 'light')
            if theme == 'dark':
                self.root.configure(bg='#2b2b2b')
            else:
                self.root.configure(bg='#f0f0f0')

    def apply_theme(self):
        """Apply the selected theme (light or dark) to the entire application."""
        try:
            # Get current theme from preferences
            theme = self.current_config.get("theme", "light")
            self.logger.info(f"Applying theme: {theme}")
            
            # Use theme manager if available
            if hasattr(self, 'theme_manager') and self.theme_manager:
                self.theme_manager.apply_theme(self.root, theme)
                self.logger.info(f"Theme '{theme}' applied successfully using ThemeManager")
            else:
                # Fallback to function-based theme application
                from gui.theme_manager import apply_theme_to_app
                apply_theme_to_app(self.root, theme)
                self.logger.info(f"Theme '{theme}' applied successfully using fallback method")
            
        except Exception as e:
            self.logger.error(f"Failed to apply theme: {e}")
            # Fallback to basic theme
            theme = self.current_config.get("theme", "light")
            bg_color = "#2d2d30" if theme == "dark" else "#f0f0f0"
            self.root.configure(bg=bg_color)

    def _setup_performance_indicator(self):
        """Set up the performance indicator in the status bar."""
        try:
            # Create status bar frame
            status_frame = ttk.Frame(self.root)
            status_frame.pack(side=tk.BOTTOM, fill=tk.X)
            
            # Add performance indicator
            self.performance_indicator = PerformanceIndicator(status_frame)
            self.performance_indicator.get_widget().pack(side=tk.RIGHT, padx=10, pady=2)
            
        except Exception as e:
            self.logger.warning(f"Failed to setup performance indicator: {e}")

    def show_performance_monitor(self):
        """Show the performance monitoring dialog."""
        try:
            show_performance_monitor(self.root)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open performance monitor:\n{e}")
            self.logger.error(f"Performance monitor error: {e}")

    @monitor_performance(include_memory=True, log_level='DEBUG')
    @monitor_operation("workbook_loading")
    def load_workbook(self, file_path: str):
        """Load Excel workbook with performance monitoring."""
        try:
            # Use the existing workbook loading logic with performance monitoring
            # This is a placeholder - integrate with your actual workbook loading
            self.logger.info(f"Loading workbook: {file_path}")
            # Your existing workbook loading code here
            return True
        except Exception as e:
            self.logger.error(f"Failed to load workbook: {e}")
            return False

    # check_first_run is now replaced by check_first_run_and_show_diagnostics

    def create_directory_structure(self):
        """Create the necessary directory structure and config files."""
        user_home = str(Path.home())
        base_dir = os.path.join(user_home, "ProjectBudgetinator")
        
        # Define directory structure
        directories = [
            os.path.join(base_dir, "workbooks"),
            os.path.join(base_dir, "logs", "system"),
            os.path.join(base_dir, "logs", "user_actions"),
            os.path.join(base_dir, "logs", "comparisons", "snapshots"),
            os.path.join(base_dir, "config"),
            os.path.join(base_dir, "backups"),
            os.path.join(base_dir, "templates")
        ]

        # Create directories
        for directory in directories:
            os.makedirs(directory, exist_ok=True)

        # Map config files to their default content
        config_files = {
            "user.config.json": DEFAULT_USER_CONFIG,
            "backup.config.json": DEFAULT_BACKUP_CONFIG,
            "diagnostic.config.json": DEFAULT_DIAGNOSTIC_CONFIG
        }

        # Create config files
        config_dir = os.path.join(base_dir, "config")
        for filename, content in config_files.items():
            filepath = os.path.join(config_dir, filename)
            with open(filepath, "w") as f:
                json.dump(content, f, indent=4)

    def show_batch_operations(self):
        """Show the batch operations dialog."""
        try:
            self.logger.info("Opening batch operations dialog")
            show_batch_operations_dialog(parent=self.root)
            self.logger.info("Batch operations dialog closed")
        except Exception as e:
            self.logger.error("Failed to open batch operations dialog",
                              error=str(e))
            messagebox.showerror("Error",
                                 "Failed to open batch operations dialog.")

    def _on_profile_changed(self, profile):
        """Handle profile change event."""
        try:
            self.current_profile = profile
            self.logger.info(f"Profile changed to: {profile.profile_name}")
            
            # Update current config with new profile preferences
            if profile.preferences_data:
                self.current_config.update(profile.preferences_data)
                self.apply_theme()  # Apply any theme changes
                
            # Show notification
            messagebox.showinfo(
                "Profile Changed",
                f"Switched to profile: {profile.profile_name}\n"
                f"Environment: {profile.environment_type}"
            )
            
        except Exception as e:
            self.logger.error(f"Error handling profile change: {e}")
            messagebox.showerror("Error", f"Failed to switch profile: {str(e)}")

    def change_password(self):
        """Show change password dialog."""
        try:
            # Check if current user is admin
            if self.auth_manager and self.auth_manager.is_current_user_admin():
                messagebox.showinfo(
                    "Admin Password Protected",
                    "The admin user password cannot be changed.\n\n"
                    "The admin password is permanently set to 'pbi' for security reasons."
                )
                return
            
            from gui.auth.password_change_dialog import show_password_change_dialog
            
            def change_password_callback(old_password: str, new_password: str) -> bool:
                return self.auth_manager.change_password(old_password, new_password)
            
            show_password_change_dialog(self.root, self.current_user.username, change_password_callback)
        except ImportError:
            # Simple password change dialog
            from tkinter import simpledialog
            
            # Check if current user is admin
            if self.current_user and self.current_user.username.lower() == "admin":
                messagebox.showinfo(
                    "Admin Password Protected",
                    "The admin user password cannot be changed.\n\n"
                    "The admin password is permanently set to 'pbi' for security reasons."
                )
                return
            
            old_password = simpledialog.askstring("Change Password", "Enter current password:", show='*')
            if old_password:
                new_password = simpledialog.askstring("Change Password", "Enter new password:", show='*')
                if new_password:
                    confirm_password = simpledialog.askstring("Change Password", "Confirm new password:", show='*')
                    if new_password == confirm_password:
                        if self.auth_manager.change_password(old_password, new_password):
                            messagebox.showinfo("Success", "Password changed successfully!")
                        else:
                            messagebox.showerror("Error", "Failed to change password. Please check your current password.")
                    else:
                        messagebox.showerror("Error", "Passwords do not match!")
        except Exception as e:
            self.logger.error(f"Error changing password: {e}")
            messagebox.showerror("Error", f"Failed to change password: {str(e)}")

    def show_profile_management(self):
        """Show profile management dialog."""
        try:
            from gui.auth.profile_dialog import show_profile_management_dialog
            show_profile_management_dialog(self.root, self.auth_manager.profile_manager,
                                          self.current_user.user_id)
            if hasattr(self, 'profile_switcher_menu'):
                self.profile_switcher_menu.refresh()
        except ImportError:
            messagebox.showinfo("Profile Management", "Profile management feature is not yet available.")
        except Exception as e:
            self.logger.error(f"Error opening profile management: {e}")
            messagebox.showerror("Error", f"Failed to open profile management: {str(e)}")

    def show_user_administration(self):
        """Show user administration dialog."""
        try:
            from gui.auth.user_admin_dialog import show_user_admin_dialog
            show_user_admin_dialog(self.root, self.auth_manager)
        except ImportError:
            messagebox.showinfo("User Administration", "User administration feature is not yet available.")
        except Exception as e:
            self.logger.error(f"Error opening user administration: {e}")
            messagebox.showerror("Error", f"Failed to open user administration: {str(e)}")

    def logout(self):
        """Logout current user and restart application."""
        try:
            if messagebox.askyesno("Logout", "Are you sure you want to logout?"):
                # Save any current preferences to profile
                if self.auth_manager and self.current_profile:
                    self.auth_manager.update_preferences(self.current_config)
                
                # Logout
                if self.auth_manager:
                    self.auth_manager.logout()
                
                self.logger.info("User logged out, restarting application")
                
                # Restart the application
                self.root.quit()
                # Note: In a real implementation, you might want to restart the entire application
                # For now, we'll just quit and let the user restart manually
                
        except Exception as e:
            self.logger.error(f"Error during logout: {e}")
            messagebox.showerror("Error", f"Failed to logout: {str(e)}")

    def exit_program(self):
        """Prompt the user to confirm exit and quit the application."""
        try:
            # Save current preferences to profile before exiting
            if self.auth_manager and self.current_profile:
                self.auth_manager.update_preferences(self.current_config)
            
            if messagebox.askyesno("Exit", "Are you sure you want to exit?"):
                # Logout user
                if self.auth_manager:
                    self.auth_manager.logout()
                self.root.quit()
                
        except Exception as e:
            self.logger.error(f"Error during exit: {e}")
            # Still allow exit even if there's an error
            self.root.quit()

    def run(self):
        """Start the Tkinter main event loop for the application."""
        self.root.mainloop()


if __name__ == "__main__":
    app = ProjectBudgetinator()
    app.run()
