"""
Batch Operations System for ProjectBudgetinator

This module provides comprehensive batch processing capabilities including:
- Multiple file selection and processing
- Drag-and-drop support
- Operation queues with prioritization
- Batch progress tracking
- Error handling and recovery
- Result aggregation and reporting
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import tkinter.dnd as dnd
import os
import threading
import queue
import time
from typing import List, Dict, Callable, Optional, Any
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path

from logger import get_structured_logger, LogContext
from gui.progress_dialog import ProgressDialog, show_progress_for_operation
from utils.error_handler import ExceptionHandler


logger = get_structured_logger("gui.batch_operations")
exception_handler = ExceptionHandler()


class OperationType(Enum):
    """Types of batch operations."""
    VALIDATE_FILES = "validate_files"
    ADD_PARTNERS = "add_partners"
    ADD_WORKPACKAGES = "add_workpackages" 
    CONVERT_FORMAT = "convert_format"
    BACKUP_FILES = "backup_files"
    CUSTOM_OPERATION = "custom_operation"


class OperationStatus(Enum):
    """Status of batch operations."""
    PENDING = "pending"
    IN_PROGRESS = "in_progress"
    COMPLETED = "completed"
    FAILED = "failed"
    CANCELLED = "cancelled"


@dataclass
class BatchOperation:
    """Represents a single operation in a batch."""
    id: str
    operation_type: OperationType
    file_path: str
    parameters: Dict[str, Any] = field(default_factory=dict)
    status: OperationStatus = OperationStatus.PENDING
    result: Optional[Any] = None
    error_message: Optional[str] = None
    start_time: Optional[float] = None
    end_time: Optional[float] = None
    priority: int = 0  # Higher numbers = higher priority


@dataclass
class BatchResult:
    """Results of a batch operation."""
    total_operations: int
    completed: int
    failed: int
    cancelled: int
    total_time: float
    operations: List[BatchOperation]
    
    @property
    def success_rate(self) -> float:
        """Calculate success rate percentage."""
        if self.total_operations == 0:
            return 0.0
        return (self.completed / self.total_operations) * 100


class OperationQueue:
    """Thread-safe operation queue with prioritization."""
    
    def __init__(self):
        self._queue = queue.PriorityQueue()
        self._operations = {}
        self._lock = threading.Lock()
        self._counter = 0  # For FIFO ordering of same-priority items
    
    def add_operation(self, operation: BatchOperation):
        """Add operation to queue."""
        with self._lock:
            # Use negative priority for max-heap behavior
            priority_item = (-operation.priority, self._counter, operation)
            self._queue.put(priority_item)
            self._operations[operation.id] = operation
            self._counter += 1
            
        logger.info("Operation added to queue", 
                   operation_id=operation.id,
                   operation_type=operation.operation_type.value,
                   priority=operation.priority)
    
    def get_operation(self, timeout=None) -> Optional[BatchOperation]:
        """Get next operation from queue."""
        try:
            _, _, operation = self._queue.get(timeout=timeout)
            return operation
        except queue.Empty:
            return None
    
    def remove_operation(self, operation_id: str) -> bool:
        """Remove operation from tracking."""
        with self._lock:
            if operation_id in self._operations:
                del self._operations[operation_id]
                return True
            return False
    
    def get_pending_operations(self) -> List[BatchOperation]:
        """Get list of pending operations."""
        with self._lock:
            return [op for op in self._operations.values() 
                   if op.status == OperationStatus.PENDING]
    
    def size(self) -> int:
        """Get queue size."""
        return self._queue.qsize()
    
    def clear(self):
        """Clear all operations."""
        with self._lock:
            while not self._queue.empty():
                try:
                    self._queue.get_nowait()
                except queue.Empty:
                    break
            self._operations.clear()


class BatchProcessor:
    """Main batch processing engine."""
    
    def __init__(self, max_workers: int = 1):
        """
        Initialize batch processor.
        
        Args:
            max_workers: Maximum number of concurrent operations
        """
        self.max_workers = max_workers
        self.operation_queue = OperationQueue()
        self.is_running = False
        self.workers = []
        self.operation_handlers = {}
        self.progress_callback = None
        self.completion_callback = None
        self.error_callback = None
        
        # Statistics
        self.start_time = None
        self.total_operations = 0
        self.completed_operations = 0
        self.failed_operations = 0
        self.cancelled_operations = 0
        
        logger.info("Batch processor initialized", max_workers=max_workers)
    
    def register_operation_handler(self, operation_type: OperationType, 
                                 handler: Callable):
        """Register handler for specific operation type."""
        self.operation_handlers[operation_type] = handler
        logger.info("Operation handler registered", 
                   operation_type=operation_type.value)
    
    def add_operation(self, operation: BatchOperation):
        """Add operation to processing queue."""
        if operation.operation_type not in self.operation_handlers:
            raise ValueError(f"No handler registered for {operation.operation_type}")
        
        self.operation_queue.add_operation(operation)
        self.total_operations += 1
    
    def start_processing(self):
        """Start batch processing."""
        if self.is_running:
            logger.warning("Batch processor already running")
            return
        
        self.is_running = True
        self.start_time = time.time()
        
        with LogContext("batch_processing_start"):
            logger.info("Starting batch processing",
                       total_operations=self.total_operations,
                       max_workers=self.max_workers)
        
        # Start worker threads
        for i in range(self.max_workers):
            worker = threading.Thread(
                target=self._worker_thread,
                name=f"BatchWorker-{i+1}",
                daemon=True
            )
            worker.start()
            self.workers.append(worker)
    
    def stop_processing(self):
        """Stop batch processing."""
        self.is_running = False
        logger.info("Batch processing stopped")
    
    def _worker_thread(self):
        """Worker thread for processing operations."""
        worker_name = threading.current_thread().name
        logger.info("Batch worker started", worker_name=worker_name)
        
        while self.is_running:
            operation = self.operation_queue.get_operation(timeout=1.0)
            if operation is None:
                continue
            
            self._process_operation(operation)
            
            # Check if all operations completed
            if (self.completed_operations + self.failed_operations + 
                self.cancelled_operations >= self.total_operations):
                self._finish_processing()
                break
        
        logger.info("Batch worker finished", worker_name=worker_name)
    
    def _process_operation(self, operation: BatchOperation):
        """Process a single operation."""
        operation.status = OperationStatus.IN_PROGRESS
        operation.start_time = time.time()
        
        with LogContext("batch_operation", 
                       operation_id=operation.id,
                       operation_type=operation.operation_type.value):
            logger.info("Processing batch operation",
                       file_path=operation.file_path)
            
            try:
                handler = self.operation_handlers[operation.operation_type]
                operation.result = handler(operation)
                operation.status = OperationStatus.COMPLETED
                self.completed_operations += 1
                
                logger.info("Batch operation completed successfully")
                
            except Exception as e:
                operation.status = OperationStatus.FAILED
                operation.error_message = str(e)
                self.failed_operations += 1
                
                logger.exception("Batch operation failed",
                               error_message=str(e))
                
                if self.error_callback:
                    self.error_callback(operation, e)
            
            finally:
                operation.end_time = time.time()
                self.operation_queue.remove_operation(operation.id)
                
                # Update progress
                if self.progress_callback:
                    self.progress_callback(self._get_progress_info())
    
    def _finish_processing(self):
        """Finish batch processing and generate results."""
        self.is_running = False
        end_time = time.time()
        total_time = end_time - self.start_time if self.start_time else 0
        
        result = BatchResult(
            total_operations=self.total_operations,
            completed=self.completed_operations,
            failed=self.failed_operations,
            cancelled=self.cancelled_operations,
            total_time=total_time,
            operations=[]  # Could be populated if needed
        )
        
        with LogContext("batch_processing_complete"):
            logger.info("Batch processing completed",
                       total_operations=result.total_operations,
                       completed=result.completed,
                       failed=result.failed,
                       success_rate=result.success_rate,
                       total_time=total_time)
        
        if self.completion_callback:
            self.completion_callback(result)
    
    def _get_progress_info(self) -> Dict[str, Any]:
        """Get current progress information."""
        total = self.total_operations
        processed = self.completed_operations + self.failed_operations + self.cancelled_operations
        
        return {
            "total": total,
            "processed": processed,
            "completed": self.completed_operations,
            "failed": self.failed_operations,
            "cancelled": self.cancelled_operations,
            "percentage": (processed / total * 100) if total > 0 else 0,
            "pending": self.operation_queue.size()
        }


class DragDropListbox(tk.Listbox):
    """Listbox with drag-and-drop support for files."""
    
    def __init__(self, parent, **kwargs):
        super().__init__(parent, **kwargs)
        self.file_paths = []
        self.drop_callback = None
        
        # Configure drag-and-drop
        self.drop_target_register(dnd.DND_FILES)
        self.dnd_bind('<<Drop>>', self._on_drop)
        
        # Configure appearance for drag-and-drop
        self.configure(
            selectmode=tk.EXTENDED,
            height=8,
            bg='white',
            relief='sunken',
            bd=2
        )
        
        # Add helpful text
        self._show_help_text()
    
    def _show_help_text(self):
        """Show help text when listbox is empty."""
        if not self.file_paths:
            self.insert(tk.END, "Drag and drop Excel files here...")
            self.insert(tk.END, "Or use 'Add Files' button")
            self.configure(fg='gray')
    
    def _clear_help_text(self):
        """Clear help text."""
        if self.size() > 0 and not self.file_paths:
            self.delete(0, tk.END)
            self.configure(fg='black')
    
    def _on_drop(self, event):
        """Handle file drop event."""
        files = self.tk.splitlist(event.data)
        excel_files = [f for f in files if f.lower().endswith(('.xlsx', '.xls'))]
        
        if excel_files:
            self._clear_help_text()
            
            for file_path in excel_files:
                if file_path not in self.file_paths:
                    self.file_paths.append(file_path)
                    filename = os.path.basename(file_path)
                    self.insert(tk.END, filename)
            
            logger.info("Files dropped into batch list", 
                       file_count=len(excel_files))
            
            if self.drop_callback:
                self.drop_callback(excel_files)
        else:
            messagebox.showwarning(
                "Invalid Files", 
                "Please drop only Excel files (.xlsx or .xls)"
            )
    
    def add_files(self, file_paths: List[str]):
        """Add files programmatically."""
        self._clear_help_text()
        
        for file_path in file_paths:
            if file_path not in self.file_paths:
                self.file_paths.append(file_path)
                filename = os.path.basename(file_path)
                self.insert(tk.END, filename)
    
    def remove_selected_files(self):
        """Remove selected files from list."""
        selected_indices = list(self.curselection())
        selected_indices.reverse()  # Remove from end to avoid index shifting
        
        for index in selected_indices:
            if index < len(self.file_paths):
                del self.file_paths[index]
                self.delete(index)
        
        if not self.file_paths:
            self._show_help_text()
    
    def clear_files(self):
        """Clear all files."""
        self.file_paths.clear()
        self.delete(0, tk.END)
        self._show_help_text()
    
    def get_file_paths(self) -> List[str]:
        """Get list of file paths."""
        return self.file_paths.copy()


class BatchOperationsDialog(tk.Toplevel):
    """Main dialog for batch operations."""
    
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Batch Operations - ProjectBudgetinator")
        self.geometry("800x600")
        self.resizable(True, True)
        
        self.batch_processor = BatchProcessor(max_workers=2)
        self.result = None
        
        # Make dialog modal
        self.transient(parent)
        self.grab_set()
        
        self._setup_ui()
        self._setup_batch_processor()
        
        logger.info("Batch operations dialog initialized")
    
    def _setup_ui(self):
        """Setup the user interface."""
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = ttk.Label(
            main_frame,
            text="Batch Operations",
            font=("TkDefaultFont", 16, "bold")
        )
        title_label.pack(pady=(0, 20))
        
        # Create notebook for tabs
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # File selection tab
        self._create_file_selection_tab(notebook)
        
        # Operation configuration tab
        self._create_operation_tab(notebook)
        
        # Progress monitoring tab
        self._create_progress_tab(notebook)
        
        # Results tab
        self._create_results_tab(notebook)
        
        # Control buttons
        self._create_control_buttons(main_frame)
    
    def _create_file_selection_tab(self, notebook):
        """Create file selection tab."""
        file_frame = ttk.Frame(notebook, padding="10")
        notebook.add(file_frame, text="File Selection")
        
        # Instructions
        instructions = ttk.Label(
            file_frame,
            text="Select Excel files for batch processing:",
            font=("TkDefaultFont", 12, "bold")
        )
        instructions.pack(pady=(0, 10))
        
        # File list with drag-drop
        list_frame = ttk.Frame(file_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        self.file_listbox = DragDropListbox(list_frame)
        self.file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.file_listbox.configure(yscrollcommand=scrollbar.set)
        scrollbar.configure(command=self.file_listbox.yview)
        
        # File management buttons
        button_frame = ttk.Frame(file_frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(
            button_frame,
            text="Add Files...",
            command=self._add_files
        ).pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Button(
            button_frame,
            text="Remove Selected",
            command=self._remove_selected_files
        ).pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Button(
            button_frame,
            text="Clear All",
            command=self._clear_all_files
        ).pack(side=tk.LEFT, padx=(0, 5))
        
        # File count label
        self.file_count_label = ttk.Label(
            button_frame,
            text="Files: 0"
        )
        self.file_count_label.pack(side=tk.RIGHT)
    
    def _create_operation_tab(self, notebook):
        """Create operation configuration tab."""
        op_frame = ttk.Frame(notebook, padding="10")
        notebook.add(op_frame, text="Operations")
        
        # Operation type selection
        type_frame = ttk.LabelFrame(op_frame, text="Operation Type", padding="10")
        type_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.operation_var = tk.StringVar(value=OperationType.VALIDATE_FILES.value)
        
        operations = [
            (OperationType.VALIDATE_FILES.value, "Validate Excel Files"),
            (OperationType.ADD_PARTNERS.value, "Add Partners (Batch)"),
            (OperationType.ADD_WORKPACKAGES.value, "Add Work Packages (Batch)"),
            (OperationType.CONVERT_FORMAT.value, "Convert File Format"),
            (OperationType.BACKUP_FILES.value, "Create Backups"),
        ]
        
        for value, text in operations:
            ttk.Radiobutton(
                type_frame,
                text=text,
                variable=self.operation_var,
                value=value
            ).pack(anchor=tk.W, pady=2)
        
        # Operation parameters
        params_frame = ttk.LabelFrame(op_frame, text="Parameters", padding="10")
        params_frame.pack(fill=tk.BOTH, expand=True)
        
        # Parameters will be dynamically updated based on operation type
        self.params_frame_content = ttk.Frame(params_frame)
        self.params_frame_content.pack(fill=tk.BOTH, expand=True)
        
        # Bind operation type change
        self.operation_var.trace('w', self._on_operation_type_change)
        self._on_operation_type_change()  # Initialize
    
    def _create_progress_tab(self, notebook):
        """Create progress monitoring tab."""
        progress_frame = ttk.Frame(notebook, padding="10")
        notebook.add(progress_frame, text="Progress")
        
        # Progress information
        info_frame = ttk.Frame(progress_frame)
        info_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.progress_label = ttk.Label(
            info_frame,
            text="Ready to start batch processing",
            font=("TkDefaultFont", 12)
        )
        self.progress_label.pack()
        
        # Progress bar
        self.progress_bar = ttk.Progressbar(
            progress_frame,
            mode='determinate',
            length=400
        )
        self.progress_bar.pack(pady=(0, 10))
        
        # Statistics frame
        stats_frame = ttk.LabelFrame(progress_frame, text="Statistics", padding="10")
        stats_frame.pack(fill=tk.X, pady=(0, 10))
        
        stats_grid = ttk.Frame(stats_frame)
        stats_grid.pack()
        
        # Statistics labels
        ttk.Label(stats_grid, text="Total:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        self.total_label = ttk.Label(stats_grid, text="0")
        self.total_label.grid(row=0, column=1, sticky=tk.W, padx=(0, 20))
        
        ttk.Label(stats_grid, text="Completed:").grid(row=0, column=2, sticky=tk.W, padx=(0, 10))
        self.completed_label = ttk.Label(stats_grid, text="0")
        self.completed_label.grid(row=0, column=3, sticky=tk.W, padx=(0, 20))
        
        ttk.Label(stats_grid, text="Failed:").grid(row=1, column=0, sticky=tk.W, padx=(0, 10))
        self.failed_label = ttk.Label(stats_grid, text="0")
        self.failed_label.grid(row=1, column=1, sticky=tk.W, padx=(0, 20))
        
        ttk.Label(stats_grid, text="Pending:").grid(row=1, column=2, sticky=tk.W, padx=(0, 10))
        self.pending_label = ttk.Label(stats_grid, text="0")
        self.pending_label.grid(row=1, column=3, sticky=tk.W, padx=(0, 20))
        
        # Current operation
        current_frame = ttk.LabelFrame(progress_frame, text="Current Operation", padding="10")
        current_frame.pack(fill=tk.X)
        
        self.current_operation_label = ttk.Label(
            current_frame,
            text="No operation in progress",
            wraplength=400
        )
        self.current_operation_label.pack()
    
    def _create_results_tab(self, notebook):
        """Create results tab."""
        results_frame = ttk.Frame(notebook, padding="10")
        notebook.add(results_frame, text="Results")
        
        # Results text area
        text_frame = ttk.Frame(results_frame)
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        self.results_text = tk.Text(
            text_frame,
            wrap=tk.WORD,
            height=15,
            state=tk.DISABLED
        )
        self.results_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        results_scrollbar = ttk.Scrollbar(
            text_frame,
            orient=tk.VERTICAL,
            command=self.results_text.yview
        )
        results_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.results_text.configure(yscrollcommand=results_scrollbar.set)
        
        # Results summary
        summary_frame = ttk.Frame(results_frame)
        summary_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Button(
            summary_frame,
            text="Export Results...",
            command=self._export_results
        ).pack(side=tk.LEFT)
        
        ttk.Button(
            summary_frame,
            text="Clear Results",
            command=self._clear_results
        ).pack(side=tk.LEFT, padx=(10, 0))
    
    def _create_control_buttons(self, parent):
        """Create control buttons."""
        button_frame = ttk.Frame(parent)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        self.start_button = ttk.Button(
            button_frame,
            text="Start Batch Processing",
            command=self._start_batch_processing
        )
        self.start_button.pack(side=tk.LEFT)
        
        self.stop_button = ttk.Button(
            button_frame,
            text="Stop Processing",
            command=self._stop_batch_processing,
            state=tk.DISABLED
        )
        self.stop_button.pack(side=tk.LEFT, padx=(10, 0))
        
        ttk.Button(
            button_frame,
            text="Close",
            command=self._close_dialog
        ).pack(side=tk.RIGHT)
    
    def _setup_batch_processor(self):
        """Setup batch processor with callbacks."""
        self.batch_processor.progress_callback = self._on_progress_update
        self.batch_processor.completion_callback = self._on_batch_complete
        self.batch_processor.error_callback = self._on_operation_error
        
        # Register operation handlers
        self.batch_processor.register_operation_handler(
            OperationType.VALIDATE_FILES, self._validate_file_handler
        )
        # Additional handlers would be registered here
    
    def _validate_file_handler(self, operation: BatchOperation) -> Dict[str, Any]:
        """Handler for file validation operations."""
        file_path = operation.file_path
        
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            
            result = {
                "file_path": file_path,
                "valid": True,
                "worksheet_count": len(workbook.sheetnames),
                "worksheets": workbook.sheetnames,
                "file_size": os.path.getsize(file_path)
            }
            
            workbook.close()
            return result
            
        except Exception as e:
            return {
                "file_path": file_path,
                "valid": False,
                "error": str(e)
            }
    
    def _add_files(self):
        """Add files through file dialog."""
        files = filedialog.askopenfilenames(
            title="Select Excel Files for Batch Processing",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if files:
            self.file_listbox.add_files(files)
            self._update_file_count()
            logger.info("Files added through dialog", file_count=len(files))
    
    def _remove_selected_files(self):
        """Remove selected files."""
        self.file_listbox.remove_selected_files()
        self._update_file_count()
    
    def _clear_all_files(self):
        """Clear all files."""
        self.file_listbox.clear_files()
        self._update_file_count()
    
    def _update_file_count(self):
        """Update file count display."""
        count = len(self.file_listbox.get_file_paths())
        self.file_count_label.configure(text=f"Files: {count}")
    
    def _on_operation_type_change(self, *args):
        """Handle operation type change."""
        # Clear current parameters
        for widget in self.params_frame_content.winfo_children():
            widget.destroy()
        
        operation_type = self.operation_var.get()
        
        # Add operation-specific parameters
        if operation_type == OperationType.VALIDATE_FILES.value:
            ttk.Label(
                self.params_frame_content,
                text="No additional parameters required for file validation."
            ).pack()
        
        elif operation_type == OperationType.ADD_PARTNERS.value:
            # Partner-specific parameters
            ttk.Label(
                self.params_frame_content,
                text="Partner data will be read from a template file."
            ).pack()
            
            ttk.Button(
                self.params_frame_content,
                text="Select Partner Template...",
                command=self._select_partner_template
            ).pack(pady=(10, 0))
        
        # Additional operation types would be handled here
    
    def _select_partner_template(self):
        """Select partner template file."""
        template_file = filedialog.askopenfilename(
            title="Select Partner Template File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if template_file:
            # Store template for later use
            self.partner_template = template_file
            logger.info("Partner template selected", template_file=template_file)
    
    def _start_batch_processing(self):
        """Start batch processing."""
        file_paths = self.file_listbox.get_file_paths()
        
        if not file_paths:
            messagebox.showwarning(
                "No Files", 
                "Please add files to process before starting."
            )
            return
        
        operation_type = OperationType(self.operation_var.get())
        
        # Create operations for each file
        for i, file_path in enumerate(file_paths):
            operation = BatchOperation(
                id=f"op_{i}_{int(time.time())}",
                operation_type=operation_type,
                file_path=file_path,
                parameters={},  # Would include operation-specific parameters
                priority=0
            )
            self.batch_processor.add_operation(operation)
        
        # Update UI state
        self.start_button.configure(state=tk.DISABLED)
        self.stop_button.configure(state=tk.NORMAL)
        
        # Start processing
        self.batch_processor.start_processing()
        
        logger.info("Batch processing started",
                   file_count=len(file_paths),
                   operation_type=operation_type.value)
    
    def _stop_batch_processing(self):
        """Stop batch processing."""
        self.batch_processor.stop_processing()
        self._update_ui_state(False)
        
        logger.info("Batch processing stopped by user")
    
    def _on_progress_update(self, progress_info: Dict[str, Any]):
        """Handle progress updates."""
        self.after(0, lambda: self._update_progress_ui(progress_info))
    
    def _update_progress_ui(self, progress_info: Dict[str, Any]):
        """Update progress UI elements."""
        total = progress_info["total"]
        processed = progress_info["processed"]
        percentage = progress_info["percentage"]
        
        # Update progress bar
        self.progress_bar.configure(value=percentage, maximum=100)
        
        # Update labels
        self.progress_label.configure(
            text=f"Processing: {processed}/{total} files ({percentage:.1f}%)"
        )
        
        self.total_label.configure(text=str(total))
        self.completed_label.configure(text=str(progress_info["completed"]))
        self.failed_label.configure(text=str(progress_info["failed"]))
        self.pending_label.configure(text=str(progress_info["pending"]))
    
    def _on_batch_complete(self, result: BatchResult):
        """Handle batch completion."""
        self.after(0, lambda: self._handle_batch_complete(result))
    
    def _handle_batch_complete(self, result: BatchResult):
        """Handle batch completion in main thread."""
        self.result = result
        self._update_ui_state(False)
        
        # Show completion message
        message = (
            f"Batch processing completed!\n\n"
            f"Total files: {result.total_operations}\n"
            f"Completed: {result.completed}\n"
            f"Failed: {result.failed}\n"
            f"Success rate: {result.success_rate:.1f}%\n"
            f"Total time: {result.total_time:.1f} seconds"
        )
        
        messagebox.showinfo("Batch Complete", message)
        
        # Add results to results tab
        self._add_result_text(f"\\n=== Batch Processing Complete ===\\n")
        self._add_result_text(message)
        
        logger.info("Batch processing completed notification shown")
    
    def _on_operation_error(self, operation: BatchOperation, error: Exception):
        """Handle operation errors."""
        error_msg = f"Error in {operation.file_path}: {str(error)}"
        self.after(0, lambda: self._add_result_text(f"ERROR: {error_msg}\\n"))
    
    def _update_ui_state(self, processing: bool):
        """Update UI state based on processing status."""
        if processing:
            self.start_button.configure(state=tk.DISABLED)
            self.stop_button.configure(state=tk.NORMAL)
        else:
            self.start_button.configure(state=tk.NORMAL)
            self.stop_button.configure(state=tk.DISABLED)
    
    def _add_result_text(self, text: str):
        """Add text to results area."""
        self.results_text.configure(state=tk.NORMAL)
        self.results_text.insert(tk.END, text)
        self.results_text.see(tk.END)
        self.results_text.configure(state=tk.DISABLED)
    
    def _export_results(self):
        """Export results to file."""
        if not hasattr(self, 'result') or not self.result:
            messagebox.showwarning("No Results", "No results to export.")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="Export Results",
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                with open(file_path, 'w') as f:
                    f.write(f"Batch Processing Results\\n")
                    f.write(f"{'='*50}\\n")
                    f.write(f"Total operations: {self.result.total_operations}\\n")
                    f.write(f"Completed: {self.result.completed}\\n")
                    f.write(f"Failed: {self.result.failed}\\n")
                    f.write(f"Success rate: {self.result.success_rate:.1f}%\\n")
                    f.write(f"Total time: {self.result.total_time:.1f} seconds\\n")
                
                messagebox.showinfo("Export Complete", f"Results exported to {file_path}")
                logger.info("Results exported", file_path=file_path)
                
            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to export results: {str(e)}")
                logger.exception("Failed to export results")
    
    def _clear_results(self):
        """Clear results area."""
        self.results_text.configure(state=tk.NORMAL)
        self.results_text.delete(1.0, tk.END)
        self.results_text.configure(state=tk.DISABLED)
    
    def _close_dialog(self):
        """Close the dialog."""
        if self.batch_processor.is_running:
            if messagebox.askyesno(
                "Processing Active", 
                "Batch processing is still running. Do you want to stop it and close?"
            ):
                self.batch_processor.stop_processing()
                self.destroy()
        else:
            self.destroy()


# Convenience functions for easy integration

def show_batch_operations_dialog(parent):
    """Show batch operations dialog."""
    dialog = BatchOperationsDialog(parent)
    parent.wait_window(dialog)
    return dialog.result


def create_file_validation_batch(file_paths: List[str]) -> List[BatchOperation]:
    """Create batch operations for file validation."""
    operations = []
    for i, file_path in enumerate(file_paths):
        operation = BatchOperation(
            id=f"validate_{i}_{int(time.time())}",
            operation_type=OperationType.VALIDATE_FILES,
            file_path=file_path
        )
        operations.append(operation)
    return operations
